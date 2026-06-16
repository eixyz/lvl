"""
lvl_refraction.py  --  First-break picking + LVL refraction analysis
====================================================================
Project : Low Velocity Layer (LVL) refraction seismic surveys

Workflow (mirrors ProMAX flow)
------------------------------
1.  Read SEG2 files -- one file per shot, sorted by FFID (file number)
2.  Assign receiver geometry from standard geometry files
    (geometry100.txt = 100 m spread,  geometry200.txt = 200 m spread)
3.  Apply bulk time static  (like ProMAX: bulk shift static -> add)
4.  Apply zero-phase Ormsby bandpass filter
    (frequency domain, 25 % zero-padding for FFT, 4 corner frequencies)
5.  Interactive first-break picking  (matplotlib GUI)
6.  Automatic refraction analysis:
    -- AIC-based breakpoint detection (or manual)
    -- T-X segment fitting -> velocities
    -- Perpendicular offset correction  (true offset = sqrt(inline^2 + perp^2))
    -- Intercept-time depth inversion (2- or 3-layer)
7.  Export:
    -- <profile>_picks.xlsx     Config + per-shot formulas + Analysis sheet
    -- <profile>_picks_clean.txt  compatible with lvl.ipynb
    -- <profile>_tx_picks.png     T-X summary with fitted lines, time downward
    -- <profile>_shot_XX_picks.png  per-shot QC image (auto-saved on 's')

Interactive picker controls
---------------------------
    Left-click / drag       place picks quickly
    Right-click / drag      delete picks quickly
    Shift + Left-click      range-fill picks between two traces
    s or n                  save immediately and go to next shot
    p                       save and go to previous shot
    a                       STA/LTA auto-pick (gain-aware)
    f                       toggle Ormsby filter on / off
    v                       invert trace polarity (display only)
    g                       cycle gain mode: norm/agc/none
    l                       toggle timeline guides
    c                       toggle top axis: channel <-> signed offset
    k                       save per-shot QC screenshot (PNG)
    q / Esc                 quit picking

Directory layout
----------------
  data/
    geometry100.txt
    geometry200.txt
    120/       Rec_00001.seg2  Rec_00002.seg2  Rec_00003.seg2
    150/  ...
  scripts/
    lvl_refraction.py   <- this file
  output/
    120/       <auto-created>

How to run
----------
  conda activate seiseng
  cd d:\\Daten\\seismic\\lvl\\scripts

  python lvl_refraction.py 120              # interactive picking for profile 120
  python lvl_refraction.py 120 --export-only  # re-export without picking
  python lvl_refraction.py --all            # pick all profiles in sequence
  python lvl_refraction.py                  # list available profiles
"""

from __future__ import annotations

import sys
import argparse
import json
import math
import importlib
import subprocess
import time
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Auto-install missing packages into the active environment
# ---------------------------------------------------------------------------
def _ensure(pkg: str, mod: str = "") -> Any:
    m = mod or pkg
    try:
        return importlib.import_module(m)
    except ImportError:
        print(f"  Installing '{pkg}' ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        return importlib.import_module(m)

np = _ensure("numpy")
pd = _ensure("pandas")
_ensure("obspy")
_ensure("matplotlib")
_ensure("openpyxl")

from obspy import read as _read_obspy          # type: ignore[import-untyped]
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.widgets import Button, Slider
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ===========================================================================
# CONFIG  --  edit this section for each project
# ===========================================================================

CWD        = Path(__file__).parent
DATA_DIR   = CWD.parent / "data"
OUTPUT_DIR = CWD.parent / "output"

# Geometry files: map trace number -> absolute receiver position (m along profile)
GEOM_FILES: dict = {
    100: DATA_DIR / "geometry100.txt",
    200: DATA_DIR / "geometry200.txt",
}

# ---------------------------------------------------------------------------
# Profile registry
# ---------------------------------------------------------------------------
# geom    : geometry type (100 or 200)
# line_no : line identifier used for labelling
# perp_m  : perpendicular distance of shot from the geophone line (m)
# shots   : {shot_id: inline_position_m}  OR  "auto"
#           "auto" = derive from geometry:
#             shot 1 = recv[0] - first_spacing     (before spread)
#             shot 2 = (recv[n//2-1] + recv[n//2]) / 2   (midpoint)
#             shot 3 = recv[-1] + last_spacing     (after spread)
#           Manual values are used when SEG2 SOURCE_LOCATION is absent.
#           SEG2 header always takes priority if present.
# ---------------------------------------------------------------------------
PROFILES: dict = {
    "120":   {"geom": 100, "line_no": 1, "perp_m": 0.0, "shots": "auto"},
    "214_A": {"geom": 100, "line_no": 2, "perp_m": 0.0, "shots": "auto"},
    "150":   {"geom": 200, "line_no": 3, "perp_m": 0.0, "shots": "auto"},
    "415_B": {"geom": 200, "line_no": 4, "perp_m": 0.0, "shots": "auto"},
}

# ---------------------------------------------------------------------------
# Ormsby bandpass  (zero-phase, frequency domain)
# ProMAX convention:  f1 - f2 - f3 - f4
#   f1 : low-cut ramp start      (Hz)
#   f2 : low-cut ramp end / passband start   (Hz)
#   f3 : passband end / high-cut ramp start  (Hz)
#   f4 : high-cut ramp end       (Hz)
# ---------------------------------------------------------------------------
BP_F1: float = 2.0    # Hz
BP_F2: float = 4.0    # Hz
BP_F3: float = 140.0  # Hz
BP_F4: float = 180.0  # Hz
BP_FFT_PAD: float = 0.25   # 25 % zero-padding (ProMAX default)
BP_REAPPLY: bool  = False  # True = apply twice (squares amplitude response)

# ---------------------------------------------------------------------------
# Bulk time static  (ProMAX: "Bulk shift static -> Add")
# The SEG2 DELAY field (e.g. -106 ms) is read from each file and used to
# build the correct absolute time axis.  Picks are stored as absolute times
# from the trigger (t = 0).  BULK_SHIFT_MS is an ADDITIONAL correction on
# top of the DELAY  (leave at 0.0 unless a systematic offset remains).
# ---------------------------------------------------------------------------
BULK_SHIFT_MS: float = 0.0   # ms  (negative = shift to earlier times)

# ---------------------------------------------------------------------------
# Display / theme
# ---------------------------------------------------------------------------
THEME: str         = "light"   # "dark"  or  "light"
T_MAX_MS: float    = 150.0    # end of display window (ms from trigger)
T_DISPLAY_PRE_MS: float = -10.0  # ms before trigger to show (headroom above first arrivals)
CLIP_FACTOR: float = 2.0      # wiggle clip level in std-dev multiples
FILTER_ON: bool    = True     # start with filter active
DISPLAY_MODE: str  = "both"   # "wiggle" | "vd" | "both"
SHOW_TIMELINES: bool = True
WIGGLE_STRETCH: float = 0.85

# ---------------------------------------------------------------------------
# Gain control for display/picking
# ---------------------------------------------------------------------------
GAIN_MODE: str        = "norm"   # "none" | "norm" | "agc"
AGC_WINDOW_MS: float  = 200.0
AGC_STAT: str         = "rms"    # "mean" | "rms"

# ---------------------------------------------------------------------------
# STA/LTA auto-picker
# ---------------------------------------------------------------------------
STA_MS: float      = 3.0
LTA_MS: float      = 20.0
STALTA_TRIG: float = 3.0

# ---------------------------------------------------------------------------
# Refraction analysis
# ---------------------------------------------------------------------------
N_LAYERS: int          = 2      # 2 or 3
AUTO_BREAKPOINTS: bool = True   # True = AIC search;  False = MANUAL_BREAKS
# MANUAL_BREAKS: {shot_id: [receiver_index_of_crossover, ...]}
# Example:  {1: [12], 2: [14], 3: [12]}
MANUAL_BREAKS: dict = {}

# ===========================================================================
# END OF CONFIG
# ===========================================================================


# ---------------------------------------------------------------------------
# Theme colour palette
# ---------------------------------------------------------------------------
def _tc() -> dict:
    """Return a colour dict for the current THEME."""
    if THEME == "dark":
        return dict(
            fig_bg="#0f1117",   ax_bg="#1a1d2e",
            trace="#4a8fc1",    fill_alpha=0.20,
            pick_clr="#e63946", text="white",
            label="#aaa",       grid="#888",
            tick="#888",        spine="#333344",
            leg_face="#1e2130", leg_edge="#555",
            fit_alpha=0.90,
        )
    return dict(
        fig_bg="white",     ax_bg="#f5f5f5",
        trace="#111111",    fill_alpha=0.18,
        pick_clr="#cc0000", text="#111111",
        label="#333",       grid="#bbb",
        tick="#333",        spine="#aaaaaa",
        leg_face="#f0f0f0", leg_edge="#aaa",
        fit_alpha=0.85,
    )


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def load_geometry(geom_type: int) -> Any:
    """Return receiver positions (m) as a 1-D ndarray."""
    path = GEOM_FILES[geom_type]
    positions: list = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            s = line.strip()
            if not s or s.lower().startswith("trace"):
                continue
            # Handle German decimal comma:  "2,5" -> 2.5
            parts = s.replace(",", ".").split()
            if len(parts) >= 2:
                try:
                    positions.append(float(parts[1]))
                except ValueError:
                    pass
    return np.array(positions, dtype=float)


def auto_shot_positions(recv_pos: Any) -> dict:
    """
    Compute standard refraction shot positions from the receiver geometry.

    Convention used by most land refraction surveys:
      Shot 1  :  one receiver-spacing BEFORE the first receiver
      Shot 2  :  midpoint between receivers n//2 and n//2+1
                 (labelled e.g. GP24.5 for a 48-channel spread)
      Shot 3  :  one receiver-spacing AFTER the last receiver

    Returns {1: pos_m, 2: pos_m, 3: pos_m}.
    """
    n          = len(recv_pos)
    dx_start   = float(recv_pos[1]    - recv_pos[0])    # first spacing
    dx_end     = float(recv_pos[-1]   - recv_pos[-2])   # last spacing
    mid_lo     = float(recv_pos[n // 2 - 1])
    mid_hi     = float(recv_pos[n // 2])
    return {
        1: round(float(recv_pos[0]) - dx_start, 4),
        2: round((mid_lo + mid_hi) / 2.0,        4),
        3: round(float(recv_pos[-1]) + dx_end,   4),
    }


def true_offset(inline_m: Any, perp_m: float) -> Any:
    """True source-receiver distance corrected for perpendicular shot offset."""
    return np.sqrt(np.asarray(inline_m, dtype=float) ** 2 + perp_m ** 2)


# ---------------------------------------------------------------------------
# SEG2 reader
# ---------------------------------------------------------------------------

def read_seg2(path: Path) -> tuple:
    """
    Read one SEG2 shot record.

    Returns
    -------
    data          : ndarray (n_traces x n_samples, float32)
    dt_s          : sample interval (s)
    n_traces      : int
    n_samples     : int
    shot_pos      : float or None  (SOURCE_LOCATION header, first component)
    ffid          : int  (SHOT_SEQUENCE_NUMBER or filename digits)
    delay_ms      : float  (DELAY header in ms; typically negative pre-trigger)
    recv_locs_m   : ndarray or None  (RECEIVER_LOCATION per trace, m)
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        st = _read_obspy(str(path), format="SEG2")

    n_traces = len(st)
    n_samp   = max(tr.stats.npts for tr in st)
    dt_s     = float(st[0].stats.delta)

    data = np.zeros((n_traces, n_samp), dtype=np.float32)
    for i, tr in enumerate(st):
        npts = tr.stats.npts
        data[i, :npts] = tr.data.astype(np.float32)

    shot_pos: Any  = None
    ffid: int      = 0
    delay_ms: float = 0.0
    recv_locs_m: Any = None

    try:
        hdr0 = dict(st[0].stats.seg2)

        sloc = str(hdr0.get("SOURCE_LOCATION", "")).strip()
        if sloc:
            # Guard against German comma decimal
            shot_pos = float(sloc.replace(",", ".").split()[0])

        ssn = str(hdr0.get("SHOT_SEQUENCE_NUMBER", "")).strip()
        if ssn.isdigit():
            ffid = int(ssn)

        # DELAY is stored in seconds; convert to ms
        delay_str = str(hdr0.get("DELAY", "0")).strip()
        delay_ms  = float(delay_str.replace(",", ".")) * 1000.0

        # RECEIVER_LOCATION per trace (m along profile)
        locs = []
        for tr in st:
            h = dict(tr.stats.seg2)
            rl = str(h.get("RECEIVER_LOCATION", "")).strip().replace(",", ".")
            locs.append(float(rl) if rl else None)
        if all(v is not None for v in locs):
            recv_locs_m = np.array(locs, dtype=float)
    except Exception:
        pass

    if ffid == 0:
        digits = "".join(c for c in path.stem if c.isdigit())
        ffid   = int(digits) if digits else 0

    return data, dt_s, n_traces, n_samp, shot_pos, ffid, delay_ms, recv_locs_m


# ---------------------------------------------------------------------------
# Ormsby bandpass  (zero-phase, frequency domain)
# ---------------------------------------------------------------------------

def _ormsby_response(freqs: Any,
                     f1: float, f2: float, f3: float, f4: float) -> Any:
    """Trapezoidal Ormsby amplitude response (linear ramps)."""
    H = np.zeros(len(freqs), dtype=float)
    m_lo = (freqs > f1) & (freqs <= f2)
    m_pb = (freqs > f2) & (freqs <= f3)
    m_hi = (freqs > f3) & (freqs < f4)
    H[m_lo] = (freqs[m_lo] - f1) / (f2 - f1)
    H[m_pb] = 1.0
    H[m_hi] = (f4 - freqs[m_hi]) / (f4 - f3)
    return H


def ormsby(trace: Any, dt_s: float,
           f1: float = BP_F1, f2: float = BP_F2,
           f3: float = BP_F3, f4: float = BP_F4,
           pad_pct: float = BP_FFT_PAD,
           reapply: bool = BP_REAPPLY) -> Any:
    """
    Zero-phase Ormsby bandpass filter  (ProMAX-compatible).

    Algorithm
    ---------
    1. Zero-pad to next power-of-2  (pad_pct=0.25 -> 25 % extra = ProMAX default)
    2. FFT -> multiply by trapezoidal amplitude response (zero-phase: real filter)
    3. IFFT -> trim to original length
    4. Optionally repeat once  (reapply=True squares the amplitude response)
    """
    n     = len(trace)
    n_pad = int(n * (1.0 + pad_pct))
    n_fft = 1 << (n_pad - 1).bit_length()          # next power of 2
    spec  = np.fft.rfft(trace.astype(np.float64), n=n_fft)
    freqs = np.fft.rfftfreq(n_fft, d=dt_s)
    H     = _ormsby_response(freqs, f1, f2, f3, f4)
    if reapply:
        H = H ** 2
    filtered = np.fft.irfft(spec * H, n=n_fft)
    return filtered[:n].astype(np.float32)


def apply_ormsby_all(data: Any, dt_s: float) -> Any:
    """Apply Ormsby bandpass to every trace in data (n_traces x n_samples)."""
    out = np.empty_like(data)
    for i in range(data.shape[0]):
        out[i] = ormsby(data[i], dt_s)
    return out


def apply_ormsby_all_params(data: Any, dt_s: float,
                            f1: float, f2: float, f3: float, f4: float) -> Any:
    """Apply Ormsby with explicit frequency parameters to all traces."""
    out = np.empty_like(data)
    for i in range(data.shape[0]):
        out[i] = ormsby(data[i], dt_s, f1=f1, f2=f2, f3=f3, f4=f4)
    return out


# ---------------------------------------------------------------------------
# Bulk static correction
# ---------------------------------------------------------------------------

def apply_bulk_static(picks_raw: dict) -> dict:
    """
    Apply BULK_SHIFT_MS to a picks dict {trace_idx: time_ms}.
    Call at analysis / export time; JSON always stores RAW picks.
    """
    if abs(BULK_SHIFT_MS) < 1e-9:
        return picks_raw
    return {k: max(0.0, float(v) + BULK_SHIFT_MS) for k, v in picks_raw.items()}


# ---------------------------------------------------------------------------
# STA/LTA auto-picker
# ---------------------------------------------------------------------------

def stalta_pick(trace: Any, dt_s: float,
                sta_ms: float = STA_MS,
                lta_ms: float = LTA_MS,
                threshold: float = STALTA_TRIG) -> Any:
    """Return first-break time (ms) via STA/LTA energy ratio, or None."""
    n_sta = max(1, int(sta_ms  / 1000.0 / dt_s))
    n_lta = max(2, int(lta_ms  / 1000.0 / dt_s))
    n_max = min(len(trace), int(T_MAX_MS / 1000.0 / dt_s) + n_sta)
    e     = trace[:n_max].astype(float) ** 2
    for k in range(n_lta, len(e) - n_sta):
        lta_v = e[k - n_lta:k].mean()
        if lta_v < 1e-30:
            continue
        if e[k:k + n_sta].mean() / lta_v >= threshold:
            return round(float(k) * dt_s * 1000.0, 2)
    return None


def apply_gain(data: Any, dt_s: float,
               mode: str = GAIN_MODE,
               window_ms: float = AGC_WINDOW_MS,
               stat: str = AGC_STAT) -> Any:
    mode_l = str(mode).lower()
    out = np.asarray(data, dtype=np.float32).copy()

    if mode_l == "none":
        return out

    if mode_l == "norm":
        for i in range(out.shape[0]):
            mx = float(np.max(np.abs(out[i])))
            if mx > 1e-20:
                out[i] /= mx
        return out

    if mode_l == "agc":
        win = max(3, int((window_ms / 1000.0) / dt_s))
        if win % 2 == 0:
            win += 1
        ker = np.ones(win, dtype=np.float64) / float(win)
        for i in range(out.shape[0]):
            tr = out[i].astype(np.float64)
            if str(stat).lower() == "mean":
                env = np.convolve(np.abs(tr), ker, mode="same")
            else:
                env = np.sqrt(np.convolve(tr * tr, ker, mode="same"))
            env = np.where(env > 1e-12, env, 1e-12)
            out[i] = (tr / env).astype(np.float32)
        return out

    return out


# ---------------------------------------------------------------------------
# Refraction analysis functions  (integrated from lvl.ipynb)
# ---------------------------------------------------------------------------

def fit_line(x: Any, y: Any) -> tuple:
    """
    Fit y = slope*x + intercept.
    Returns (slope [ms/m], intercept [ms], r^2).
    Velocity (m/s) = 1000 / slope  when x in m, y in ms.
    """
    if len(x) < 2:
        return float("nan"), float("nan"), float("nan")
    x_a = np.asarray(x, dtype=float)
    y_a = np.asarray(y, dtype=float)
    slope, intercept = np.polyfit(x_a, y_a, 1)
    y_hat  = slope * x_a + intercept
    ss_res = float(np.sum((y_a - y_hat) ** 2))
    ss_tot = float(np.sum((y_a - y_a.mean()) ** 2))
    r2     = 1.0 - ss_res / ss_tot if ss_tot > 1e-30 else 1.0
    return float(slope), float(intercept), float(r2)


def depth_2layer(ti_ms: float, V1: float, V2: float) -> Any:
    """
    Intercept-time depth, 2-layer model.
    h = (ti_s x V1 x V2) / (2 x sqrt(V2^2 - V1^2))  [m]
    Returns None if V2 <= V1.
    """
    if V2 <= V1 or V1 <= 0.0:
        return None
    return (ti_ms / 1000.0) * V1 * V2 / (2.0 * math.sqrt(V2 ** 2 - V1 ** 2))


def depth_3layer(ti2_ms: float, V1: float, V2: float,
                 V3: float, h1: float) -> Any:
    """
    Intercept-time depth to second refractor, 3-layer model.
    Accounts for first-layer delay time.  Returns None if geometry is invalid.
    """
    if V3 <= V2 or V2 <= V1 or V1 <= 0.0:
        return None
    cos_ic12 = math.sqrt(max(0.0, 1.0 - (V1 / V2) ** 2))
    delay_ms = 2.0 * h1 * cos_ic12 / V1 * 1000.0
    ti2_eff  = ti2_ms - delay_ms
    if ti2_eff <= 0.0:
        return None
    return (ti2_eff / 1000.0) * V1 * V3 / (2.0 * math.sqrt(V3 ** 2 - V1 ** 2))


def _find_break_aic(abs_off: Any, t_ms: Any) -> list:
    """Find up to (N_LAYERS-1) T-X crossover indices using AIC on segment fit residuals."""
    breaks: list = []
    seg_start = 0
    for _ in range(N_LAYERS - 1):
        seg_x = abs_off[seg_start:]
        seg_t = t_ms[seg_start:]
        seg_n = len(seg_x)
        if seg_n < 6:
            break
        best_k, best_aic = seg_n // 2, float("inf")
        for k in range(3, seg_n - 3):
            s1, b1, _ = fit_line(seg_x[:k], seg_t[:k])
            s2, b2, _ = fit_line(seg_x[k:], seg_t[k:])
            res1 = np.asarray(seg_t[:k]) - (s1 * np.asarray(seg_x[:k]) + b1)
            res2 = np.asarray(seg_t[k:]) - (s2 * np.asarray(seg_x[k:]) + b2)
            rss  = float(np.sum(res1 ** 2) + np.sum(res2 ** 2))
            aic  = seg_n * math.log(rss / seg_n + 1e-30) + 8.0
            if aic < best_aic:
                best_aic, best_k = aic, k
        breaks.append(seg_start + best_k)
        seg_start += best_k
    return breaks


def analyse_shots(shot_results: list, perp_m: float = 0.0) -> dict:
    """
    Segmented T-X fitting + intercept-time depth inversion for each shot.

    Parameters
    ----------
    shot_results : list of dicts, keys: shot_id, shot_pos, recv_pos, picks_ms
    perp_m       : perpendicular shot-to-line distance (m)

    Returns  {shot_id: result_dict or None}
    """
    results: dict = {}
    for sr in shot_results:
        sid       = sr["shot_id"]
        picks     = apply_bulk_static(sr["picks_ms"])   # apply static for analysis

        if len(picks) < 4:
            results[sid] = None
            continue

        recv_pos = sr["recv_pos"]
        shot_pos = sr["shot_pos"]
        idxs     = sorted(picks)
        inline   = np.array([recv_pos[i] - shot_pos for i in idxs])
        t_ms_arr = np.array([picks[i]               for i in idxs])
        abs_off  = np.abs(true_offset(inline, perp_m))

        if N_LAYERS > 1:
            break_idxs = (_find_break_aic(abs_off, t_ms_arr) if AUTO_BREAKPOINTS
                          else list(MANUAL_BREAKS.get(sid, [len(abs_off) // 2])))
        else:
            break_idxs = []

        boundaries = [0] + break_idxs + [len(abs_off)]
        segments: list = []
        for bi in range(len(boundaries) - 1):
            xs = abs_off[boundaries[bi]:boundaries[bi + 1]]
            ts = t_ms_arr[boundaries[bi]:boundaries[bi + 1]]
            if len(xs) < 2:
                continue
            slope, intercept, r2 = fit_line(xs, ts)
            vel = 1000.0 / slope if slope > 1e-9 else float("inf")
            segments.append({
                "x_start":      float(xs[0]),
                "x_end":        float(xs[-1]),
                "slope_ms_m":   slope,
                "intercept_ms": intercept,
                "velocity_m_s": vel,
                "r2":           r2,
                "n_picks":      len(xs),
            })

        depths: dict = {}
        if len(segments) >= 2:
            V1, V2 = segments[0]["velocity_m_s"], segments[1]["velocity_m_s"]
            ti1    = segments[1]["intercept_ms"]
            h1     = depth_2layer(ti1, V1, V2)
            depths = {"V1": V1, "V2": V2, "ti1_ms": ti1, "h1_m": h1}
        if len(segments) >= 3 and depths.get("h1_m"):
            V3  = segments[2]["velocity_m_s"]
            ti2 = segments[2]["intercept_ms"]
            h2  = depth_3layer(ti2, depths["V1"], depths["V2"], V3, depths["h1_m"])
            depths.update({"V3": V3, "ti2_ms": ti2, "h2_m": h2})

        t_pred = np.zeros(len(abs_off))
        for bi in range(len(boundaries) - 1):
            sl = segments[bi]["slope_ms_m"]   if bi < len(segments) else 0.0
            ic = segments[bi]["intercept_ms"] if bi < len(segments) else 0.0
            t_pred[boundaries[bi]:boundaries[bi + 1]] = (
                sl * abs_off[boundaries[bi]:boundaries[bi + 1]] + ic)
        rms_ms = float(np.sqrt(np.mean((t_ms_arr - t_pred) ** 2)))

        results[sid] = {
            "shot_id":    sid,
            "shot_pos":   shot_pos,
            "abs_off":    abs_off,
            "times_ms":   t_ms_arr,
            "segments":   segments,
            "break_idxs": break_idxs,
            "depths":     depths,
            "rms_ms":     rms_ms,
        }
    return results


# ---------------------------------------------------------------------------
# Interactive first-break picker
# ---------------------------------------------------------------------------

class FirstBreakPicker:
    """
    Matplotlib interactive first-break picker for one shot record.

    Key design decisions
    --------------------
    - run() uses plt.show(block=False) + a manual polling loop.
      This avoids the double-open artefact caused by plt.show(block=True)
      on Windows (figure re-appearing after plt.close was called by a key handler).
    - Pressing 's' first saves a per-shot QC PNG, then closes the window.
      The run() loop detects the closure and returns the picks.
    - The JSON always stores RAW picks; BULK_SHIFT_MS is applied later.
    """

    def __init__(self, data_raw: Any, data_filt: Any, dt_s: float,
                 recv_abs: Any, shot_id: int, profile_name: str,
                 shot_pos_m: float = 0.0,
                 delay_ms: float = 0.0,
                 existing_picks: dict | None = None,
                 qc_dir: Path | None = None,
                 save_callback: Any = None,
                 header_info: dict | None = None,
                 show_controls: bool = True,
                 control_file: Path | None = None):
        self.data_raw   = data_raw
        self.data_filt  = data_filt
        self.dt_s       = dt_s
        self.recv_abs   = recv_abs      # absolute receiver positions (m)
        self.shot_pos_m = shot_pos_m
        self.delay_ms   = delay_ms      # time of first sample relative to trigger (ms)
        self.shot_id    = shot_id
        self.profile    = profile_name
        self.qc_dir     = qc_dir or (OUTPUT_DIR / profile_name)

        self._picks: dict = {}
        self._saved          = False
        self._cancelled      = False
        self._done           = False
        self._filter_on      = FILTER_ON
        self._inverted       = False
        self._top_mode       = "channel"   # "channel" or "offset" (toggle with 'c')
        self.ax_top: Any     = None
        self._save_callback  = save_callback   # called immediately on 's' with picks dict
        self._header_info    = header_info or {}
        self._nav_action     = "stay"   # "next" | "prev" | "quit" | "stay"
        self._drag_pick      = False
        self._drag_delete    = False
        self._last_drag_idx: Any = None
        self._range_anchor: Any = None
        self._gain_mode      = GAIN_MODE
        self._agc_window_ms  = AGC_WINDOW_MS
        self._agc_stat       = AGC_STAT
        self._display_mode   = DISPLAY_MODE
        self._show_timelines = SHOW_TIMELINES
        self._wiggle_stretch = WIGGLE_STRETCH
        self._f1 = BP_F1
        self._f2 = BP_F2
        self._f3 = BP_F3
        self._f4 = BP_F4
        self._show_controls = bool(show_controls)
        self._control_file = Path(control_file) if control_file else None
        self._last_control_id: int = -1

        if existing_picks:
            self._picks = {int(k): float(v) for k, v in existing_picks.items()}

        self.n_traces = data_raw.shape[0]
        self.n_samp   = data_raw.shape[1]
        dt_ms         = dt_s * 1000.0

        # Absolute time axis:  first sample at delay_ms, increases by dt_ms
        self.times_ms = delay_ms + np.arange(self.n_samp) * dt_ms

        # Display window: T_DISPLAY_PRE_MS (< 0) to T_MAX_MS
        self.t_disp_start = max(delay_ms, T_DISPLAY_PRE_MS)
        self.t_disp_end   = T_MAX_MS
        t_start_idx = max(0, int((self.t_disp_start - delay_ms) / dt_ms))
        t_end_idx   = min(int((self.t_disp_end   - delay_ms) / dt_ms) + 2,
                          self.n_samp)
        self.t_slice = slice(t_start_idx, t_end_idx)

        # Trace spacing for wiggle display
        off_range = float(np.ptp(recv_abs)) if len(recv_abs) > 1 else 1.0
        self._dx  = off_range / max(len(recv_abs) - 1, 1)

        self._build_figure()

    # ---- figure ------------------------------------------------------------

    def _build_figure(self):
        c = _tc()
        self.fig, self.ax = plt.subplots(figsize=(16, 8), constrained_layout=False)
        if self._show_controls:
            self.fig.subplots_adjust(left=0.06, right=0.76, bottom=0.13, top=0.91)
        else:
            self.fig.subplots_adjust(left=0.06, right=0.98, bottom=0.13, top=0.91)
        self.fig.patch.set_facecolor(c["fig_bg"])
        self.ax.set_facecolor(c["ax_bg"])
        try:
            self.fig.canvas.manager.set_window_title(
                f"LVL Picker  --  {self.profile}  Shot {self.shot_id}")
        except Exception:
            pass
        self.fig.canvas.mpl_connect("button_press_event", self._on_click)
        self.fig.canvas.mpl_connect("motion_notify_event", self._on_motion)
        self.fig.canvas.mpl_connect("button_release_event", self._on_release)
        self.fig.canvas.mpl_connect("key_press_event",    self._on_key)
        if self._show_controls:
            self._build_controls()
        self._redraw()   # creates ax_top secondary axis inside

    def _build_controls(self):
        c = _tc()
        x0 = 0.78
        w  = 0.20
        row_h = 0.032
        gap = 0.005
        row3_w = (w - 2 * gap) / 3.0

        self.ax_btn_prev = self.fig.add_axes([x0 + 0 * (row3_w + gap), 0.86, row3_w, row_h])
        self.ax_btn_save = self.fig.add_axes([x0 + 1 * (row3_w + gap), 0.86, row3_w, row_h])
        self.ax_btn_next = self.fig.add_axes([x0 + 2 * (row3_w + gap), 0.86, row3_w, row_h])

        self.ax_btn_auto = self.fig.add_axes([x0 + 0 * (row3_w + gap), 0.82, row3_w, row_h])
        self.ax_btn_inv  = self.fig.add_axes([x0 + 1 * (row3_w + gap), 0.82, row3_w, row_h])
        self.ax_btn_tl   = self.fig.add_axes([x0 + 2 * (row3_w + gap), 0.82, row3_w, row_h])

        self.ax_btn_apply= self.fig.add_axes([x0, 0.78, w, row_h])

        self.btn_prev = Button(self.ax_btn_prev, "Prev", color="#444", hovercolor="#666")
        self.btn_save = Button(self.ax_btn_save, "Save", color="#444", hovercolor="#666")
        self.btn_next = Button(self.ax_btn_next, "Next", color="#444", hovercolor="#666")
        self.btn_auto = Button(self.ax_btn_auto, "Auto", color="#444", hovercolor="#666")
        self.btn_inv  = Button(self.ax_btn_inv, "Invert", color="#444", hovercolor="#666")
        self.btn_tl   = Button(self.ax_btn_tl, "Timeline", color="#444", hovercolor="#666")
        self.btn_apply= Button(self.ax_btn_apply, "Apply Filter", color="#444", hovercolor="#666")

        self.btn_prev.on_clicked(lambda _e: self._go_prev())
        self.btn_save.on_clicked(lambda _e: self._save_and_finish("next"))
        self.btn_next.on_clicked(lambda _e: self._save_and_finish("next"))
        self.btn_auto.on_clicked(lambda _e: self._auto_then_redraw())
        self.btn_inv.on_clicked(lambda _e: self._toggle_invert())
        self.btn_tl.on_clicked(lambda _e: self._toggle_timelines())
        self.btn_apply.on_clicked(lambda _e: self._apply_filter_controls())

        self.fig.text(x0, 0.755, "Gain", fontsize=8, color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.717, "AGC stat", fontsize=8, color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.679, "Display", fontsize=8, color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.603, "Filter (Hz)", fontsize=8, color=c["text"], ha="left", va="bottom")

        self._gain_labels = ("none", "norm", "agc")
        self._gain_btns = self._make_mode_buttons(
            x0=x0, y=0.725, w=w, h=row_h, labels=self._gain_labels,
            callback=self._on_gain_mode,
        )

        self._stat_labels = ("rms", "mean")
        self._stat_btns = self._make_mode_buttons(
            x0=x0, y=0.687, w=w, h=row_h, labels=self._stat_labels,
            callback=self._on_agc_stat,
        )

        self._disp_labels = ("wiggle", "vd", "both")
        self._disp_btns = self._make_mode_buttons(
            x0=x0, y=0.649, w=w, h=row_h, labels=self._disp_labels,
            callback=self._on_display_mode,
        )

        self.ax_agc = self.fig.add_axes([x0, 0.628, w, 0.016], facecolor=c["ax_bg"])
        self.sl_agc = Slider(self.ax_agc, "AGC", 5.0, 500.0,
                             valinit=float(self._agc_window_ms), valstep=1.0)
        self.sl_agc.on_changed(self._on_agc_window)

        self.ax_wig = self.fig.add_axes([x0, 0.608, w, 0.016], facecolor=c["ax_bg"])
        self.sl_wig = Slider(self.ax_wig, "Scale", 0.20, 5.00,
                     valinit=float(self._wiggle_stretch), valstep=0.01)
        self.sl_wig.on_changed(self._on_wiggle_stretch)

        self.ax_f1 = self.fig.add_axes([x0, 0.582, w, 0.014], facecolor=c["ax_bg"])
        self.ax_f2 = self.fig.add_axes([x0, 0.562, w, 0.014], facecolor=c["ax_bg"])
        self.ax_f3 = self.fig.add_axes([x0, 0.542, w, 0.014], facecolor=c["ax_bg"])
        self.ax_f4 = self.fig.add_axes([x0, 0.522, w, 0.014], facecolor=c["ax_bg"])
        self.sl_f1 = Slider(self.ax_f1, "f1", 0.0, 50.0, valinit=float(self._f1), valstep=0.5)
        self.sl_f2 = Slider(self.ax_f2, "f2", 0.5, 80.0, valinit=float(self._f2), valstep=0.5)
        self.sl_f3 = Slider(self.ax_f3, "f3", 20.0, 220.0, valinit=float(self._f3), valstep=1.0)
        self.sl_f4 = Slider(self.ax_f4, "f4", 40.0, 260.0, valinit=float(self._f4), valstep=1.0)

        # Header/parameter info panel
        self.ax_info = self.fig.add_axes([x0, 0.03, w, 0.47], facecolor=c["ax_bg"])
        self.ax_info.set_xticks([])
        self.ax_info.set_yticks([])
        for sp in self.ax_info.spines.values():
            sp.set_edgecolor(c["spine"])
        self.info_text = self.ax_info.text(0.02, 0.98, "", va="top", ha="left",
                                           fontsize=8, color=c["text"],
                                           transform=self.ax_info.transAxes)
        self._refresh_mode_buttons()

    def _make_mode_buttons(self, x0: float, y: float, w: float, h: float,
                           labels: tuple, callback: Any) -> list:
        gap = 0.004
        n = max(1, len(labels))
        bw = (w - gap * (n - 1)) / n
        buttons: list = []
        for i, lbl in enumerate(labels):
            axb = self.fig.add_axes([x0 + i * (bw + gap), y, bw, h])
            btn = Button(axb, str(lbl), color="#444", hovercolor="#666")
            btn.label.set_fontsize(8)
            btn.on_clicked(lambda _e, _lbl=str(lbl): callback(_lbl))
            buttons.append(btn)
        return buttons

    def _style_mode_buttons(self, buttons: list, labels: tuple, active_label: str):
        for btn, lbl in zip(buttons, labels):
            is_active = (str(lbl) == str(active_label))
            btn.ax.set_facecolor("#2a7fff" if is_active else "#444444")
            btn.label.set_color("white")

    def _refresh_mode_buttons(self):
        if not self._show_controls:
            return
        if hasattr(self, "_gain_btns"):
            self._style_mode_buttons(self._gain_btns, self._gain_labels, self._gain_mode)
        if hasattr(self, "_stat_btns"):
            self._style_mode_buttons(self._stat_btns, self._stat_labels, self._agc_stat)
        if hasattr(self, "_disp_btns"):
            self._style_mode_buttons(self._disp_btns, self._disp_labels, self._display_mode)

    def _on_gain_mode(self, label: str):
        self._gain_mode = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _on_agc_stat(self, label: str):
        self._agc_stat = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _on_agc_window(self, val: float):
        self._agc_window_ms = float(val)
        if self._gain_mode == "agc":
            self._redraw()

    def _on_wiggle_stretch(self, val: float):
        self._wiggle_stretch = float(val)
        self._redraw()

    def _on_display_mode(self, label: str):
        self._display_mode = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _toggle_invert(self):
        self._inverted = not self._inverted
        self._redraw()

    def _toggle_timelines(self):
        self._show_timelines = not self._show_timelines
        self._redraw()

    def _apply_filter_controls(self):
        if not self._show_controls:
            return
        f1 = float(self.sl_f1.val)
        f2 = float(self.sl_f2.val)
        f3 = float(self.sl_f3.val)
        f4 = float(self.sl_f4.val)
        if not (f1 < f2 < f3 < f4):
            print("     [WARN] Need f1 < f2 < f3 < f4")
            return
        self._f1, self._f2, self._f3, self._f4 = f1, f2, f3, f4
        self._recompute_filter()
        self._redraw()

    # ---- data --------------------------------------------------------------

    def _active_data(self) -> Any:
        d = self.data_filt if self._filter_on else self.data_raw
        d = apply_gain(d, self.dt_s, mode=self._gain_mode,
                       window_ms=self._agc_window_ms, stat=self._agc_stat)
        return -d if self._inverted else d

    def _recompute_filter(self):
        self.data_filt = apply_ormsby_all_params(
            self.data_raw, self.dt_s, self._f1, self._f2, self._f3, self._f4
        )

    # ---- drawing -----------------------------------------------------------

    def _draw_traces(self):
        c    = _tc()
        data = self._active_data()
        t_ms = self.times_ms[self.t_slice]
        dx   = self._dx

        if self._display_mode in ("vd", "both"):
            d_img = data[:, self.t_slice]
            vmax = float(np.percentile(np.abs(d_img), 98)) if d_img.size else 1.0
            if vmax < 1e-12:
                vmax = 1.0
            self.ax.imshow(
                d_img.T,
                aspect="auto",
                cmap="seismic",
                vmin=-vmax,
                vmax=vmax,
                extent=[self.recv_abs.min(), self.recv_abs.max(), t_ms[-1], t_ms[0]],
                alpha=0.45,
                interpolation="nearest",
                zorder=1,
            )

        if self._display_mode == "vd":
            return

        # Shared-median normalisation with per-trace clamping.
        # Using the global median keeps the filter effect visible (filtered
        # data appears quieter overall vs unfiltered).
        # Clamping each trace's own std to [median*0.3, median*3] prevents
        # loud near-offset traces from being over-squashed and dead traces
        # from being amplified into noise.
        stds  = np.array([data[i, self.t_slice].astype(float).std()
                          for i in range(self.n_traces)])
        valid = stds[stds > 1e-20]
        med   = float(np.median(valid)) if len(valid) else 1.0
        norms = np.clip(stds, med * 0.3, med * 3.0)
        norms = np.where(norms > 1e-20, norms, med)

        scale = norms * CLIP_FACTOR   # shape (n_traces,)

        for i, off in enumerate(self.recv_abs):
            tr   = data[i, self.t_slice].astype(float)
            tr_n = np.clip(tr / scale[i], -1.0, 1.0)
            x_w  = off + tr_n * dx * self._wiggle_stretch
            self.ax.plot(x_w, t_ms, color=c["trace"], lw=0.5, alpha=0.85)
            pos = np.where(tr_n > 0.0, tr_n, 0.0)
            self.ax.fill_betweenx(t_ms, off, off + pos * dx * self._wiggle_stretch,
                                  color=c["trace"], alpha=c["fill_alpha"])

    def _draw_picks(self):
        c = _tc()
        for idx, t in self._picks.items():
            if idx < len(self.recv_abs):
                self.ax.plot(self.recv_abs[idx], t, "v",
                             color=c["pick_clr"], ms=7, zorder=10,
                             markeredgecolor=c["pick_clr"],
                             markeredgewidth=0.6)

    def _redraw(self):
        # Remove the previous top axis BEFORE clearing the main axis;
        # twiny() axes are siblings in the figure and are NOT cleared by ax.clear().
        if getattr(self, 'ax_top', None) is not None:
            try:
                self.ax_top.remove()
            except Exception:
                pass
            self.ax_top = None

        c = _tc()
        self.ax.clear()
        self.ax.set_facecolor(c["ax_bg"])
        self._draw_traces()
        self._draw_picks()
        # Shot position marker (vertical) and trigger time (horizontal)
        self.ax.axvline(self.shot_pos_m, color="#ffcc00", lw=1.2,
                        ls="--", alpha=0.70, zorder=3)
        if self._show_timelines:
            self.ax.axhline(0.0, color="#00aa66", lw=1.0,
                            ls=":", alpha=0.75, zorder=3)
            for ms in range(10, int(self.t_disp_end) + 1, 10):
                if self.t_disp_start <= float(ms) <= self.t_disp_end:
                    major = (ms % 20 == 0)
                    self.ax.axhline(
                        float(ms),
                        color="#666666" if THEME == "light" else "#aaaaaa",
                        lw=0.50 if major else 0.30,
                        ls=":" if major else "--",
                        alpha=0.35 if major else 0.20,
                        zorder=2,
                    )

        margin = self._dx * 0.5
        self.ax.set_xlim(self.recv_abs.min() - margin,
                         self.recv_abs.max() + margin)
        # Y-axis: time increases downward; 0 ms at top, T_MAX_MS at bottom
        self.ax.set_ylim(self.t_disp_end, self.t_disp_start)

        dt_ms  = self.dt_s * 1000.0
        n_samp = self.n_samp
        t_end  = self.delay_ms + (n_samp - 1) * dt_ms
        filt_str = (f"Ormsby {self._f1:.0f}-{self._f2:.0f}-{self._f3:.0f}-{self._f4:.0f} Hz  "
                    + ("ON" if self._filter_on else "OFF")
                    + ("  [INV]" if self._inverted else ""))
        title = (f"Profile: {self.profile}  |  Shot {self.shot_id}  |"
                 f"  {n_samp} smp  dt={dt_ms:.4f} ms  "
                 f"delay={self.delay_ms:.1f} ms  end={t_end:.1f} ms  |"
                 f"  {len(self._picks)}/{self.n_traces} picks  |  {filt_str}"
                 f"  | gain={self._gain_mode}")
        self.ax.set_title(title, color=c["text"], fontsize=9)
        self.ax.set_xlabel(
            "Receiver position  (m)    "
            "[L] pick/drag   [R] del/drag   [Shift+L] range fill   [a] auto   [f] filter   "
            "[g] gain mode   [u/j] f2±1   [i/m] f3±5   [v] invert   [l] timelines   [t] theme   [c] ch↔offset   "
            "[s]/[n] save+next   [p] prev shot   [k] save PNG   [q] quit",
            color=c["label"], fontsize=8)
        self.ax.set_ylabel("Time  (ms)", color=c["label"])
        self.ax.grid(True, lw=0.3, alpha=0.30, color=c["grid"])
        self.ax.tick_params(colors=c["tick"])
        for sp in self.ax.spines.values():
            sp.set_edgecolor(c["spine"])
        handle = Line2D([0], [0], marker="v", linestyle="none",
                        color=c["pick_clr"], markersize=7,
                        label="First-break pick")
        self.ax.legend(handles=[handle], fontsize=8,
                       facecolor=c["leg_face"], edgecolor=c["leg_edge"],
                       labelcolor=c["text"], loc="lower right")

        # ---- Top axis: channel numbers (default) or signed offset from shot ----
        # twiny() shares the y-axis, giving an independent x-axis at the top.
        # We set matching xlim and place ticks at actual receiver positions so
        # the labels are exact regardless of whether spacing is uniform or not.
        ax_top = self.ax.twiny()
        ax_top.set_xlim(self.ax.get_xlim())
        n_recv = len(self.recv_abs)
        step   = max(1, n_recv // 8)          # aim for ~8-10 tick labels
        idxs   = list(range(0, n_recv, step))
        if (n_recv - 1) not in idxs:
            idxs.append(n_recv - 1)

        if self._top_mode == "channel":
            ax_top.set_xticks(self.recv_abs[idxs])
            ax_top.set_xticklabels([str(i + 1) for i in idxs], fontsize=7)
            ax_top.set_xlabel("Channel   [c → offset]",
                              color=c["label"], fontsize=8)
        else:
            spm = self.shot_pos_m
            ax_top.set_xticks(self.recv_abs[idxs])
            ax_top.set_xticklabels(
                [f"{self.recv_abs[i] - spm:+.0f}" for i in idxs], fontsize=7)
            ax_top.set_xlabel("Signed offset from shot  (m)   [c → channel]",
                              color=c["label"], fontsize=8)

        ax_top.tick_params(colors=c["tick"], labelsize=7)
        for sp in ax_top.spines.values():
            sp.set_edgecolor(c["spine"])
        self.ax_top = ax_top

        ffid = self._header_info.get("ffid", "?")
        ntr = self._header_info.get("n_tr", self.n_traces)
        nsamp = self._header_info.get("n_samp", self.n_samp)
        s_hdr = self._header_info.get("shot_pos_hdr", None)
        info_block = (
            f"Header info\n"
            f"FFID: {ffid}\n"
            f"Traces: {ntr}\n"
            f"Samples: {nsamp}\n"
            f"dt(ms): {dt_ms:.4f}\n"
            f"delay(ms): {self.delay_ms:.1f}  shot:{self.shot_pos_m:.2f}\n"
            f"gain:{self._gain_mode}({self._agc_stat},{self._agc_window_ms:.0f}ms)\n"
            f"disp:{self._display_mode}  scale:{self._wiggle_stretch:.2f}\n"
            f"f:{self._f1:.1f}-{self._f2:.1f}-{self._f3:.1f}-{self._f4:.1f}\n"
            f"inv:{'y' if self._inverted else 'n'} tl:{'y' if self._show_timelines else 'n'}"
        )
        if hasattr(self, "info_text"):
            self.info_text.set_text(info_block)

        self._refresh_mode_buttons()
        self.fig.patch.set_facecolor(c["fig_bg"])
        self.fig.canvas.draw_idle()

    # ---- events ------------------------------------------------------------

    def _nearest_idx(self, x: float) -> int:
        return int(np.abs(self.recv_abs - x).argmin())

    def _on_click(self, event: Any):
        # twiny() creates an overlay axis (ax_top) at the same figure position as
        # self.ax.  On Windows/TkAgg, mouse events are routed to whichever axes
        # was created last, so event.inaxes is often ax_top, not self.ax.
        # Fix: accept either axis, then convert pixel coordinates to self.ax
        # data space so xdata/ydata are always in the correct (geometry) units.
        ax_top_ref  = self.ax_top          # may be None before first _redraw
        valid_axes  = (self.ax,) if ax_top_ref is None else (self.ax, ax_top_ref)
        if event.inaxes not in valid_axes or event.x is None or event.y is None:
            return
        try:
            xdata, ydata = self.ax.transData.inverted().transform(
                (event.x, event.y))
        except Exception:
            if event.xdata is None or event.ydata is None:
                return
            xdata, ydata = event.xdata, event.ydata
        idx = self._nearest_idx(xdata)
        if event.key == "shift" and event.button == 1:
            self._range_fill(idx, float(ydata))
            self._redraw()
            return
        if event.button == 1:
            t = round(float(ydata), 2)
            if self.t_disp_start <= t <= self.t_disp_end:
                self._picks[idx] = t
                self._drag_pick = True
                self._last_drag_idx = idx
        elif event.button == 3:
            self._picks.pop(idx, None)
            self._drag_delete = True
            self._last_drag_idx = idx
        self._redraw()

    def _on_motion(self, event: Any):
        if not (self._drag_pick or self._drag_delete):
            return
        ax_top_ref = self.ax_top
        valid_axes = (self.ax,) if ax_top_ref is None else (self.ax, ax_top_ref)
        if event.inaxes not in valid_axes or event.x is None or event.y is None:
            return
        try:
            xdata, ydata = self.ax.transData.inverted().transform((event.x, event.y))
        except Exception:
            return
        idx = self._nearest_idx(xdata)
        if self._last_drag_idx == idx:
            return
        self._last_drag_idx = idx
        if self._drag_pick:
            t = round(float(ydata), 2)
            if self.t_disp_start <= t <= self.t_disp_end:
                self._picks[idx] = t
        elif self._drag_delete:
            self._picks.pop(idx, None)
        self._redraw()

    def _on_release(self, _event: Any):
        self._drag_pick = False
        self._drag_delete = False
        self._last_drag_idx = None

    def _range_fill(self, idx: int, ydata: float):
        t = round(float(ydata), 2)
        if not (self.t_disp_start <= t <= self.t_disp_end):
            return
        if self._range_anchor is None:
            self._range_anchor = (idx, t)
            print(f"     Range anchor set at trace {idx+1}, t={t:.2f} ms")
            return
        i0, t0 = self._range_anchor
        i1, t1 = idx, t
        if i0 == i1:
            self._picks[i0] = t1
            self._range_anchor = None
            return
        lo, hi = (i0, i1) if i0 < i1 else (i1, i0)
        for ii in range(lo, hi + 1):
            frac = (ii - i0) / float(i1 - i0)
            tt = round(float(t0 + frac * (t1 - t0)), 2)
            if self.t_disp_start <= tt <= self.t_disp_end:
                self._picks[ii] = tt
        self._range_anchor = None
        print(f"     Range fill: traces {lo+1}-{hi+1}")

    def _save_and_finish(self, nav: str):
        if self._save_callback is not None:
            self._save_callback(self._picks)
        try:
            self._save_qc_image()
        except Exception:
            pass
        self._saved = True
        self._nav_action = nav
        self._done = True
        try:
            self.fig.canvas.stop_event_loop()
        except Exception:
            pass

    def _go_prev(self):
        self._save_and_finish("prev")

    def _auto_then_redraw(self):
        self._auto_pick()
        self._redraw()

    def _on_key(self, event: Any):
        global THEME
        key = event.key
        if key in ("q", "escape"):
            self._cancelled = True
            self._nav_action = "quit"
            self._done = True
            try: self.fig.canvas.stop_event_loop()
            except Exception: pass
        elif key == "s":
            self._save_and_finish("next")
        elif key == "n":
            self._save_and_finish("next")
        elif key == "p":
            self._go_prev()
        elif key == "k":
            # Save a per-shot QC screenshot (separate from pick-saving)
            self._save_qc_image()
        elif key == "f":
            self._filter_on = not self._filter_on
            self._redraw()
        elif key == "u":
            self._f2 = max(self._f1 + 0.5, self._f2 + 1.0)
            self._recompute_filter()
            self._redraw()
        elif key == "j":
            self._f2 = max(self._f1 + 0.5, self._f2 - 1.0)
            self._recompute_filter()
            self._redraw()
        elif key == "i":
            self._f3 = min(self._f4 - 1.0, self._f3 + 5.0)
            self._recompute_filter()
            self._redraw()
        elif key == "m":
            self._f3 = max(self._f2 + 1.0, self._f3 - 5.0)
            self._recompute_filter()
            self._redraw()
        elif key == "g":
            self._gain_mode = ({"none": "norm", "norm": "agc", "agc": "none"}
                               .get(self._gain_mode, "none"))
            self._redraw()
        elif key == "v":
            self._inverted = not self._inverted
            self._redraw()
        elif key == "l":
            self._show_timelines = not self._show_timelines
            self._redraw()
        elif key == "t":
            THEME = "light" if THEME == "dark" else "dark"
            self._redraw()
        elif key == "a":
            self._auto_pick()
            self._redraw()
        elif key == "c":
            self._top_mode = "offset" if self._top_mode == "channel" else "channel"
            self._redraw()

    def _auto_pick(self):
        data = self._active_data()
        count = 0
        n_sta = max(1, int(STA_MS / 1000.0 / self.dt_s))
        n_lta = max(3, int(LTA_MS / 1000.0 / self.dt_s))
        start_idx = max(0, int((0.0 - self.delay_ms) / (self.dt_s * 1000.0)))
        end_idx = min(self.n_samp - n_sta - 1,
                      int((self.t_disp_end - self.delay_ms) / (self.dt_s * 1000.0)))

        for i in range(self.n_traces):
            tr = data[i].astype(float)
            tr = tr - np.mean(tr[:max(1, n_lta)])
            char = np.maximum(tr, 0.0)
            picked_idx = None
            for k in range(max(n_lta, start_idx), max(n_lta + 1, end_idx)):
                lta = char[k - n_lta:k].mean()
                if lta < 1e-30:
                    continue
                sta = char[k:k + n_sta].mean()
                ratio = sta / lta
                if ratio >= STALTA_TRIG:
                    picked_idx = k
                    break
            if picked_idx is None:
                continue

            t_abs = round(self.delay_ms + picked_idx * self.dt_s * 1000.0, 2)
            if self.t_disp_start <= t_abs <= self.t_disp_end:
                self._picks[i] = t_abs
                count += 1

        print(f"     Auto-pick: {count}/{self.n_traces} placed (gain-aware).")

    def _save_qc_image(self):
        """Save the main plot area only (without control widgets) as PNG."""
        self.qc_dir.mkdir(parents=True, exist_ok=True)
        out = self.qc_dir / f"{self.profile}_shot{self.shot_id:02d}_picks.png"
        self.fig.canvas.draw()
        renderer = self.fig.canvas.get_renderer()
        bbox = self.ax.get_tightbbox(renderer).expanded(1.02, 1.04)
        bbox_inches = bbox.transformed(self.fig.dpi_scale_trans.inverted())
        self.fig.savefig(str(out), dpi=180, bbox_inches=bbox_inches,
                         facecolor=self.ax.get_facecolor())
        print(f"\n     QC image -> {out.name}")

    def _execute_control_command(self, cmd: dict):
        action = str(cmd.get("action", "")).strip().lower()
        if not action:
            return

        if action == "next":
            self._save_and_finish("next")
            return
        if action == "prev":
            self._go_prev()
            return
        if action == "save":
            self._save_and_finish("next")
            return
        if action == "quit":
            self._cancelled = True
            self._nav_action = "quit"
            self._done = True
            try:
                self.fig.canvas.stop_event_loop()
            except Exception:
                pass
            return
        if action == "auto":
            self._auto_pick()
            self._redraw()
            return
        if action == "invert":
            self._toggle_invert()
            return
        if action == "timeline":
            self._toggle_timelines()
            return
        if action == "filter_toggle":
            self._filter_on = not self._filter_on
            self._redraw()
            return
        if action == "save_image":
            self._save_qc_image()
            return
        if action == "gain":
            mode = str(cmd.get("value", "")).strip().lower()
            if mode in ("none", "norm", "agc"):
                self._on_gain_mode(mode)
            return
        if action == "agc_stat":
            stat = str(cmd.get("value", "")).strip().lower()
            if stat in ("rms", "mean"):
                self._on_agc_stat(stat)
            return
        if action == "display":
            mode = str(cmd.get("value", "")).strip().lower()
            if mode in ("wiggle", "vd", "both"):
                self._on_display_mode(mode)
            return
        if action == "agc_window":
            try:
                self._on_agc_window(float(cmd.get("value")))
            except Exception:
                pass
            return
        if action == "wiggle_scale":
            try:
                self._on_wiggle_stretch(float(cmd.get("value")))
            except Exception:
                pass
            return
        if action == "set_filter":
            try:
                f1 = float(cmd.get("f1", self._f1))
                f2 = float(cmd.get("f2", self._f2))
                f3 = float(cmd.get("f3", self._f3))
                f4 = float(cmd.get("f4", self._f4))
            except Exception:
                return
            if f1 < f2 < f3 < f4:
                self._f1, self._f2, self._f3, self._f4 = f1, f2, f3, f4
                self._recompute_filter()
                self._redraw()

    def _poll_control_file(self):
        if self._control_file is None:
            return
        try:
            if not self._control_file.exists():
                return
            with open(self._control_file, encoding="utf-8") as fh:
                payload = json.load(fh)
            if isinstance(payload, dict):
                cmd_id = int(payload.get("id", -1))
                if cmd_id <= self._last_control_id:
                    return
                self._last_control_id = cmd_id
                self._execute_control_command(payload)
        except Exception:
            return

    # ---- public ------------------------------------------------------------

    @property
    def picks_ms(self) -> dict:
        return self._picks

    def run(self) -> Any:
        """
        Show exactly ONE picker window and block until it is done.

        Why this approach:
        - plt.show(block=True) on Windows/TkAgg re-enters mainloop().
          If the first mainloop() was already stopped by plt.close(), the
          second call may not block, causing all 3 shot figures to be created
          before any are shown (6-window bug).
        - canvas.start_event_loop(timeout) runs the backend event loop for
          exactly `timeout` seconds per call, then returns.  This gives a
          tightly-controlled poll loop that is fully responsive.
        - Key handlers set self._done = True and call stop_event_loop() to
          return the current start_event_loop call immediately (no plt.close
          inside key handlers).
        - plt.close() is called by run() AFTER the loop exits, exactly once.

        Returns raw picks {trace_idx: ms from trigger} or None if cancelled.
        """
        self._done = False

        def _on_close(_evt: Any):
            # Fired when the user clicks the window X button
            if not (self._saved or self._cancelled):
                self._cancelled = True
            self._done = True
            try: self.fig.canvas.stop_event_loop()
            except Exception: pass

        self.fig.canvas.mpl_connect("close_event", _on_close)
        plt.show(block=False)   # display this figure (non-blocking)

        # Poll in short chunks; each chunk runs the native event loop so the
        # window remains fully interactive
        while not self._done:
            try:
                self.fig.canvas.start_event_loop(0.05)
            except Exception:
                self._cancelled = True
                break
            self._poll_control_file()

        # Close the figure if still open (X button may have already closed it)
        try:
            if plt.fignum_exists(self.fig.number):
                plt.close(self.fig)
        except Exception:
            pass

        if self._cancelled:
            return {"status": "quit", "picks": self._picks}
        return {"status": self._nav_action or "next", "picks": self._picks}


# ---------------------------------------------------------------------------
# Picks persistence  (JSON -- always RAW, no static applied)
# ---------------------------------------------------------------------------

def _picks_json_path(profile_name: str) -> Path:
    return OUTPUT_DIR / profile_name / "picks.json"


def load_picks_json(profile_name: str) -> dict:
    """Load raw picks: {shot_id: {trace_idx: ms}}."""
    p = _picks_json_path(profile_name)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as fh:
            raw = json.load(fh)
        return {int(k): {int(ti): float(tv) for ti, tv in v.items()}
                for k, v in raw.items()}
    except Exception as exc:
        print(f"  [WARN] Could not load picks.json: {exc}")
        return {}


def save_picks_json(profile_name: str, all_picks: dict):
    p = _picks_json_path(profile_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump({str(k): {str(ti): tv for ti, tv in v.items()}
                   for k, v in all_picks.items()},
                  fh, indent=2)
    print(f"     Picks saved -> {p.relative_to(CWD.parent)}")


# ---------------------------------------------------------------------------
# Export: picks_clean.txt  (compatible with lvl.ipynb)
# ---------------------------------------------------------------------------

def export_picks_txt(profile_name: str, shots_info: list,
                     all_picks: dict, recv_positions: Any) -> Path:
    """Write Ensemble / SOURCE / CHAN / OFFSET / FB_PICK text file."""
    out_dir  = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    txt_path = out_dir / f"{profile_name}_picks_clean.txt"
    header   = (f"{'Ensemble':>10} {'#':>4} {'SOURCE':>8} "
                f"{'CHAN':>6} {'OFFSET':>10} {'FB_PICK':>10}")

    with open(txt_path, "w", encoding="utf-8") as fh:
        fh.write(header + "\n")
        ens = 1
        for shot_idx, (shot_id, shot_pos_m) in enumerate(shots_info):
            corr = apply_bulk_static(all_picks.get(shot_id, {}))
            if not corr:
                continue
            if shot_idx > 0:
                fh.write("\n")
            for trace_idx in sorted(corr):
                if trace_idx >= len(recv_positions):
                    continue
                offset_m = recv_positions[trace_idx] - shot_pos_m
                fh.write(f"{ens:>10} {ens:>4} {shot_id:>8} "
                         f"{trace_idx+1:>6} {offset_m:>10.2f} "
                         f"{corr[trace_idx]:>10.3f}\n")
                ens += 1

    print(f"  picks_txt -> {txt_path.relative_to(CWD.parent)}")
    return txt_path


# ---------------------------------------------------------------------------
# Export: Excel workbook  (Config + per-shot formulas + Analysis)
# ---------------------------------------------------------------------------

_FILL_HDR  = PatternFill("solid", fgColor="2E4057")
_FILL_IN   = PatternFill("solid", fgColor="FFFACD")   # lemon  = editable input
_FILL_FORM = PatternFill("solid", fgColor="E8F4F8")   # blue   = Excel formula
_FILL_OK   = PatternFill("solid", fgColor="DDFFDD")   # green  = depth result
_FONT_HDR  = Font(bold=True, color="FFFFFF")
_FONT_BOLD = Font(bold=True)
_FONT_ITA  = Font(italic=True, color="888888")


def _chdr(ws: Any, row: int, col: int, val: str):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _FILL_HDR
    c.font = _FONT_HDR
    c.alignment = Alignment(horizontal="center")


def _autofit_xl(ws: Any, min_w: int = 6, max_w: int = 30):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None),
                   default=min_w)
        ws.column_dimensions[col[0].column_letter].width = min(best + 3, max_w)


def export_excel(profile_name: str, shots_info: list,
                 all_picks: dict, recv_positions: Any,
                 analysis: dict, cfg: dict) -> Path:
    """
    Write <profile>_picks.xlsx with four sheets.

    Config    Yellow cells = user input.
              B5 = perp_m, B6 = bulk_static_ms.
              All per-shot formula cells reference these two cells so
              changing perp_m or the static updates everything instantly.

    Shot_N    Per-shot picks with Excel formulas:
                True_off  = SQRT(Inline_off^2 + Config!$B$5^2)
                FB_corr   = FB_raw + Config!$B$6
              Summary rows: SLOPE / INTERCEPT / RSQ on corrected picks.

    Combined  All shots concatenated.

    Analysis  Python-computed velocities, intercept times, depths (green),
              R^2, RMS.  Depth formula reference below the table.
    """
    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    xl_path = out_dir / f"{profile_name}_picks.xlsx"
    wb      = openpyxl.Workbook()

    # -- Config sheet -------------------------------------------------------
    ws_cfg = wb.active
    ws_cfg.title = "Config"
    ws_cfg["A1"] = f"LVL Refraction Analysis  --  Profile {profile_name}"
    ws_cfg["A1"].font = Font(bold=True, size=13)
    ws_cfg.merge_cells("A1:D1")

    def _inp(row: int, name: str, value: Any, unit: str = "", note: str = ""):
        ws_cfg.cell(row=row, column=1, value=name).font = _FONT_BOLD
        c = ws_cfg.cell(row=row, column=2, value=value)
        c.fill = _FILL_IN
        c.alignment = Alignment(horizontal="right")
        ws_cfg.cell(row=row, column=3, value=unit)
        if note:
            ws_cfg.cell(row=row, column=4, value=note).font = _FONT_ITA

    _inp(3,  "Profile name",      profile_name,           "",   "Folder name under data/")
    _inp(4,  "Geometry type",     cfg.get("geom", "?"),   "m",  "100 = 100m spread,  200 = 200m spread")
    _inp(5,  "Perp. distance",    cfg.get("perp_m", 0.0), "m",
             "Perpendicular shot-to-line offset.  "
             "True_off = SQRT(inline^2 + perp^2)  <- referenced in Shot sheets col G")
    _inp(6,  "Bulk static",       BULK_SHIFT_MS,          "ms",
             "Added to raw FB times.  ProMAX: bulk shift static -> add.  "
             "Negative = earlier.  <- referenced in Shot sheets col I")
    _inp(7,  "N layers",          N_LAYERS,                "",   "2 or 3")
    _inp(8,  "Ormsby f1",         BP_F1,                  "Hz", "Low-cut ramp start")
    _inp(9,  "Ormsby f2",         BP_F2,                  "Hz", "Low-cut ramp end / pass start")
    _inp(10, "Ormsby f3",         BP_F3,                  "Hz", "Pass end / high-cut ramp start")
    _inp(11, "Ormsby f4",         BP_F4,                  "Hz", "High-cut ramp end")
    _inp(12, "FFT zero-padding",  int(BP_FFT_PAD * 100),  "%",  "25 = ProMAX default")
    _inp(13, "STA window",        STA_MS,                 "ms", "Short-term average (auto-pick)")
    _inp(14, "LTA window",        LTA_MS,                 "ms", "Long-term average (auto-pick)")
    _inp(15, "STA/LTA threshold", STALTA_TRIG,             "",  "Trigger ratio")

    ws_cfg.column_dimensions["A"].width = 24
    ws_cfg.column_dimensions["B"].width = 14
    ws_cfg.column_dimensions["C"].width = 6
    ws_cfg.column_dimensions["D"].width = 60

    # -- Per-shot sheets ----------------------------------------------------
    combined_rows: list = []

    for shot_id, shot_pos_m in shots_info:
        raw_picks = all_picks.get(shot_id, {})
        if not raw_picks:
            continue

        ws = wb.create_sheet(title=f"Shot_{shot_id}")
        col_headers = ["Trace", "FFID", "Recv_pos_m", "Shot_pos_m",
                       "Inline_off_m", "Perp_m", "True_off_m",
                       "FB_raw_ms", "FB_corr_ms"]
        for ci, h in enumerate(col_headers, 1):
            _chdr(ws, 1, ci, h)

        sorted_idxs = sorted(raw_picks)
        for row_i, trace_idx in enumerate(sorted_idxs, start=2):
            if trace_idx >= len(recv_positions):
                continue
            recv_pos = float(recv_positions[trace_idx])
            inline   = recv_pos - shot_pos_m
            fb_raw   = raw_picks[trace_idx]

            ws.cell(row=row_i, column=1, value=trace_idx + 1)           # Trace
            ws.cell(row=row_i, column=2, value=shot_id)                  # FFID
            ws.cell(row=row_i, column=3, value=round(recv_pos, 3))       # Recv_pos_m
            ws.cell(row=row_i, column=4, value=round(shot_pos_m, 3))     # Shot_pos_m
            ws.cell(row=row_i, column=5, value=round(inline, 3))         # Inline_off_m
            # F: perp from Config (user edits it there; formula picks it up here)
            ws.cell(row=row_i, column=6, value="=Config!$B$5").fill = _FILL_FORM
            # G: True offset = SQRT(inline^2 + perp^2)
            ws.cell(row=row_i, column=7,
                    value=f"=SQRT(E{row_i}^2+F{row_i}^2)").fill = _FILL_FORM
            # H: raw FB pick (from picker, no static)
            ws.cell(row=row_i, column=8, value=round(fb_raw, 3))
            # I: corrected FB = raw + bulk static
            ws.cell(row=row_i, column=9,
                    value=f"=H{row_i}+Config!$B$6").fill = _FILL_FORM

            combined_rows.append({
                "shot_id": shot_id, "trace":    trace_idx + 1,
                "recv_pos": recv_pos, "shot_pos": shot_pos_m,
                "inline":   inline,   "fb_raw":   fb_raw,
            })

        # Summary: SLOPE / INTERCEPT / RSQ on corrected picks vs true offset
        n_data  = len(sorted_idxs)
        sum_row = n_data + 3
        _chdr(ws, sum_row, 1, "Metric")
        _chdr(ws, sum_row, 2, "Value")
        _chdr(ws, sum_row, 3, "Range used")

        g_rng = f"G2:G{n_data + 1}"
        i_rng = f"I2:I{n_data + 1}"
        for ri, (label, formula, basis) in enumerate([
            ("Velocity (m/s)",
             f"=IFERROR(1000/SLOPE({i_rng},{g_rng}),\"n/a\")",
             "SLOPE(FB_corr, True_off)"),
            ("Intercept (ms)",
             f"=IFERROR(INTERCEPT({i_rng},{g_rng}),\"n/a\")",
             "INTERCEPT(FB_corr, True_off)"),
            ("R^2",
             f"=IFERROR(RSQ({i_rng},{g_rng}),\"n/a\")",
             "RSQ(FB_corr, True_off)"),
        ], start=sum_row + 1):
            ws.cell(row=ri, column=1, value=label).font = _FONT_BOLD
            ws.cell(row=ri, column=2, value=formula).fill = _FILL_FORM
            ws.cell(row=ri, column=3, value=basis).font  = _FONT_ITA

        _autofit_xl(ws)

    # -- Combined sheet -----------------------------------------------------
    if combined_rows:
        ws_c = wb.create_sheet(title="Combined")
        for ci, h in enumerate(["Shot_ID", "Trace", "Recv_pos_m",
                                 "Shot_pos_m", "Inline_off_m", "FB_raw_ms"], 1):
            _chdr(ws_c, 1, ci, h)
        for ri, row in enumerate(combined_rows, start=2):
            ws_c.cell(row=ri, column=1, value=row["shot_id"])
            ws_c.cell(row=ri, column=2, value=row["trace"])
            ws_c.cell(row=ri, column=3, value=round(row["recv_pos"], 3))
            ws_c.cell(row=ri, column=4, value=round(row["shot_pos"], 3))
            ws_c.cell(row=ri, column=5, value=round(row["inline"], 3))
            ws_c.cell(row=ri, column=6, value=round(row["fb_raw"], 3))
        _autofit_xl(ws_c)

    # -- Analysis sheet -----------------------------------------------------
    ws_a = wb.create_sheet(title="Analysis")
    a_hdrs = ["Shot_ID", "FFID", "V1 (m/s)", "V2 (m/s)", "V3 (m/s)",
              "ti_1 (ms)", "ti_2 (ms)", "h1 (m)", "h2 (m)", "RMS (ms)", "n_picks"]
    for ci, h in enumerate(a_hdrs, 1):
        _chdr(ws_a, 1, ci, h)

    row_a = 2
    for shot_id, _ in shots_info:
        res = (analysis or {}).get(shot_id)
        if res is None:
            continue
        deps = res.get("depths", {})
        segs = res.get("segments", [])

        def _vel(i: int) -> Any:
            return round(segs[i]["velocity_m_s"], 1) if i < len(segs) else ""

        def _dep(key: str) -> Any:
            val = deps.get(key)
            return round(val, 2) if val else ""

        ws_a.cell(row=row_a, column=1,  value=shot_id)
        ws_a.cell(row=row_a, column=2,  value=shot_id)
        ws_a.cell(row=row_a, column=3,  value=_vel(0))
        ws_a.cell(row=row_a, column=4,  value=_vel(1))
        ws_a.cell(row=row_a, column=5,  value=_vel(2))
        ws_a.cell(row=row_a, column=6,
                  value=round(deps["ti1_ms"], 3) if deps.get("ti1_ms") else "")
        ws_a.cell(row=row_a, column=7,
                  value=round(deps["ti2_ms"], 3) if deps.get("ti2_ms") else "")
        ws_a.cell(row=row_a, column=8,  value=_dep("h1_m"))
        ws_a.cell(row=row_a, column=9,  value=_dep("h2_m"))
        ws_a.cell(row=row_a, column=10, value=round(res.get("rms_ms", 0.0), 3))
        ws_a.cell(row=row_a, column=11, value=len(all_picks.get(shot_id, {})))

        for col in (8, 9):
            cell = ws_a.cell(row=row_a, column=col)
            if cell.value != "":
                cell.fill = _FILL_OK
        row_a += 1

    note_row = row_a + 2
    notes = [
        "Depth formulas (intercept-time method):",
        "  2-layer:  h1 = (ti1_ms/1000 x V1 x V2) / (2 x SQRT(V2^2 - V1^2))",
        "  3-layer:  h2 = (ti2_eff_ms/1000 x V1 x V3) / (2 x SQRT(V3^2 - V1^2))"
        "    where  ti2_eff = ti2 - 2 x h1 x cos(ic12) / V1 x 1000",
    ]
    for oi, note in enumerate(notes):
        r = note_row + oi
        ws_a.cell(row=r, column=1, value=note).font = (
            _FONT_BOLD if oi == 0 else _FONT_ITA)
        ws_a.merge_cells(f"A{r}:K{r}")

    _autofit_xl(ws_a)
    wb.save(str(xl_path))
    print(f"  Excel     -> {xl_path.relative_to(CWD.parent)}")
    return xl_path


# ---------------------------------------------------------------------------
# Export: T-X picks plot  (time increases downward + analysis overlay)
# ---------------------------------------------------------------------------

def export_tx_plot(profile_name: str, shots_info: list,
                   all_picks: dict, recv_positions: Any,
                   analysis: dict, perp_m: float = 0.0,
                   shot_label_pos: dict | None = None) -> Path:
    """
    Final T-X summary plot.
    - Scatter of corrected picks (absolute true offset vs corrected pick time)
    - Dashed fitted-line overlay with velocity labels
    - Y-axis inverted: t = 0 at top  (seismic convention: time grows downward)
    - Theme follows the current THEME setting at export time
    """
    c = _tc()
    fig, ax = plt.subplots(figsize=(10, 6))
    fig.patch.set_facecolor(c["fig_bg"])
    ax.set_facecolor(c["ax_bg"])

    pal = ["#e63946", "#2a9d8f", "#e9c46a", "#a8dadc", "#f4a261", "#457b9d"]
    mks = ["o", "s", "^", "D", "P", "X"]

    for shot_id, shot_pos_m in shots_info:
        corr = apply_bulk_static(all_picks.get(shot_id, {}))
        if not corr:
            continue
        col = pal[(shot_id - 1) % len(pal)]
        mk  = mks[(shot_id - 1) % len(mks)]
        label_pos_m = (shot_label_pos or {}).get(shot_id, shot_pos_m)

        x_geom, t_vals = [], []
        for idx in sorted(corr):
            if idx < len(recv_positions):
                x_geom.append(float(recv_positions[idx]))
                t_vals.append(corr[idx])

        ax.scatter(x_geom, t_vals, color=col, marker=mk, s=28, zorder=5,
               label=f"Shot {shot_id}  (@ {label_pos_m:.1f} m)")

        # Dashed fitted T-X lines (only when analysis data is available)
        res = (analysis or {}).get(shot_id)
        if res and res.get("segments"):
            for seg in res["segments"]:
                # Convert segment x-range (absolute offset) back to geometry x
                # using shot position so lines align with geometry x-axis
                xg0 = shot_pos_m + seg["x_start"]
                xg1 = shot_pos_m + seg["x_end"]
                sl  = seg["slope_ms_m"]
                ic  = seg["intercept_ms"]
                vel = seg["velocity_m_s"]
                xl  = np.array([xg0, xg1])
                # Predicted time at geometry x using offset from shot
                t_line = sl * np.abs(xl - shot_pos_m) + ic
                ax.plot(xl, t_line, "--", color=col,
                        lw=1.6, alpha=c["fit_alpha"])
                xm = (xg0 + xg1) / 2.0
                tm = sl * abs(xm - shot_pos_m) + ic
                ax.text(xm, tm - 3.5,
                        f"{vel:.0f} m/s", color=col, fontsize=7,
                        ha="center", fontweight="bold")

    # Seismic convention: t = 0 at top, time grows downward
    ax.set_ylim(T_MAX_MS, 0.0)
    ax.set_xlabel("Receiver position  (m)", color=c["label"])
    ax.set_ylabel("First-break time  (ms)", color=c["label"])
    ax.set_title(f"T-X first-break picks  --  Profile {profile_name}",
                 color=c["text"], fontsize=11)
    ax.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax.tick_params(colors=c["tick"])
    for sp in ax.spines.values():
        sp.set_edgecolor(c["spine"])
    ax.legend(fontsize=9, facecolor=c["leg_face"],
              edgecolor=c["leg_edge"], labelcolor=c["text"])
    fig.tight_layout()

    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{profile_name}_tx_picks.png"
    fig.savefig(str(out), dpi=180, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  T-X plot  -> {out.relative_to(CWD.parent)}")
    return out


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------

def process_profile(profile_name: str, pick_mode: bool = True,
                    geom_override: int | None = None,
                    control_file: Path | None = None,
                    show_plot_controls: bool = True):
    cfg = PROFILES.get(profile_name)
    if cfg is None:
        print(f"[ERROR] Profile '{profile_name}' not in PROFILES. "
              f"Known: {list(PROFILES)}")
        return

    data_dir = DATA_DIR / profile_name
    if not data_dir.exists():
        print(f"[ERROR] Data folder not found: {data_dir}")
        return

    geom_type      = int(geom_override) if geom_override is not None else int(cfg["geom"])
    perp_m         = float(cfg.get("perp_m", 0.0))
    recv_positions = load_geometry(geom_type)

    # Resolve shot positions: "auto" derives them from geometry
    shots_cfg = cfg.get("shots", "auto")
    if shots_cfg == "auto" or not isinstance(shots_cfg, dict):
        shots_cfg = auto_shot_positions(recv_positions)
        print(f"  Shot positions (auto from geometry): "
              + "  ".join(f"Shot{k}={v:.3f}m" for k, v in shots_cfg.items()))

    print(f"  Geometry {geom_type} m : {len(recv_positions)} receivers, "
          f"{recv_positions[0]:.2f} - {recv_positions[-1]:.2f} m  "
          f"|  perp = {perp_m:.1f} m  "
          f"|  bulk static = {BULK_SHIFT_MS:+.1f} ms")

    def _ffid(p: Path) -> int:
        digits = "".join(c for c in p.stem if c.isdigit())
        return int(digits) if digits else 0

    seg2_candidates = list(data_dir.glob("*.seg2")) + list(data_dir.glob("*.SEG2"))
    seg2_unique = {str(p.resolve()).lower(): p for p in seg2_candidates}
    seg2_files = sorted(seg2_unique.values(), key=_ffid)
    if not seg2_files:
        print(f"[ERROR] No .seg2 files found in {data_dir}")
        return
    print(f"  Found {len(seg2_files)} SEG2 file(s)")

    all_picks: dict   = load_picks_json(profile_name)
    shots_meta: list = []
    qc_dir = OUTPUT_DIR / profile_name

    shot_cache: list = []
    for file_idx, seg2_path in enumerate(seg2_files):
        shot_id = file_idx + 1
        data_raw, dt_s, n_tr, n_samp, shot_pos_hdr, ffid_hdr, \
            delay_ms, recv_locs_hdr = read_seg2(seg2_path)
        n_show = min(n_tr, len(recv_positions))
        shot_pos_cfg = shots_cfg.get(shot_id)
        if shot_pos_hdr is not None:
            shot_pos_m = float(shot_pos_hdr)
        elif shot_pos_cfg is not None:
            shot_pos_m = float(shot_pos_cfg)
        else:
            shot_pos_m = 0.0
        shot_pos_nominal = (float(shot_pos_cfg)
                            if shot_pos_cfg is not None else float(shot_pos_m))
        shot_cache.append({
            "shot_id": shot_id,
            "seg2_path": seg2_path,
            "data_raw": data_raw,
            "dt_s": dt_s,
            "n_tr": n_tr,
            "n_samp": n_samp,
            "shot_pos_hdr": shot_pos_hdr,
            "ffid_hdr": ffid_hdr,
            "delay_ms": delay_ms,
            "n_show": n_show,
            "shot_pos_m": shot_pos_m,
            "shot_pos_nominal": shot_pos_nominal,
        })

    idx = 0
    while 0 <= idx < len(shot_cache):
        shot = shot_cache[idx]
        shot_id = int(shot["shot_id"])
        seg2_path = shot["seg2_path"]
        print(f"\n  -- Shot {shot_id}  ({seg2_path.name}) --")

        data_raw = shot["data_raw"]
        dt_s = float(shot["dt_s"])
        n_tr = int(shot["n_tr"])
        n_samp = int(shot["n_samp"])
        shot_pos_hdr = shot["shot_pos_hdr"]
        ffid_hdr = shot["ffid_hdr"]
        delay_ms = float(shot["delay_ms"])
        n_show = int(shot["n_show"])
        shot_pos_m = float(shot["shot_pos_m"])

        dt_ms  = dt_s * 1000.0
        t_end  = delay_ms + (n_samp - 1) * dt_ms
        print(f"     FFID={ffid_hdr}  |  {n_tr} traces  |  "
              f"dt={dt_ms:.4f} ms  |  delay={delay_ms:.1f} ms  |  "
              f"{n_samp} smp  ({delay_ms:.1f} to {t_end:.1f} ms)")

        if shot_pos_hdr is not None:
            print(f"     Shot pos : {shot_pos_m:.2f} m  (SEG2 header)")
        elif shots_cfg.get(shot_id) is not None:
            print(f"     Shot pos : {shot_pos_m:.2f} m  (CONFIG)")
        else:
            print("     Shot pos : 0.0 m  [WARN: defaulting to 0]")

        if pick_mode:
            # Always use geometry file positions for the x-axis display and picking.
            # SEG2 RECEIVER_LOCATION may differ from the geometry file (e.g. a
            # 200-m spread recorded with 0-94 m header values).
            geom_slice = recv_positions[:n_show]

            data_slice = data_raw[:n_show]
            data_filt  = apply_ormsby_all(data_slice, dt_s)

            def _save_cb(picks_for_shot: dict, _sid: int = shot_id) -> None:
                """Called immediately when 's' is pressed; writes JSON to disk."""
                all_picks[_sid] = picks_for_shot
                save_picks_json(profile_name, all_picks)
                print(f"     ✓ {len(picks_for_shot)} pick(s) saved to disk immediately.")

            picker = FirstBreakPicker(
                data_slice, data_filt, dt_s,
                geom_slice, shot_id, profile_name,
                shot_pos_m=shot_pos_m,
                delay_ms=delay_ms,
                existing_picks=all_picks.get(shot_id, {}),
                qc_dir=qc_dir,
                save_callback=_save_cb,
                show_controls=show_plot_controls,
                control_file=control_file,
                header_info={
                    "ffid": ffid_hdr,
                    "n_tr": n_tr,
                    "n_samp": n_samp,
                    "shot_pos_hdr": shot_pos_hdr,
                },
            )
            result = picker.run() or {"status": "quit", "picks": all_picks.get(shot_id, {})}
            status = result.get("status", "next")

            if all_picks.get(shot_id):
                print(f"     Shot {shot_id}: {len(all_picks[shot_id])} pick(s) in memory.")
            else:
                print(f"     Picking cancelled / no picks for shot {shot_id}.")

            if status == "prev":
                idx = max(0, idx - 1)
                continue
            if status == "quit":
                break

        idx += 1

    for shot in shot_cache:
        shot_id = int(shot["shot_id"])
        shots_meta.append((shot_id, float(shot["shot_pos_m"]), float(shot["shot_pos_nominal"])))

    total_picks = sum(len(v) for v in all_picks.values())
    if total_picks == 0:
        print("\n  No picks to export.")
        return

    # --- Refraction analysis (deferred: run once picking is finalised) ----
    # Uncomment the block below when ready to compute velocities and depths.
    # analysis = analyse_shots(shot_results, perp_m=perp_m)
    # for sid, res in analysis.items(): ...
    analysis: dict = {}   # placeholder until analysis phase is activated

    # Exports
    print(f"\n  -- Exporting  ({total_picks} picks) --")
    shots_info_proc = [(sid, sp_proc) for sid, sp_proc, _ in shots_meta]
    shot_label_pos  = {sid: sp_nom for sid, _, sp_nom in shots_meta}

    export_picks_txt(profile_name, shots_info_proc, all_picks, recv_positions)
    export_excel(profile_name, shots_info_proc, all_picks, recv_positions,
                 analysis, cfg)
    export_tx_plot(profile_name, shots_info_proc, all_picks, recv_positions,
                   analysis, perp_m=perp_m, shot_label_pos=shot_label_pos)
    print(f"\n  Output -> {(OUTPUT_DIR / profile_name).relative_to(CWD.parent)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        prog="lvl_refraction.py",
        description="LVL refraction seismic: interactive picking + analysis.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python lvl_refraction.py 120\n"
            "  python lvl_refraction.py 120 --export-only\n"
            "  python lvl_refraction.py --all\n"
        ),
    )
    parser.add_argument("profile", nargs="?", default=None,
                        help="Profile folder name (e.g. 120 or 214_A)")
    parser.add_argument("geometry", nargs="?", default=None,
                        help="Optional geometry override: 100, 200, geometry100, geometry200")
    parser.add_argument("--all", action="store_true",
                        help="Process all profiles in PROFILES")
    parser.add_argument("--export-only", action="store_true",
                        help="Skip picking; re-analyse and re-export existing picks")
    parser.add_argument("--control-file", default=None,
                        help="Path to JSON command file for external UI control")
    parser.add_argument("--minimal-plot-controls", action="store_true",
                        help="Hide in-plot buttons/widgets; keep keyboard interactions")
    args = parser.parse_args()

    geom_override: int | None = None
    if args.geometry is not None:
        g = str(args.geometry).strip().lower().replace(" ", "")
        if g in ("100", "geometry100"):
            geom_override = 100
        elif g in ("200", "geometry200"):
            geom_override = 200
        else:
            print("[ERROR] Invalid geometry override. Use one of:")
            print("        100, 200, geometry100, geometry200")
            return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.all:
        targets = list(PROFILES)
    elif args.profile:
        targets = [args.profile]
    else:
        chosen_profile = None
        try:
            root = tk.Tk()
            root.withdraw()
            sel = filedialog.askdirectory(
                title="Select profile data folder",
                initialdir=str(DATA_DIR),
                mustexist=True,
            )
            root.destroy()
            if sel:
                p = Path(sel)
                if p.parent.resolve() == DATA_DIR.resolve():
                    chosen_profile = p.name
        except Exception:
            chosen_profile = None

        if chosen_profile:
            print(f"  Selected folder -> profile '{chosen_profile}'")
            targets = [chosen_profile]
        else:
            print("\nAvailable profiles:")
            print(f"  {'Name':<12}  {'Geom':>6}  {'Perp_m':>7}  Data folder")
            print(f"  {'-'*12}  {'-'*6}  {'-'*7}  {'-'*30}")
            for pname, pcfg in PROFILES.items():
                folder = DATA_DIR / pname
                status = "found" if folder.exists() else "MISSING"
                n_seg2 = len(list(folder.glob("*.seg2"))) if folder.exists() else 0
                print(f"  {pname:<12}  {pcfg['geom']:>5}m  "
                      f"{pcfg.get('perp_m', 0.0):>7.1f}m  "
                      f"{status}  ({n_seg2} SEG2 files)")
            print(f"\nUsage:  python lvl_refraction.py <profile>  [--export-only]")
            return

    control_file = Path(args.control_file) if args.control_file else None

    for pname in targets:
        print(f"\n{'='*60}\nProfile : {pname}\n{'='*60}")
        process_profile(pname, pick_mode=not args.export_only,
                        geom_override=geom_override,
                        control_file=control_file,
                        show_plot_controls=not args.minimal_plot_controls)

    print("\nAll done.")


if __name__ == "__main__":
    main()
