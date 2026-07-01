"""
lvl_refraction.py  --  First-break picking + LVL refraction analysis
====================================================================
Project : Low Velocity Layer (LVL) refraction seismic surveys

Workflow (mirrors ProMAX flow)
------------------------------
1.  Read SEG2 files -- one file per shot, sorted by FFID (file number)
2.  Assign receiver geometry from standard geometry files
    (geometry100.txt = 100 m spread,  geometry200.txt = 200 m spread)
3.  Apply bulk time static  (like ProMAX: bulk shift static -> add)
4.  Apply zero-phase Ormsby bandpass filter
    (frequency domain, 25 % zero-padding for FFT, 4 corner frequencies)
5.  Interactive first-break picking  (matplotlib GUI)
6.  Automatic refraction analysis:
    -- AIC-based breakpoint detection (or manual)
    -- T-X segment fitting -> velocities
    -- Perpendicular offset correction  (true offset = sqrt(inline^2 + perp^2))
    -- Intercept-time depth inversion (2- or 3-layer)
7.  Export:
    -- <profile>_picks.xlsx     Config + per-shot formulas + Analysis sheet
    -- <profile>_picks_clean.txt  compatible with lvl.ipynb
    -- <profile>_tx_picks.png     T-X summary with fitted lines, time downward
    -- <profile>_shot_XX_picks.png  per-shot QC image (saved on shot navigation/finalize)

Interactive picker controls
---------------------------
    Left-click / drag       place picks quickly (interpolates across skipped traces)
    Right-click / drag      delete picks quickly (range delete while dragging)
    Shift + Left-click      range-fill picks between two traces
    n                       save to session and go to next shot
    p                       save to session and go to previous shot
    s                       Save/Close (finalize picks + exports)
    a                       STA/LTA auto-pick (gain-aware)
    f                       toggle Ormsby filter on / off
    v                       invert trace polarity (display only)
    g                       cycle gain mode: norm/agc/none
    l                       toggle timeline guides
    c                       toggle top axis: channel <-> signed offset
    k                       save per-shot QC screenshot (PNG)
    q / Esc                 quit picking

Directory layout
----------------
  data/
    geometry100.txt
    geometry200.txt
    120/       Rec_00001.seg2  Rec_00002.seg2  Rec_00003.seg2
    150/  ...
  scripts/
    lvl_refraction.py   <- this file
  output/
    120/       <auto-created>

How to run
----------
  conda activate seiseng
  cd d:\\Daten\\seismic\\lvl\\scripts

  python lvl_refraction.py 120              # interactive picking for profile 120
  python lvl_refraction.py 120 --export-only  # re-export without picking
  python lvl_refraction.py --all            # pick all profiles in sequence
  python lvl_refraction.py                  # list available profiles
"""

from __future__ import annotations

import sys
import argparse
import json
import math
import importlib
import subprocess
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Auto-install missing packages into the active environment
# ---------------------------------------------------------------------------
def _ensure(pkg: str, mod: str = "") -> Any:
    m = mod or pkg
    try:
        return importlib.import_module(m)
    except ImportError:
        print(f"  Installing '{pkg}' ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        return importlib.import_module(m)

np = _ensure("numpy")
pd = _ensure("pandas")
_ensure("obspy")
_ensure("matplotlib")
_ensure("openpyxl")
_ensure("scipy")
_ensure("xlrd")

from obspy import read as _read_obspy          # type: ignore[import-untyped]
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.widgets import Button, Slider, TextBox
from scipy.signal import butter as _scipy_butter, sosfiltfilt as _scipy_sosfiltfilt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


# ===========================================================================
# CONFIG  --  edit this section for each project
# ===========================================================================

CWD        = Path(__file__).parent
DATA_DIR   = CWD.parent / "data"
OUTPUT_DIR = CWD.parent / "output"

# Geometry files: map trace number -> absolute receiver position (m along profile)
GEOM_FILES: dict = {
    100: DATA_DIR / "geometry100.txt",
    200: DATA_DIR / "geometry200.txt",
}

# ---------------------------------------------------------------------------
# Profile registry
# ---------------------------------------------------------------------------
# geom    : geometry type (100 or 200)
# line_no : line identifier used for labelling
# perp_m  : perpendicular distance of shot from the geophone line (m)
# shots   : {shot_id: inline_position_m}  OR  "auto"
#           "auto" = derive from geometry:
#             shot 1 = recv[0] - first_spacing     (before spread)
#             shot 2 = (recv[n//2-1] + recv[n//2]) / 2   (midpoint)
#             shot 3 = recv[-1] + last_spacing     (after spread)
#           Manual values are used when SEG2 SOURCE_LOCATION is absent.
#           SEG2 header always takes priority if present.
# ---------------------------------------------------------------------------
PROFILES: dict = {
    "120":   {"geom": 100, "line_no": 1, "perp_m": 0.0, "shots": "auto"},
    "214_A": {"geom": 100, "line_no": 2, "perp_m": 0.0, "shots": "auto"},
    "150":   {"geom": 200, "line_no": 3, "perp_m": 0.0, "shots": "auto"},
    "415_B": {"geom": 200, "line_no": 4, "perp_m": 0.0, "shots": "auto"},
}

# Shot position source policy:
# False -> use geometry/config shot positions by default (recommended)
# True  -> allow SEG2 SOURCE_LOCATION to override when available
USE_SEG2_SHOT_POSITION: bool = False

# ---------------------------------------------------------------------------
# Ormsby bandpass  (zero-phase, frequency domain)
# ProMAX convention:  f1 - f2 - f3 - f4
#   f1 : low-cut ramp start      (Hz)
#   f2 : low-cut ramp end / passband start   (Hz)
#   f3 : passband end / high-cut ramp start  (Hz)
#   f4 : high-cut ramp end       (Hz)
# ---------------------------------------------------------------------------
BP_F1: float = 2.0    # Hz
BP_F2: float = 4.0    # Hz
BP_F3: float = 140.0  # Hz
BP_F4: float = 180.0  # Hz
BP_FFT_PAD: float = 0.25   # 25 % zero-padding (ProMAX default)
BP_REAPPLY: bool  = False  # True = apply twice (squares amplitude response)
BUTTER_ORDER: int = 4      # Butterworth order (zero-phase via sosfiltfilt)
FILTER_DEBOUNCE_MS: int = 40  # 0 = immediate live update, >0 = debounce while dragging

# ---------------------------------------------------------------------------
# Bulk time static  (ProMAX: "Bulk shift static -> Add")
# The SEG2 DELAY field (e.g. -106 ms) is read from each file and used to
# build the correct absolute time axis.  Picks are stored as absolute times
# from the trigger (t = 0).  BULK_SHIFT_MS is an ADDITIONAL correction on
# top of the DELAY  (leave at 0.0 unless a systematic offset remains).
# ---------------------------------------------------------------------------
BULK_SHIFT_MS: float = 0.0   # ms  (negative = shift to earlier times)

# ---------------------------------------------------------------------------
# Display / theme
# ---------------------------------------------------------------------------
THEME: str         = "light"   # "dark"  or  "light"
T_MAX_MS: float    = 150.0    # end of display window (ms from trigger)
T_DISPLAY_PRE_MS: float = -10.0  # ms before trigger to show (headroom above first arrivals)
CLIP_FACTOR: float = 2.0      # wiggle clip level in std-dev multiples
FILTER_ON: bool    = True     # start with filter active
DISPLAY_MODE: str  = "wiggle"   # "wiggle" | "vd" | "both"
SHOW_TIMELINES: bool = True
WIGGLE_STRETCH: float = 1.0

# ---------------------------------------------------------------------------
# Gain control for display/picking
# ---------------------------------------------------------------------------
GAIN_MODE: str        = "norm"   # "none" | "norm" | "agc"
AGC_WINDOW_MS: float  = 200.0
AGC_STAT: str         = "rms"    # "mean" | "rms"

# ---------------------------------------------------------------------------
# STA/LTA auto-picker
# ---------------------------------------------------------------------------
STA_MS: float      = 3.0
LTA_MS: float      = 20.0
STALTA_TRIG: float = 3.0

# ---------------------------------------------------------------------------
# Refraction analysis
# ---------------------------------------------------------------------------
N_LAYERS: int          = 2      # 2 or 3
AUTO_BREAKPOINTS: bool = True   # True = AIC search;  False = MANUAL_BREAKS
# MANUAL_BREAKS: {shot_id: [receiver_index_of_crossover, ...]}
# Example:  {1: [12], 2: [14], 3: [12]}
MANUAL_BREAKS: dict = {}

# ===========================================================================
# END OF CONFIG
# ===========================================================================


# ---------------------------------------------------------------------------
# Theme colour palette
# ---------------------------------------------------------------------------
def _tc() -> dict:
    """Return a colour dict for the current THEME."""
    if THEME == "dark":
        return dict(
            fig_bg="#0f1117",   ax_bg="#1a1d2e",
            trace="#4a8fc1",    fill_alpha=0.20,
            pick_clr="#e63946", text="white",
            label="#aaa",       grid="#888",
            tick="#888",        spine="#333344",
            leg_face="#1e2130", leg_edge="#555",
            fit_alpha=0.90,
        )
    return dict(
        fig_bg="white",     ax_bg="#f5f5f5",
        trace="#111111",    fill_alpha=0.18,
        pick_clr="#cc0000", text="#111111",
        label="#333",       grid="#bbb",
        tick="#333",        spine="#aaaaaa",
        leg_face="#f0f0f0", leg_edge="#aaa",
        fit_alpha=0.85,
    )


def _ensure_interactive_backend() -> bool:
    """Ensure matplotlib uses an interactive backend if possible."""
    def _is_non_interactive(name: str) -> bool:
        n = str(name).strip().lower()
        non_interactive = {
            "agg", "pdf", "ps", "svg", "template", "cairo",
            "module://matplotlib_inline.backend_inline",
            "module://ipykernel.pylab.backend_inline",
        }
        return n in non_interactive

    try:
        backend = str(plt.get_backend()).lower()
    except Exception:
        backend = ""
    if backend and not _is_non_interactive(backend):
        return True

    for candidate in ("TkAgg", "QtAgg", "Qt5Agg"):
        try:
            plt.switch_backend(candidate)
            break
        except Exception:
            continue

    try:
        new_backend = str(plt.get_backend()).lower()
    except Exception:
        new_backend = ""

    if (not new_backend) or _is_non_interactive(new_backend):
        print("  [WARN] Matplotlib backend is non-interactive (Agg); UI plots cannot open.")
        return False

    print(f"  [INFO] Switched matplotlib backend to interactive mode: {plt.get_backend()}")
    return True


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def load_geometry(geom_type: int) -> Any:
    """Return receiver positions (m) as a 1-D ndarray."""
    path = GEOM_FILES[geom_type]
    positions: list = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            s = line.strip()
            if not s or s.lower().startswith("trace"):
                continue
            # Handle German decimal comma:  "2,5" -> 2.5
            parts = s.replace(",", ".").split()
            if len(parts) >= 2:
                try:
                    positions.append(float(parts[1]))
                except ValueError:
                    pass
    return np.array(positions, dtype=float)


def auto_shot_positions(recv_pos: Any) -> dict:
    """
    Compute standard refraction shot positions from the receiver geometry.

    Convention used by most land refraction surveys:
      Shot 1  :  one receiver-spacing BEFORE the first receiver
      Shot 2  :  midpoint between receivers n//2 and n//2+1
                 (labelled e.g. GP24.5 for a 48-channel spread)
      Shot 3  :  one receiver-spacing AFTER the last receiver

    Returns {1: pos_m, 2: pos_m, 3: pos_m}.
    """
    n          = len(recv_pos)
    dx_start   = float(recv_pos[1]    - recv_pos[0])    # first spacing
    dx_end     = float(recv_pos[-1]   - recv_pos[-2])   # last spacing
    mid_lo     = float(recv_pos[n // 2 - 1])
    mid_hi     = float(recv_pos[n // 2])
    return {
        1: round(float(recv_pos[0]) - dx_start, 4),
        2: round((mid_lo + mid_hi) / 2.0,        4),
        3: round(float(recv_pos[-1]) + dx_end,   4),
    }


def _excel_col_to_index(col: str) -> int:
    """Convert Excel column label (A, D, AA, ...) to zero-based index."""
    s = str(col or "").strip().upper()
    if not s or not all("A" <= ch <= "Z" for ch in s):
        raise ValueError(f"Invalid Excel column label '{col}'")
    out = 0
    for ch in s:
        out = out * 26 + (ord(ch) - ord("A") + 1)
    return out - 1


def _normalize_profile_token(value: Any) -> str:
    """
    Normalize profile strings for robust matching.

    Examples:
      "LVL150" -> "150"
      "150" -> "150"
      "LVL214_A" -> "214_A"
    """
    s = str(value or "").strip().upper().replace(" ", "")
    s = s.replace("_", "").replace("-", "")
    if s.startswith("LVL"):
        s = s[3:]
    return s


def discover_field_report_excel(data_dir: Path) -> Path | None:
    """Find likely field-report Excel files in data directory; return newest match."""
    pats = [
        "field_report*.xls",
        "field_report*.xlsx",
        "fieldreport*.xls",
        "fieldreport*.xlsx",
    ]
    cand: list = []
    for pat in pats:
        cand.extend(list(data_dir.glob(pat)))
    if not cand:
        return None
    cand_sorted = sorted(cand, key=lambda p: p.stat().st_mtime, reverse=True)
    return cand_sorted[0]


def _read_excel_table(path: Path, sheet_name: str | int | None = None) -> Any:
    """Read Excel with fallback engines for both .xls and .xlsx files."""
    sheet = (0 if sheet_name is None else sheet_name)

    # First try pandas default engine resolution.
    try:
        return pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    except Exception:
        pass

    suffix = str(path.suffix).lower()
    last_exc: Exception | None = None

    if suffix == ".xls":
        try:
            return pd.read_excel(path, sheet_name=sheet, header=None, dtype=object, engine="xlrd")
        except Exception as exc:
            last_exc = exc
    else:
        try:
            return pd.read_excel(path, sheet_name=sheet, header=None, dtype=object, engine="openpyxl")
        except Exception as exc:
            last_exc = exc

    if last_exc is not None:
        raise last_exc
    raise ValueError(f"Could not read Excel file: {path}")


def infer_geometry_from_field_report(path: Path,
                                     profile_name: str,
                                     sheet_name: str | int | None = None) -> int | None:
    """
    Infer geometry from field report notes near profile markers.

        Rule:
            - find row containing LVL<profile>
            - inspect nearby rows (+/- 1..2)
      - if any cell contains 'short' (or 'short profile'), geometry = 100
      - otherwise return None
    """
    if not path.exists():
        return None

    df = _read_excel_table(path, sheet_name=sheet_name)
    if df is None or df.empty:
        return None

    prof_tok = _normalize_profile_token(profile_name)
    lvl_tag = f"LVL{prof_tok}"
    n_rows = int(df.shape[0])

    def _row_text(ridx: int) -> str:
        vals = [str(v) for v in df.iloc[ridx].tolist() if pd.notna(v)]
        return " ".join(vals).strip().lower()

    for ridx in range(n_rows):
        row_txt_up = _row_text(ridx).upper().replace(" ", "")
        if lvl_tag not in row_txt_up:
            continue

        for delta in (-2, -1, 1, 2):
            rr = ridx + delta
            if rr >= n_rows:
                continue
            if rr < 0:
                continue
            txt = _row_text(rr)
            if "short profile" in txt or "short" in txt:
                return 100
    return None


def load_profile_offsets_from_excel(
    path: Path,
    profile_name: str,
    ffid_by_shot: dict | None = None,
    sheet_name: str | int | None = None,
    ffid_col: str = "A",
    perp_col: str = "D",
    profile_col: str = "F",
    inline_shift_col: str | None = None,
) -> tuple:
    """
    Load per-shot perpendicular offsets (and optional inline shift) from Excel.

        Expected table convention:
            - vertically organized profile blocks
            - first row contains LVLXXX in profile column
            - continuation rows follow until next LVL row
            - shot numbering often lives in column A as 1/2/3

    Returns
    -------
    (perp_by_shot, inline_shift_by_shot)
    """
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    df = _read_excel_table(path, sheet_name=sheet_name)
    if df.empty:
        return {}, {}

    i_ffid = _excel_col_to_index(ffid_col)
    i_perp = _excel_col_to_index(perp_col)
    i_prof = _excel_col_to_index(profile_col)
    i_shift = _excel_col_to_index(inline_shift_col) if inline_shift_col else None
    need_cols = [i_ffid, i_perp, i_prof] + ([i_shift] if i_shift is not None else [])
    if max(need_cols) >= df.shape[1]:
        raise ValueError(
            f"Excel has {df.shape[1]} columns; requested column index {max(need_cols)+1} is out of range"
        )

    target = _normalize_profile_token(profile_name)
    ffid_to_shots: dict = {}
    for sid in sorted(ffid_by_shot.keys() if ffid_by_shot else []):
        try:
            ff = int((ffid_by_shot or {}).get(sid, 0))
            ss = int(sid)
        except Exception:
            continue
        if ff <= 0:
            continue
        ffid_to_shots.setdefault(ff, []).append(ss)

    rows_target: list = []
    n_rows = int(df.shape[0])
    lvl_starts: list = []
    target_starts: list = []
    for ridx in range(n_rows):
        if i_prof >= df.shape[1]:
            continue
        p_raw = df.iat[ridx, i_prof]
        p_txt = "" if pd.isna(p_raw) else str(p_raw).strip()
        p_up = p_txt.upper().replace(" ", "")
        if "LVL" not in p_up:
            continue
        lvl_starts.append(ridx)
        if _normalize_profile_token(p_txt) == target:
            target_starts.append(ridx)

    for start in target_starts:
        next_starts = [r for r in lvl_starts if r > start]
        end = min(next_starts) if next_starts else n_rows
        for ridx in range(start, end):
            ffid_raw = df.iat[ridx, i_ffid]
            perp_raw = df.iat[ridx, i_perp]
            shift_raw = df.iat[ridx, i_shift] if i_shift is not None else 0.0

            ffid = None
            perp = None
            shift = 0.0
            try:
                if pd.notna(ffid_raw) and str(ffid_raw).strip() != "":
                    ffid = int(float(str(ffid_raw).replace(",", ".")))
            except Exception:
                ffid = None
            try:
                if pd.notna(perp_raw) and str(perp_raw).strip() != "":
                    perp = float(str(perp_raw).replace(",", "."))
            except Exception:
                perp = None
            try:
                if pd.notna(shift_raw) and str(shift_raw).strip() != "":
                    shift = float(str(shift_raw).replace(",", "."))
            except Exception:
                shift = 0.0

            if perp is None:
                continue

            rows_target.append({"ffid": ffid, "perp": float(perp), "shift": float(shift)})

    if not rows_target:
        return {}, {}

    perp_by_shot: dict = {}
    shift_by_shot: dict = {}

    # Many field sheets use column A as shot index (1/2/3).
    shot_numbers = [int(r["ffid"]) for r in rows_target if r.get("ffid") is not None]
    if shot_numbers:
        uniq = sorted(set(shot_numbers))
        if uniq and min(uniq) >= 1 and max(uniq) <= max(10, len(rows_target) + 1):
            for row in rows_target:
                sid = row.get("ffid")
                if sid is None:
                    continue
                perp_by_shot[int(sid)] = float(row["perp"])
                shift_by_shot[int(sid)] = float(row["shift"])
            return perp_by_shot, shift_by_shot

    if ffid_to_shots:
        ffid_cursor: dict = {k: 0 for k in ffid_to_shots}
        for row in rows_target:
            ffid = row["ffid"]
            if ffid is None:
                continue
            ff_key = int(ffid)
            shot_list = ffid_to_shots.get(ff_key, [])
            if not shot_list:
                continue
            cur = int(ffid_cursor.get(ff_key, 0))
            sid = shot_list[min(cur, len(shot_list) - 1)]
            ffid_cursor[ff_key] = cur + 1
            perp_by_shot[int(sid)] = float(row["perp"])
            shift_by_shot[int(sid)] = float(row["shift"])
    else:
        for i, row in enumerate(rows_target, start=1):
            perp_by_shot[int(i)] = float(row["perp"])
            shift_by_shot[int(i)] = float(row["shift"])

    return perp_by_shot, shift_by_shot

def read_perpendicular_offsets(path: Path) -> dict:
    """
    Read a simple text file with per-shot perpendicular offsets.

    Format:
        # shot_id  perp_m
        1         0.0
        2         3.5
        3         0.0
    """
    out: dict = {}
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            s = str(line).strip()
            if not s or s.startswith("#"):
                continue
            parts = s.replace(",", ".").split()
            if len(parts) < 2:
                continue
            try:
                sid = int(parts[0])
                pov = float(parts[1])
            except Exception:
                continue
            out[int(sid)] = float(pov)
    return out

def true_offset(inline_m: Any, perp_m: float) -> Any:
    """True source-receiver distance corrected for perpendicular shot offset."""
    return np.sqrt(np.asarray(inline_m, dtype=float) ** 2 + perp_m ** 2)


def parse_shot_value_map(text: str | None) -> dict:
    """
    Parse mappings like "1:0,2:3.5,3:0" into {1: 0.0, 2: 3.5, 3: 0.0}.
    """
    out: dict = {}
    if not text:
        return out
    for chunk in str(text).split(","):
        part = chunk.strip()
        if not part:
            continue
        if ":" not in part:
            raise ValueError(f"Invalid mapping '{part}'. Expected shot:value")
        k_s, v_s = part.split(":", 1)
        shot_id = int(k_s.strip())
        val = float(v_s.strip().replace(",", "."))
        out[shot_id] = val
    return out


def resolve_perp_by_shot(cfg: dict, shot_ids: list,
                         override_map: dict | None = None) -> dict:
    """
    Resolve perpendicular offsets per shot from profile config + CLI overrides.
    cfg['perp_m'] can be either scalar or dict {shot_id: perp_m}.
    """
    base = cfg.get("perp_m", 0.0)
    per_shot: dict = {}

    if isinstance(base, dict):
        for sid in shot_ids:
            per_shot[int(sid)] = float(base.get(int(sid), 0.0))
    else:
        base_val = float(base)
        for sid in shot_ids:
            per_shot[int(sid)] = base_val

    for sid, val in (override_map or {}).items():
        per_shot[int(sid)] = float(val)

    return per_shot


def prompt_offset_model_by_shot(perp_by_shot: dict,
                                inline_shift_by_shot: dict | None = None,
                                shots_info: list | None = None,
                                all_picks: dict | None = None,
                                recv_positions: Any | None = None) -> tuple:
    """
    Interactive UI to edit per-shot geometric corrections.

    - PO (m): perpendicular source-to-line distance.
    - X-shift (m): optional inline distance shift for far-offset acquisition geometry.

    A live preview panel redraws all shots together when values change.
    """
    out_po = {int(k): float(v) for k, v in (perp_by_shot or {}).items()}
    out_shift = {int(k): float(v) for k, v in (inline_shift_by_shot or {}).items()}
    for sid in out_po:
        out_shift.setdefault(int(sid), 0.0)

    if not out_po:
        return out_po, out_shift

    try:
        backend = str(plt.get_backend()).lower()
    except Exception:
        backend = ""
    if "agg" in backend and not _ensure_interactive_backend():
        print("  [INFO] Using default PO values (interactive PO window unavailable).")
        return out_po, out_shift

    c = _tc()
    shot_ids = sorted(out_po)
    n = len(shot_ids)
    fig_h = max(5.6, min(11.0, 3.3 + 0.55 * n))
    fig = plt.figure(figsize=(11.2, fig_h))
    fig.patch.set_facecolor(c["fig_bg"])
    fig.text(0.04, 0.95, "Shot Geometry Setup (PO + X-shift)",
             color=c["text"], fontsize=12, fontweight="bold", va="top")
    fig.text(0.04, 0.90,
             "Adjust PO and optional inline shift. Press Enter in a field for live update.",
             color=c["label"], fontsize=9, va="top")

    top = 0.82
    row_h = 0.055
    textboxes_po: dict = {}
    textboxes_shift: dict = {}
    for i, sid in enumerate(shot_ids):
        y = top - i * row_h
        fig.text(0.04, y + 0.017, f"Shot {sid}", color=c["text"], fontsize=9,
                 ha="left", va="center")
        ax_tb_po = fig.add_axes([0.11, y, 0.12, 0.038])
        tb_po = TextBox(ax_tb_po, "PO", initial=f"{out_po[sid]:.3f}")
        tb_po.label.set_color(c["text"])
        tb_po.label.set_fontsize(8)
        tb_po.text_disp.set_color(c["text"])
        tb_po.text_disp.set_fontsize(9)
        ax_tb_po.set_facecolor(c["ax_bg"])
        for sp in ax_tb_po.spines.values():
            sp.set_edgecolor(c["spine"])

        ax_tb_shift = fig.add_axes([0.25, y, 0.12, 0.038])
        tb_shift = TextBox(ax_tb_shift, "X-shift", initial=f"{out_shift.get(sid, 0.0):.3f}")
        tb_shift.label.set_color(c["text"])
        tb_shift.label.set_fontsize(8)
        tb_shift.text_disp.set_color(c["text"])
        tb_shift.text_disp.set_fontsize(9)
        ax_tb_shift.set_facecolor(c["ax_bg"])
        for sp in ax_tb_shift.spines.values():
            sp.set_edgecolor(c["spine"])

        textboxes_po[sid] = tb_po
        textboxes_shift[sid] = tb_shift

    ax_prev = fig.add_axes([0.41, 0.16, 0.55, 0.70])
    ax_prev.set_facecolor(c["ax_bg"])
    for sp in ax_prev.spines.values():
        sp.set_edgecolor(c["spine"])
    ax_prev.grid(True, lw=0.3, alpha=0.30, color=c["grid"])
    ax_prev.tick_params(colors=c["tick"])
    ax_prev.set_xlabel("True offset (m)", color=c["label"], fontsize=8)
    ax_prev.set_ylabel("FB time (ms)", color=c["label"], fontsize=8)
    ax_prev.set_title("Live preview: all shots", color=c["text"], fontsize=9)

    def _read_float(tb: Any, fallback: float) -> float:
        try:
            txt = str(tb.text).strip()
        except Exception:
            txt = ""
        if not txt:
            return float(fallback)
        try:
            return float(txt.replace(",", "."))
        except Exception:
            return float(fallback)

    def _read_ui_values() -> tuple:
        po_new = dict(out_po)
        sh_new = dict(out_shift)
        for sid in shot_ids:
            po_new[sid] = _read_float(textboxes_po[sid], po_new.get(sid, 0.0))
            sh_new[sid] = _read_float(textboxes_shift[sid], sh_new.get(sid, 0.0))
        return po_new, sh_new

    def _redraw_preview(po_map: dict, sh_map: dict):
        ax_prev.clear()
        ax_prev.set_facecolor(c["ax_bg"])
        ax_prev.grid(True, lw=0.3, alpha=0.30, color=c["grid"])
        ax_prev.tick_params(colors=c["tick"])
        ax_prev.set_xlabel("True offset (m)", color=c["label"], fontsize=8)
        ax_prev.set_ylabel("FB time (ms)", color=c["label"], fontsize=8)
        ax_prev.set_title("Live preview: all shots", color=c["text"], fontsize=9)

        if not (shots_info and all_picks and recv_positions is not None):
            ax_prev.text(0.5, 0.5, "Preview unavailable (missing picks/geometry)",
                         transform=ax_prev.transAxes, ha="center", va="center",
                         color=c["label"], fontsize=8)
            fig.canvas.draw_idle()
            return

        pal = ["#e63946", "#2a9d8f", "#e9c46a", "#457b9d", "#f4a261", "#8d99ae"]
        plotted = 0
        for i, (sid, shot_pos) in enumerate(shots_info):
            sid_i = int(sid)
            picks = apply_bulk_static((all_picks or {}).get(sid_i, {}))
            if not picks:
                continue
            po = float(po_map.get(sid_i, 0.0))
            x_shift = float(sh_map.get(sid_i, 0.0))
            xvals: list = []
            tvals: list = []
            for tr in sorted(picks):
                if tr >= len(recv_positions):
                    continue
                inline_abs = abs(float(recv_positions[tr]) - float(shot_pos))
                inline_corr = max(0.0, inline_abs + x_shift)
                x_true = float(true_offset(inline_corr, po))
                xvals.append(x_true)
                tvals.append(float(picks[tr]))
            if len(xvals) < 2:
                continue
            col = pal[i % len(pal)]
            order = np.argsort(np.asarray(xvals, dtype=float))
            xs = np.asarray(xvals, dtype=float)[order]
            ts = np.asarray(tvals, dtype=float)[order]
            ax_prev.plot(xs, ts, "o-", ms=3.5, lw=1.0, alpha=0.9, color=col,
                         label=f"Shot {sid_i}  PO={po:.2f}m  dX={x_shift:.2f}m")
            plotted += 1

        if plotted > 0:
            ax_prev.legend(fontsize=7, facecolor=c["leg_face"], edgecolor=c["leg_edge"],
                           labelcolor=c["text"], loc="best")
            ax_prev.set_ylim(T_MAX_MS, 0.0)
        else:
            ax_prev.text(0.5, 0.5, "No picks available for preview", transform=ax_prev.transAxes,
                         ha="center", va="center", color=c["label"], fontsize=8)
        fig.canvas.draw_idle()

    def _apply_preview(_evt: Any = None):
        po_map, sh_map = _read_ui_values()
        _redraw_preview(po_map, sh_map)

    _apply_preview()

    status_ax = fig.add_axes([0.04, 0.08, 0.58, 0.05])
    status_ax.set_facecolor(c["ax_bg"])
    status_ax.set_xticks([])
    status_ax.set_yticks([])
    for sp in status_ax.spines.values():
        sp.set_edgecolor(c["spine"])
    status_txt = status_ax.text(0.02, 0.5, "Ready", color=c["text"],
                                fontsize=8, va="center", ha="left",
                                transform=status_ax.transAxes)

    ax_ok = fig.add_axes([0.64, 0.08, 0.15, 0.05])
    ax_ref = fig.add_axes([0.81, 0.08, 0.15, 0.05])
    ax_def = fig.add_axes([0.81, 0.14, 0.15, 0.05])
    btn_ok = Button(ax_ok, "Use Values", color=c["ax_bg"],
                    hovercolor="#e8e8e8" if THEME == "light" else "#2a2d3d")
    btn_ref = Button(ax_ref, "Refresh", color=c["ax_bg"],
                     hovercolor="#e8e8e8" if THEME == "light" else "#2a2d3d")
    btn_def = Button(ax_def, "Use Defaults", color=c["ax_bg"],
                     hovercolor="#e8e8e8" if THEME == "light" else "#2a2d3d")
    for btn in (btn_ok, btn_ref, btn_def):
        btn.label.set_color(c["text"])
        btn.label.set_fontsize(8)
        for sp in btn.ax.spines.values():
            sp.set_edgecolor(c["spine"])

    state = {"done": False, "use_defaults": False}

    def _finish(use_defaults: bool):
        state["use_defaults"] = use_defaults
        state["done"] = True
        try:
            fig.canvas.stop_event_loop()
        except Exception:
            pass

    btn_ok.on_clicked(lambda _e: _finish(False))
    btn_ref.on_clicked(_apply_preview)
    btn_def.on_clicked(lambda _e: _finish(True))

    for sid in shot_ids:
        textboxes_po[sid].on_submit(lambda _txt, _sid=sid: _apply_preview())
        textboxes_shift[sid].on_submit(lambda _txt, _sid=sid: _apply_preview())
        try:
            textboxes_po[sid].on_text_change(lambda _txt, _sid=sid: _apply_preview())
            textboxes_shift[sid].on_text_change(lambda _txt, _sid=sid: _apply_preview())
        except Exception:
            pass

    def _on_key(evt: Any):
        if evt.key == "enter":
            _finish(False)
        elif evt.key in ("escape", "q"):
            _finish(True)

    def _on_close(_evt: Any):
        _finish(True)

    fig.canvas.mpl_connect("key_press_event", _on_key)
    fig.canvas.mpl_connect("close_event", _on_close)
    plt.show(block=False)
    while not state["done"] and plt.fignum_exists(fig.number):
        try:
            fig.canvas.start_event_loop(0.05)
        except Exception:
            break

    if state["use_defaults"]:
        try:
            if plt.fignum_exists(fig.number):
                plt.close(fig)
        except Exception:
            pass
        return out_po, out_shift

    for sid in shot_ids:
        cur_po = out_po[sid]
        cur_shift = out_shift.get(sid, 0.0)
        out_po[sid] = _read_float(textboxes_po[sid], cur_po)
        out_shift[sid] = _read_float(textboxes_shift[sid], cur_shift)
        if not np.isfinite(out_po[sid]):
            status_txt.set_text(f"Invalid PO for Shot {sid}; keeping {cur_po:.3f}")
            out_po[sid] = cur_po
        if not np.isfinite(out_shift[sid]):
            status_txt.set_text(f"Invalid X-shift for Shot {sid}; keeping {cur_shift:.3f}")
            out_shift[sid] = cur_shift

    try:
        if plt.fignum_exists(fig.number):
            plt.close(fig)
    except Exception:
        pass
    return out_po, out_shift


def build_corrected_pick_data(shots_info: list, all_picks: dict,
                              recv_positions: Any,
                              perp_by_shot: dict | None = None,
                              inline_shift_by_shot: dict | None = None) -> dict:
    """
    Build per-shot corrected pick rows with true offset and interpolated travel-time.

    Returns
    -------
    {shot_id: [
        {
          trace_idx, trace_no, recv_pos_m, shot_pos_m,
          inline_signed_m, inline_abs_m, perp_m, true_off_m,
          fb_raw_ms, fb_bulk_ms, fb_interp_inline_ms, side
        }, ...
    ]}
    """
    corrected: dict = {}
    for shot_id, shot_pos_m in shots_info:
        raw_picks = all_picks.get(shot_id, {})
        if not raw_picks:
            corrected[int(shot_id)] = []
            continue

        bulk_picks = apply_bulk_static(raw_picks)
        po = float((perp_by_shot or {}).get(int(shot_id), 0.0))
        x_shift = float((inline_shift_by_shot or {}).get(int(shot_id), 0.0))
        rows: list = []

        for trace_idx in sorted(raw_picks):
            if trace_idx >= len(recv_positions):
                continue
            recv_pos = float(recv_positions[trace_idx])
            inline_signed = recv_pos - float(shot_pos_m)
            inline_abs = abs(inline_signed)
            inline_corr = max(0.0, float(inline_abs + x_shift))
            true_off = float(true_offset(inline_corr, po))
            side = "L" if inline_signed < 0.0 else "R"

            rows.append({
                "trace_idx": int(trace_idx),
                "trace_no": int(trace_idx) + 1,
                "recv_pos_m": recv_pos,
                "shot_pos_m": float(shot_pos_m),
                "inline_signed_m": float(inline_signed),
                "inline_abs_m": float(inline_abs),
                "inline_shift_m": float(x_shift),
                "inline_corr_m": float(inline_corr),
                "perp_m": po,
                "true_off_m": true_off,
                "fb_raw_ms": float(raw_picks[trace_idx]),
                "fb_bulk_ms": float(bulk_picks.get(trace_idx, raw_picks[trace_idx])),
                "fb_interp_inline_ms": float(bulk_picks.get(trace_idx, raw_picks[trace_idx])),
                "fb_interp_geom_ms": float(bulk_picks.get(trace_idx, raw_picks[trace_idx])),
                "side": side,
            })

        for side in ("L", "R"):
            part = [r for r in rows if r["side"] == side]
            if len(part) < 2:
                continue

            part_sorted = sorted(part, key=lambda r: (r["inline_abs_m"], r["trace_idx"]))
            x_true = np.array([r["true_off_m"] for r in part_sorted], dtype=float)
            t_bulk = np.array([r["fb_bulk_ms"] for r in part_sorted], dtype=float)
            x_inline = np.array([r["inline_corr_m"] for r in part_sorted], dtype=float)

            if np.allclose(x_true, x_true[0]):
                t_interp = t_bulk.copy()
            else:
                order = np.argsort(x_true)
                xs = x_true[order]
                ts = t_bulk[order]
                t_interp = np.interp(x_inline, xs, ts)

            for rr, ti in zip(part_sorted, t_interp):
                rr["fb_interp_inline_ms"] = float(ti)
                rr["fb_interp_geom_ms"] = float(ti)

        corrected[int(shot_id)] = sorted(rows, key=lambda r: r["trace_idx"])

    return corrected


class AnalysisWorkflow:
    """
    Dedicated post-picking analysis workflow.

    Responsibilities
    ----------------
    - Collect/edit per-shot PO values in a UI.
    - Compute corrected offsets per trace using:
        XO_k = sqrt((x_k - SP)^2 + PO^2)
      where x_k comes from geometry, SP is shot position, and PO is
      perpendicular shot-to-line offset.
    - Show layer-window picking plots and return fitted layer metrics.
    """

    def __init__(self, profile_name: str, cfg: dict,
                 shots_info: list, shot_label_pos: dict,
                 all_picks: dict, recv_positions: Any,
                 perp_override: dict | None = None,
                 inline_shift_override: dict | None = None,
                 enable_layer_pick: bool = True,
                 existing_layer_results: dict | None = None):
        self.profile_name = profile_name
        self.cfg = cfg
        self.shots_info = shots_info
        self.shot_label_pos = shot_label_pos
        self.all_picks = all_picks
        self.recv_positions = recv_positions
        self.perp_override = perp_override or {}
        self.inline_shift_override = inline_shift_override or {}
        self.enable_layer_pick = bool(enable_layer_pick)
        self.existing_layer_results = existing_layer_results or {}

        self.perp_by_shot: dict = {}
        self.inline_shift_by_shot: dict = {}
        self.corrected_by_shot: dict = {}
        self.layer_results: dict = {}

    def run(self) -> dict:
        shot_ids = [sid for sid, _ in self.shots_info]
        self.perp_by_shot = resolve_perp_by_shot(self.cfg, shot_ids, self.perp_override)
        self.inline_shift_by_shot = {int(sid): 0.0 for sid in shot_ids}
        for sid, val in (self.inline_shift_override or {}).items():
            self.inline_shift_by_shot[int(sid)] = float(val)

        print("  Step 1/2: Perpendicular offset (PO) setup")
        self.perp_by_shot, self.inline_shift_by_shot = prompt_offset_model_by_shot(
            self.perp_by_shot,
            inline_shift_by_shot=self.inline_shift_by_shot,
            shots_info=self.shots_info,
            all_picks=self.all_picks,
            recv_positions=self.recv_positions,
        )
        print("  Perp offsets by shot: "
              + ", ".join(f"S{sid}={self.perp_by_shot.get(sid, 0.0):.2f}m" for sid in shot_ids))
        print("  Inline shift by shot: "
              + ", ".join(f"S{sid}={self.inline_shift_by_shot.get(sid, 0.0):.2f}m" for sid in shot_ids))

        print("  Step 2/2: Corrected picks and layer windows")
        self.corrected_by_shot = build_corrected_pick_data(
            self.shots_info, self.all_picks, self.recv_positions,
            perp_by_shot=self.perp_by_shot,
            inline_shift_by_shot=self.inline_shift_by_shot,
        )

        if self.enable_layer_pick:
            self.layer_results = pick_layer_windows_interactive(
                self.profile_name, self.corrected_by_shot,
                existing_results=self.existing_layer_results,
            )
        else:
            self.layer_results = self.existing_layer_results

        return {
            "perp_by_shot": self.perp_by_shot,
            "inline_shift_by_shot": self.inline_shift_by_shot,
            "corrected_by_shot": self.corrected_by_shot,
            "layer_results": self.layer_results,
        }


# ---------------------------------------------------------------------------
# SEG2 reader
# ---------------------------------------------------------------------------

def read_seg2(path: Path) -> tuple:
    """
    Read one SEG2 shot record.

    Returns
    -------
    data          : ndarray (n_traces x n_samples, float32)
    dt_s          : sample interval (s)
    n_traces      : int
    n_samples     : int
    shot_pos      : float or None  (SOURCE_LOCATION header, first component)
    ffid          : int  (SHOT_SEQUENCE_NUMBER or filename digits)
    delay_ms      : float  (DELAY header in ms; typically negative pre-trigger)
    recv_locs_m   : ndarray or None  (RECEIVER_LOCATION per trace, m)
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        st = _read_obspy(str(path), format="SEG2")

    n_traces = len(st)
    n_samp   = max(tr.stats.npts for tr in st)
    dt_s     = float(st[0].stats.delta)

    data = np.zeros((n_traces, n_samp), dtype=np.float32)
    for i, tr in enumerate(st):
        npts = tr.stats.npts
        data[i, :npts] = tr.data.astype(np.float32)

    shot_pos: Any  = None
    ffid: int      = 0
    delay_ms: float = 0.0
    recv_locs_m: Any = None

    try:
        hdr0 = dict(st[0].stats.seg2)

        sloc = str(hdr0.get("SOURCE_LOCATION", "")).strip()
        if sloc:
            # Guard against German comma decimal
            shot_pos = float(sloc.replace(",", ".").split()[0])

        ssn = str(hdr0.get("SHOT_SEQUENCE_NUMBER", "")).strip()
        if ssn.isdigit():
            ffid = int(ssn)

        # DELAY is stored in seconds; convert to ms
        delay_str = str(hdr0.get("DELAY", "0")).strip()
        delay_ms  = float(delay_str.replace(",", ".")) * 1000.0

        # RECEIVER_LOCATION per trace (m along profile)
        locs = []
        for tr in st:
            h = dict(tr.stats.seg2)
            rl = str(h.get("RECEIVER_LOCATION", "")).strip().replace(",", ".")
            locs.append(float(rl) if rl else None)
        if all(v is not None for v in locs):
            recv_locs_m = np.array(locs, dtype=float)
    except Exception:
        pass

    if ffid == 0:
        digits = "".join(c for c in path.stem if c.isdigit())
        ffid   = int(digits) if digits else 0

    return data, dt_s, n_traces, n_samp, shot_pos, ffid, delay_ms, recv_locs_m


# ---------------------------------------------------------------------------
# Ormsby bandpass  (zero-phase, frequency domain)
# ---------------------------------------------------------------------------

def _ormsby_response(freqs: Any,
                     f1: float, f2: float, f3: float, f4: float) -> Any:
    """Trapezoidal Ormsby amplitude response (linear ramps)."""
    H = np.zeros(len(freqs), dtype=float)
    m_lo = (freqs > f1) & (freqs <= f2)
    m_pb = (freqs > f2) & (freqs <= f3)
    m_hi = (freqs > f3) & (freqs < f4)
    H[m_lo] = (freqs[m_lo] - f1) / (f2 - f1)
    H[m_pb] = 1.0
    H[m_hi] = (f4 - freqs[m_hi]) / (f4 - f3)
    return H


def ormsby(trace: Any, dt_s: float,
           f1: float = BP_F1, f2: float = BP_F2,
           f3: float = BP_F3, f4: float = BP_F4,
           pad_pct: float = BP_FFT_PAD,
           reapply: bool = BP_REAPPLY) -> Any:
    """
    Zero-phase Ormsby bandpass filter  (ProMAX-compatible).

    Algorithm
    ---------
    1. Zero-pad to next power-of-2  (pad_pct=0.25 -> 25 % extra = ProMAX default)
    2. FFT -> multiply by trapezoidal amplitude response (zero-phase: real filter)
    3. IFFT -> trim to original length
    4. Optionally repeat once  (reapply=True squares the amplitude response)
    """
    n     = len(trace)
    n_pad = int(n * (1.0 + pad_pct))
    n_fft = 1 << (n_pad - 1).bit_length()          # next power of 2
    spec  = np.fft.rfft(trace.astype(np.float64), n=n_fft)
    freqs = np.fft.rfftfreq(n_fft, d=dt_s)
    H     = _ormsby_response(freqs, f1, f2, f3, f4)
    if reapply:
        H = H ** 2
    filtered = np.fft.irfft(spec * H, n=n_fft)
    return filtered[:n].astype(np.float32)


def apply_ormsby_all(data: Any, dt_s: float) -> Any:
    """Apply Ormsby bandpass to every trace in data (n_traces x n_samples)."""
    out = np.empty_like(data)
    for i in range(data.shape[0]):
        out[i] = ormsby(data[i], dt_s)
    return out


def apply_ormsby_all_params(data: Any, dt_s: float,
                            f1: float, f2: float, f3: float, f4: float) -> Any:
    """Apply Ormsby with explicit frequency parameters to all traces."""
    out = np.empty_like(data)
    for i in range(data.shape[0]):
        out[i] = ormsby(data[i], dt_s, f1=f1, f2=f2, f3=f3, f4=f4)
    return out


def butterworth_bandpass(trace: Any, dt_s: float,
                         low_hz: float, high_hz: float,
                         order: int = BUTTER_ORDER) -> Any:
    """Zero-phase Butterworth bandpass using scipy SOS filter."""
    x = np.asarray(trace, dtype=np.float64)
    if x.size < 8:
        return x.astype(np.float32)

    fs = 1.0 / max(float(dt_s), 1e-12)
    nyq = 0.5 * fs
    lo = max(0.001, float(low_hz))
    hi = min(float(high_hz), nyq * 0.999)
    if not (0.0 < lo < hi < nyq):
        return x.astype(np.float32)

    sos = _scipy_butter(int(max(1, order)), [lo, hi], btype="bandpass", fs=fs, output="sos")
    try:
        y = _scipy_sosfiltfilt(sos, x)
    except Exception:
        y = x
    return np.asarray(y, dtype=np.float32)


def apply_butterworth_all_params(data: Any, dt_s: float,
                                 low_hz: float, high_hz: float,
                                 order: int = BUTTER_ORDER) -> Any:
    """Apply Butterworth bandpass to every trace in data."""
    out = np.empty_like(data)
    for i in range(data.shape[0]):
        out[i] = butterworth_bandpass(data[i], dt_s, low_hz=low_hz, high_hz=high_hz, order=order)
    return out


# ---------------------------------------------------------------------------
# Bulk static correction
# ---------------------------------------------------------------------------

def apply_bulk_static(picks_raw: dict) -> dict:
    """
    Apply BULK_SHIFT_MS to a picks dict {trace_idx: time_ms}.
    Call at analysis / export time; JSON always stores RAW picks.
    """
    if abs(BULK_SHIFT_MS) < 1e-9:
        return picks_raw
    return {k: max(0.0, float(v) + BULK_SHIFT_MS) for k, v in picks_raw.items()}


def apply_gain(data: Any, dt_s: float,
               mode: str = GAIN_MODE,
               window_ms: float = AGC_WINDOW_MS,
               stat: str = AGC_STAT) -> Any:
    mode_l = str(mode).lower()
    out = np.asarray(data, dtype=np.float32).copy()

    if mode_l == "none":
        return out

    if mode_l == "norm":
        for i in range(out.shape[0]):
            mx = float(np.max(np.abs(out[i])))
            if mx > 1e-20:
                out[i] /= mx
        return out

    if mode_l == "agc":
        win = max(3, int((window_ms / 1000.0) / dt_s))
        if win % 2 == 0:
            win += 1
        ker = np.ones(win, dtype=np.float64) / float(win)
        for i in range(out.shape[0]):
            tr = out[i].astype(np.float64)
            if str(stat).lower() == "mean":
                env = np.convolve(np.abs(tr), ker, mode="same")
            else:
                env = np.sqrt(np.convolve(tr * tr, ker, mode="same"))
            env = np.where(env > 1e-12, env, 1e-12)
            out[i] = (tr / env).astype(np.float32)
        return out

    return out


# ---------------------------------------------------------------------------
# Refraction analysis functions  (integrated from lvl.ipynb)
# ---------------------------------------------------------------------------

def fit_line(x: Any, y: Any) -> tuple:
    """
    Fit y = slope*x + intercept.
    Returns (slope [ms/m], intercept [ms], r^2).
    Velocity (m/s) = 1000 / slope  when x in m, y in ms.
    """
    if len(x) < 2:
        return float("nan"), float("nan"), float("nan")
    x_a = np.asarray(x, dtype=float)
    y_a = np.asarray(y, dtype=float)
    slope, intercept = np.polyfit(x_a, y_a, 1)
    y_hat  = slope * x_a + intercept
    ss_res = float(np.sum((y_a - y_hat) ** 2))
    ss_tot = float(np.sum((y_a - y_a.mean()) ** 2))
    r2     = 1.0 - ss_res / ss_tot if ss_tot > 1e-30 else 1.0
    return float(slope), float(intercept), float(r2)


def depth_2layer(ti_ms: float, V1: float, V2: float) -> Any:
    """
    Intercept-time depth, 2-layer model.
    h = (ti_s x V1 x V2) / (2 x sqrt(V2^2 - V1^2))  [m]
    Returns None if V2 <= V1.
    """
    if V2 <= V1 or V1 <= 0.0:
        return None
    return (ti_ms / 1000.0) * V1 * V2 / (2.0 * math.sqrt(V2 ** 2 - V1 ** 2))


def depth_3layer(ti2_ms: float, V1: float, V2: float,
                 V3: float, h1: float) -> Any:
    """
    Intercept-time depth to second refractor, 3-layer model.
    Accounts for first-layer delay time.  Returns None if geometry is invalid.
    """
    if V3 <= V2 or V2 <= V1 or V1 <= 0.0:
        return None
    cos_ic12 = math.sqrt(max(0.0, 1.0 - (V1 / V2) ** 2))
    delay_ms = 2.0 * h1 * cos_ic12 / V1 * 1000.0
    ti2_eff  = ti2_ms - delay_ms
    if ti2_eff <= 0.0:
        return None
    return (ti2_eff / 1000.0) * V1 * V3 / (2.0 * math.sqrt(V3 ** 2 - V1 ** 2))


def _pick_layer_windows_from_plot(x_vals: Any, t_vals: Any,
                                  title: str) -> dict:
    """
        Let user pick up to 6 x-boundaries on a T-X plot:
            x1-x2 -> layer 1
            x3-x4 -> layer 2
            x5-x6 -> layer 3
    Dedicated UI controls:
      left click = add point, Undo = remove last, Save = continue, Skip = ignore side
    """
    try:
        backend = str(plt.get_backend()).lower()
    except Exception:
        backend = ""
    if "agg" in backend and not _ensure_interactive_backend():
        return {"windows": [None, None, None], "picked_points": []}

    c = _tc()
    fig, ax = plt.subplots(figsize=(8.6, 5.4), constrained_layout=False)
    fig.subplots_adjust(left=0.08, right=0.74, bottom=0.12, top=0.90)
    fig.patch.set_facecolor(c["fig_bg"])
    ax.set_facecolor(c["ax_bg"])
    ax.plot(x_vals, t_vals, "o-", ms=4, lw=1.2, color="#1f77b4")
    ax.set_xlabel("Corrected true offset XO (m)", color=c["label"])
    ax.set_ylabel("Corrected travel-time (ms)", color=c["label"])
    ax.set_title(title + "\nPick x1..x6 with mouse (LMB). Use buttons on right.",
                 color=c["text"], fontsize=10)
    ax.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax.tick_params(colors=c["tick"])
    for sp in ax.spines.values():
        sp.set_edgecolor(c["spine"])

    pick_names = ["x1", "x2", "x3", "x4", "x5", "x6"]
    state = {"xs": [], "ts": [], "done": False, "skip": False}
    click_lines: list = []
    click_labels: list = []

    info_ax = fig.add_axes([0.76, 0.62, 0.22, 0.28])
    info_ax.set_facecolor(c["ax_bg"])
    info_ax.set_xticks([])
    info_ax.set_yticks([])
    for sp in info_ax.spines.values():
        sp.set_edgecolor(c["spine"])
    info_txt = info_ax.text(
        0.05, 0.95,
        "Layer windows\n"
        "x1-x2: layer1\n"
        "x3-x4: layer2\n"
        "x5-x6: layer3\n\n"
        "Selected: none",
        transform=info_ax.transAxes,
        color=c["text"], fontsize=8, va="top", ha="left",
    )

    ax_undo = fig.add_axes([0.76, 0.50, 0.22, 0.07])
    ax_save = fig.add_axes([0.76, 0.41, 0.22, 0.07])
    ax_skip = fig.add_axes([0.76, 0.32, 0.22, 0.07])

    hover = "#e8e8e8" if THEME == "light" else "#2a2d3d"
    btn_undo = Button(ax_undo, "Undo last", color=c["ax_bg"], hovercolor=hover)
    btn_save = Button(ax_save, "Save side", color=c["ax_bg"], hovercolor=hover)
    btn_skip = Button(ax_skip, "Skip side", color=c["ax_bg"], hovercolor=hover)
    for btn in (btn_undo, btn_save, btn_skip):
        btn.label.set_color(c["text"])
        btn.label.set_fontsize(8)
        for sp in btn.ax.spines.values():
            sp.set_edgecolor(c["spine"])

    def _update_info():
        if not state["xs"]:
            sel = "none"
        else:
            parts = []
            for i, xv in enumerate(state["xs"]):
                nm = pick_names[i] if i < len(pick_names) else f"x{i}"
                tv = state["ts"][i] if i < len(state["ts"]) else float("nan")
                parts.append(f"{nm}={xv:.2f}, t={tv:.2f}")
            sel = ", ".join(parts)
        info_txt.set_text(
            "Layer windows\n"
            "x1-x2: layer1\n"
            "x3-x4: layer2\n"
            "x5-x6: layer3\n\n"
            f"Selected: {sel}"
        )

    def _finish(skip: bool):
        if state["done"]:
            return
        state["skip"] = skip
        state["done"] = True
        try:
            fig.canvas.stop_event_loop()
        except Exception:
            pass

    def _undo(_evt: Any):
        if not state["xs"]:
            return
        state["xs"].pop()
        if state["ts"]:
            state["ts"].pop()
        if click_lines:
            ln = click_lines.pop()
            try:
                ln.remove()
            except Exception:
                pass
        if click_labels:
            tx = click_labels.pop()
            try:
                tx.remove()
            except Exception:
                pass
        _update_info()
        fig.canvas.draw_idle()

    def _save(_evt: Any):
        _finish(False)

    def _skip(_evt: Any):
        _finish(True)

    btn_undo.on_clicked(_undo)
    btn_save.on_clicked(_save)
    btn_skip.on_clicked(_skip)

    def _on_click(evt: Any):
        if evt.inaxes is not ax or evt.xdata is None or evt.ydata is None:
            return
        if len(state["xs"]) >= 6:
            return
        xv = float(evt.xdata)
        tv = float(evt.ydata)
        idx = len(state["xs"])
        state["xs"].append(xv)
        state["ts"].append(tv)
        nm = pick_names[idx]
        ln = ax.axvline(x=xv, color="#e63946", lw=1.0, ls="--", alpha=0.85)
        tx = ax.text(xv, tv, f"{nm}\n{tv:.1f}ms", color="#e63946", fontsize=8,
                 rotation=90, va="bottom", ha="center")
        click_lines.append(ln)
        click_labels.append(tx)
        _update_info()
        fig.canvas.draw_idle()

    def _on_key(evt: Any):
        if evt.key == "enter":
            _finish(False)
        elif evt.key in ("escape", "q"):
            _finish(True)
        elif evt.key in ("backspace", "delete"):
            _undo(evt)

    def _on_close(_evt: Any):
        _finish(True)

    fig.canvas.mpl_connect("button_press_event", _on_click)
    fig.canvas.mpl_connect("key_press_event", _on_key)
    fig.canvas.mpl_connect("close_event", _on_close)
    _update_info()
    plt.show(block=False)
    while not state["done"] and plt.fignum_exists(fig.number):
        try:
            fig.canvas.start_event_loop(0.05)
        except Exception:
            break

    xs = state["xs"]
    try:
        if plt.fignum_exists(fig.number):
            plt.close(fig)
    except Exception:
        pass

    if state["skip"] or not xs:
        return {"windows": [None, None, None], "picked_points": []}

    windows: list = []
    for i in range(0, 6, 2):
        if i + 1 < len(xs):
            a, b = xs[i], xs[i + 1]
            windows.append((min(a, b), max(a, b)))
        else:
            windows.append(None)
    picked_points = [(float(xv), float(tv)) for xv, tv in zip(state.get("xs", []), state.get("ts", []))]
    return {"windows": windows[:3], "picked_points": picked_points}


def _fit_layers_from_windows(x_vals: Any, t_vals: Any, windows: list) -> dict:
    """Fit up to 3 linear layers from user-selected x-windows."""
    x = np.asarray(x_vals, dtype=float)
    t = np.asarray(t_vals, dtype=float)

    # Handle cases where the shallowest visible layer is not picked (e.g. direct to L2):
    # shift non-empty windows to the left so fitting always starts at first available layer.
    wins_in = list((windows or []))[:3]
    wins_in += [None] * (3 - len(wins_in))
    normalized_windows = [w for w in wins_in if w is not None]
    normalized_windows += [None] * (3 - len(normalized_windows))

    segs: list = []
    for layer_idx in range(3):
        win = normalized_windows[layer_idx] if layer_idx < len(normalized_windows) else None
        if not win:
            segs.append({
                "layer": layer_idx + 1,
                "x0": 0.0,
                "x1": 0.0,
                "velocity_m_s": 0.0,
                "intercept_ms": 0.0,
                "slope_ms_m": 0.0,
                "r2": 0.0,
                "n": 0,
            })
            continue

        x0, x1 = float(win[0]), float(win[1])
        mask = (x >= x0) & (x <= x1)
        if int(mask.sum()) < 2:
            segs.append({
                "layer": layer_idx + 1,
                "x0": x0,
                "x1": x1,
                "velocity_m_s": 0.0,
                "intercept_ms": 0.0,
                "slope_ms_m": 0.0,
                "r2": 0.0,
                "n": int(mask.sum()),
            })
            continue

        slope, intercept, r2 = fit_line(x[mask], t[mask])
        vel = (1000.0 / slope) if slope > 1e-9 else 0.0
        segs.append({
            "layer": layer_idx + 1,
            "x0": x0,
            "x1": x1,
            "velocity_m_s": float(max(0.0, vel)),
            "intercept_ms": float(intercept),
            "slope_ms_m": float(slope),
            "r2": float(r2),
            "n": int(mask.sum()),
        })

    v0 = float(segs[0]["velocity_m_s"])
    v1 = float(segs[1]["velocity_m_s"])
    v2 = float(segs[2]["velocity_m_s"])
    ti1 = float(segs[1]["intercept_ms"]) if v1 > 0.0 else 0.0
    ti2 = float(segs[2]["intercept_ms"]) if v2 > 0.0 else 0.0

    h1 = depth_2layer(ti1, v0, v1) if (v0 > 0.0 and v1 > 0.0 and ti1 > 0.0) else None
    h2 = depth_3layer(ti2, v0, v1, v2, h1) if (h1 and v2 > 0.0 and ti2 > 0.0) else None

    return {
        "segments": segs,
        "V0_m_s": v0,
        "V1_m_s": v1,
        "V2_m_s": v2,
        "ti1_ms": ti1,
        "ti2_ms": ti2,
        "h1_m": float(h1) if h1 else 0.0,
        "h2_m": float(h2) if h2 else 0.0,
    }


def _compute_fit_rms(x_vals: Any, t_vals: Any, fit_res: dict) -> float:
    """Compute RMS(ms) between observed and fitted times for given x-values."""
    x = np.asarray(x_vals, dtype=float)
    t = np.asarray(t_vals, dtype=float)
    if x.size == 0 or t.size == 0:
        return 0.0
    pred: list = []
    obs: list = []
    for xv, tv in zip(x, t):
        tp = _predict_time_from_fit(float(xv), fit_res)
        if tp is None:
            continue
        pred.append(float(tp))
        obs.append(float(tv))
    if not pred:
        return 0.0
    pa = np.asarray(pred, dtype=float)
    oa = np.asarray(obs, dtype=float)
    return float(np.sqrt(np.mean((oa - pa) ** 2)))


def _review_layer_fit_interactive(x_vals: Any, t_vals: Any,
                                  fit_res: dict, title: str) -> str:
    """
    Review layer fit and choose action.

    Returns one of: "accept", "repick", "skip".
    """
    try:
        backend = str(plt.get_backend()).lower()
    except Exception:
        backend = ""
    if "agg" in backend and not _ensure_interactive_backend():
        return "accept"

    c = _tc()
    x = np.asarray(x_vals, dtype=float)
    t = np.asarray(t_vals, dtype=float)
    if x.size < 2 or t.size < 2:
        return "accept"

    rms = _compute_fit_rms(x, t, fit_res)
    nfit = int(sum(_predict_time_from_fit(float(xx), fit_res) is not None for xx in x))

    fig, (ax_tx, ax_cmp) = plt.subplots(1, 2, figsize=(12.2, 5.0), constrained_layout=False)
    fig.subplots_adjust(left=0.06, right=0.82, bottom=0.12, top=0.88, wspace=0.22)
    fig.patch.set_facecolor(c["fig_bg"])
    ax_tx.set_facecolor(c["ax_bg"])
    ax_cmp.set_facecolor(c["ax_bg"])

    ax_tx.plot(x, t, "o", ms=4, color="#1f77b4", alpha=0.9, label="Observed")
    wins = list((fit_res or {}).get("windows", []) or [])
    segs = list((fit_res or {}).get("segments", []) or [])
    for i, seg in enumerate(segs):
        sl = float(seg.get("slope_ms_m", 0.0))
        ic = float(seg.get("intercept_ms", 0.0))
        if sl <= 0.0:
            continue
        w = wins[i] if i < len(wins) else None
        if w is not None:
            x0, x1 = float(w[0]), float(w[1])
        else:
            x0, x1 = float(seg.get("x0", 0.0)), float(seg.get("x1", 0.0))
        if x1 <= x0:
            continue
        xl = np.linspace(x0, x1, 40)
        tl = sl * xl + ic
        ax_tx.plot(xl, tl, "-", lw=1.6, alpha=0.9, color="#e63946")
        xm = 0.5 * (x0 + x1)
        tm = float(sl * xm + ic)
        ax_tx.text(xm, tm - 2.0, f"V={seg.get('velocity_m_s', 0.0):.0f} m/s",
                   fontsize=7, color="#e63946", ha="center")

    pred_all: list = []
    obs_all: list = []
    for xv, tv in zip(x, t):
        tp = _predict_time_from_fit(float(xv), fit_res)
        if tp is None:
            continue
        obs_all.append(float(tv))
        pred_all.append(float(tp))

    if obs_all:
        oa = np.asarray(obs_all, dtype=float)
        pa = np.asarray(pred_all, dtype=float)
        ax_cmp.scatter(oa, pa, s=24, color="#2a9d8f", alpha=0.9)
        lo = float(min(oa.min(), pa.min()))
        hi = float(max(oa.max(), pa.max()))
        pad = max(1.0, 0.03 * (hi - lo))
        lo -= pad
        hi += pad
        ax_cmp.plot([lo, hi], [lo, hi], "--", lw=1.1, color="#555555")
        ax_cmp.set_xlim(lo, hi)
        ax_cmp.set_ylim(lo, hi)
    else:
        ax_cmp.text(0.5, 0.5, "No predicted points", transform=ax_cmp.transAxes,
                    ha="center", va="center", color=c["label"], fontsize=8)

    ax_tx.set_xlabel("Corrected true offset XO (m)", color=c["label"], fontsize=8)
    ax_tx.set_ylabel("Corrected travel-time (ms)", color=c["label"], fontsize=8)
    ax_tx.set_title("Observed with fitted segments", color=c["text"], fontsize=9)
    ax_tx.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax_tx.tick_params(colors=c["tick"], labelsize=8)

    ax_cmp.set_xlabel("Observed (ms)", color=c["label"], fontsize=8)
    ax_cmp.set_ylabel("Computed (ms)", color=c["label"], fontsize=8)
    ax_cmp.set_title("Observed vs Computed", color=c["text"], fontsize=9)
    ax_cmp.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax_cmp.tick_params(colors=c["tick"], labelsize=8)

    for ax in (ax_tx, ax_cmp):
        for sp in ax.spines.values():
            sp.set_edgecolor(c["spine"])

    fig.suptitle(
        f"{title} | RMS={rms:.3f} ms (Nfit={nfit})",
        color=c["text"], fontsize=10,
    )

    panel = fig.add_axes([0.84, 0.28, 0.14, 0.40])
    panel.set_facecolor(c["ax_bg"])
    panel.set_xticks([])
    panel.set_yticks([])
    for sp in panel.spines.values():
        sp.set_edgecolor(c["spine"])
    panel.text(0.06, 0.95,
               "Fit review\n\n"
               "Accept: keep this fit\n"
               "Repick: choose windows again\n"
               "Skip: use previous/empty\n\n"
               "Keys: Enter=Accept, r=Repick, Esc=Skip",
               transform=panel.transAxes, va="top", ha="left",
               fontsize=8, color=c["text"])

    ax_acc = fig.add_axes([0.84, 0.20, 0.14, 0.06])
    ax_rep = fig.add_axes([0.84, 0.12, 0.14, 0.06])
    ax_skp = fig.add_axes([0.84, 0.04, 0.14, 0.06])
    hover = "#e8e8e8" if THEME == "light" else "#2a2d3d"
    b_acc = Button(ax_acc, "Accept", color=c["ax_bg"], hovercolor=hover)
    b_rep = Button(ax_rep, "Repick", color=c["ax_bg"], hovercolor=hover)
    b_skp = Button(ax_skp, "Skip", color=c["ax_bg"], hovercolor=hover)
    for b in (b_acc, b_rep, b_skp):
        b.label.set_color(c["text"])
        b.label.set_fontsize(8)
        for sp in b.ax.spines.values():
            sp.set_edgecolor(c["spine"])

    state = {"done": False, "choice": "accept"}

    def _finish(choice: str):
        if state["done"]:
            return
        state["done"] = True
        state["choice"] = str(choice)
        try:
            fig.canvas.stop_event_loop()
        except Exception:
            pass

    b_acc.on_clicked(lambda _e: _finish("accept"))
    b_rep.on_clicked(lambda _e: _finish("repick"))
    b_skp.on_clicked(lambda _e: _finish("skip"))

    def _on_key(evt: Any):
        if evt.key == "enter":
            _finish("accept")
        elif evt.key in ("r", "R"):
            _finish("repick")
        elif evt.key in ("escape", "q"):
            _finish("skip")

    def _on_close(_evt: Any):
        _finish("skip")

    fig.canvas.mpl_connect("key_press_event", _on_key)
    fig.canvas.mpl_connect("close_event", _on_close)
    plt.show(block=False)
    while not state["done"] and plt.fignum_exists(fig.number):
        try:
            fig.canvas.start_event_loop(0.05)
        except Exception:
            break
    try:
        if plt.fignum_exists(fig.number):
            plt.close(fig)
    except Exception:
        pass
    return str(state.get("choice", "accept"))


def pick_layer_windows_interactive(profile_name: str,
                                   corrected_by_shot: dict,
                                   existing_results: dict | None = None) -> dict:
    """
    Interactive layer picking for each shot side (L/R).
    Returns {shot_id: {"L": fit_result, "R": fit_result}}.
    """
    existing = existing_results or {}
    results: dict = {}
    print("\n  -- Layer windows picking (x1..x6) --")
    print("     For each shot side: pick x1,x2,x3,x4,x5,x6; Enter to skip side.")

    for shot_id in sorted(corrected_by_shot):
        rows = corrected_by_shot.get(shot_id, [])
        if not rows:
            continue
        per_side: dict = {}
        side_candidates: list = []
        left_rows = [r for r in rows if r.get("side") == "L"]
        right_rows = [r for r in rows if r.get("side") == "R"]
        min_side_points = 2
        if len(left_rows) >= min_side_points:
            side_candidates.append(("L", left_rows))
        if len(right_rows) >= min_side_points:
            side_candidates.append(("R", right_rows))
        if side_candidates and len(side_candidates) < 2 and (len(left_rows) > 0 and len(right_rows) > 0):
            print(f"     Shot {shot_id}: one side has too few picks for standalone fit "
                  f"(L={len(left_rows)}, R={len(right_rows)}).")

        if not side_candidates and len(rows) >= 3:
            side_candidates.append(("ALL", rows))
            print(f"     Shot {shot_id}: using combined side (ALL) for layer-window picking.")

        if not side_candidates:
            print(f"     Shot {shot_id}: skipped layer-window plot (need >=3 corrected picks).")
            continue

        for side, side_rows in side_candidates:
            side_rows = sorted(side_rows, key=lambda r: r.get("true_off_m", 0.0))
            x = np.array([r.get("true_off_m", 0.0) for r in side_rows], dtype=float)
            t = np.array([r["fb_interp_inline_ms"] for r in side_rows], dtype=float)
            title = f"Profile {profile_name} | Shot {shot_id} | Side {side}"
            print(f"     Opening layer-window picker: Shot {shot_id} Side {side}")

            while True:
                pick_payload = _pick_layer_windows_from_plot(x, t, title)
                windows = pick_payload.get("windows", [None, None, None])
                picked_points = pick_payload.get("picked_points", [])

                if all(w is None for w in windows):
                    prev = ((existing.get(int(shot_id), {}) or {}).get(side))
                    if prev:
                        per_side[side] = prev
                        print(f"     Shot {shot_id} Side {side}: skipped by user, kept previous layer picks.")
                    else:
                        fit_res = _fit_layers_from_windows(x, t, windows)
                        fit_res["windows"] = windows
                        fit_res["picked_points"] = picked_points
                        fit_res["n_points"] = int(len(x))
                        fit_res["skipped"] = True
                        fit_res["rms_ms"] = 0.0
                        per_side[side] = fit_res
                        print(f"     Shot {shot_id} Side {side}: skipped by user, saved empty layer selection.")
                    break

                fit_res = _fit_layers_from_windows(x, t, windows)
                fit_res["windows"] = windows
                fit_res["picked_points"] = picked_points
                fit_res["n_points"] = int(len(x))
                fit_res["skipped"] = False
                fit_res["rms_ms"] = _compute_fit_rms(x, t, fit_res)

                choice = _review_layer_fit_interactive(x, t, fit_res, title)
                if choice == "repick":
                    print(f"     Shot {shot_id} Side {side}: re-pick requested.")
                    continue
                if choice == "skip":
                    prev = ((existing.get(int(shot_id), {}) or {}).get(side))
                    if prev:
                        per_side[side] = prev
                        print(f"     Shot {shot_id} Side {side}: skipped in review, kept previous layer picks.")
                    else:
                        fit_res_empty = _fit_layers_from_windows(x, t, [None, None, None])
                        fit_res_empty["windows"] = [None, None, None]
                        fit_res_empty["picked_points"] = []
                        fit_res_empty["n_points"] = int(len(x))
                        fit_res_empty["skipped"] = True
                        fit_res_empty["rms_ms"] = 0.0
                        per_side[side] = fit_res_empty
                        print(f"     Shot {shot_id} Side {side}: skipped in review.")
                    break

                per_side[side] = fit_res
                print(f"     Shot {shot_id} Side {side}: accepted fit, RMS={fit_res.get('rms_ms', 0.0):.3f} ms")
                break

        if per_side:
            results[int(shot_id)] = per_side

    return results


def compute_layer_averages(layer_results: dict, shot_pos_by_id: dict) -> dict:
    """
    Compute layer averages from off-end and center shots.

    Returns
    -------
    {
      "V0": {"off":..,"center":..,"avg":..},
      "V1": {"off":..,"center":..,"avg":..},
      "V2": {"off":..,"center":..,"avg":..},
      "ti1_ms": {"off":..,"center":..,"avg":..},
      "ti2_ms": {"off":..,"center":..,"avg":..},
      "h1_m": ..., "h2_m": ...,
      "off_shots": [...], "center_shots": [...]
    }
    """
    if not layer_results:
        return {}

    shot_ids = [sid for sid in shot_pos_by_id if sid in layer_results]
    if not shot_ids:
        return {}

    shot_ids_sorted = sorted(shot_ids, key=lambda sid: float(shot_pos_by_id.get(sid, sid)))
    if len(shot_ids_sorted) >= 2:
        off_shots = [shot_ids_sorted[0], shot_ids_sorted[-1]]
    else:
        off_shots = [shot_ids_sorted[0]]

    center_candidates = [sid for sid in shot_ids_sorted if sid not in off_shots]
    if not center_candidates:
        center_candidates = [sid for sid in shot_ids_sorted if sid in off_shots]

    def _collect(shots: list, key: str) -> list:
        vals: list = []
        for sid in shots:
            for side_res in (layer_results.get(sid, {}) or {}).values():
                v = float(side_res.get(key, 0.0))
                if v > 0.0:
                    vals.append(v)
        return vals

    def _mean(vals: list) -> float:
        return float(np.mean(vals)) if vals else 0.0

    def _avg_two(off_v: float, cen_v: float) -> float:
        picks = [v for v in (off_v, cen_v) if v > 0.0]
        return float(np.mean(picks)) if picks else 0.0

    out: dict = {"off_shots": off_shots, "center_shots": center_candidates}
    for key in ("V0_m_s", "V1_m_s", "V2_m_s", "ti1_ms", "ti2_ms"):
        off_v = _mean(_collect(off_shots, key))
        cen_v = _mean(_collect(center_candidates, key))
        avg_v = _avg_two(off_v, cen_v)
        short = key.replace("_m_s", "").replace("_ms", "_ms")
        out[short] = {"off": off_v, "center": cen_v, "avg": avg_v}

    V0 = out.get("V0", {}).get("avg", 0.0)
    V1 = out.get("V1", {}).get("avg", 0.0)
    V2 = out.get("V2", {}).get("avg", 0.0)
    ti1 = out.get("ti1_ms", {}).get("avg", 0.0)
    ti2 = out.get("ti2_ms", {}).get("avg", 0.0)

    h1 = depth_2layer(ti1, V0, V1) if (V0 > 0 and V1 > 0 and ti1 > 0) else None
    h2 = depth_3layer(ti2, V0, V1, V2, h1) if (h1 and V2 > 0 and ti2 > 0) else None
    out["h1_m"] = float(h1) if h1 else 0.0
    out["h2_m"] = float(h2) if h2 else 0.0
    return out


def _predict_time_from_fit(x_abs: float, fit_res: dict) -> float | None:
    """Predict t(ms) at corrected true offset x_abs from fitted layer windows."""
    windows = list((fit_res or {}).get("windows", []) or [])
    windows += [None] * (3 - len(windows))
    segs = list((fit_res or {}).get("segments", []) or [])
    segs += [{}] * (3 - len(segs))

    x = float(x_abs)
    for i in range(3):
        w = windows[i]
        seg = segs[i]
        if w is None:
            continue
        x0, x1 = float(w[0]), float(w[1])
        if not (x0 <= x <= x1):
            continue
        sl = float(seg.get("slope_ms_m", 0.0))
        ic = float(seg.get("intercept_ms", 0.0))
        if sl <= 0.0:
            return None
        return float(sl * x + ic)

    return None


def _choose_shot_fit(layer_by_side: dict) -> tuple[str | None, dict | None]:
    """Choose one representative fit per shot for summary export."""
    if not layer_by_side:
        return None, None
    if "ALL" in layer_by_side:
        return "ALL", layer_by_side.get("ALL")

    best_side = None
    best_res = None
    best_n = -1
    for side, res in layer_by_side.items():
        n = int((res or {}).get("n_points", 0))
        if n > best_n:
            best_n = n
            best_side = side
            best_res = res
    return best_side, best_res


def build_analysis_from_layers(corrected_by_shot: dict,
                               layer_results: dict) -> dict:
    """
    Build per-shot analysis payload for Excel/T-X export.

    Output shape per shot:
      {
        "segments": [...],
        "depths": {...},
        "rms_ms": float,
        "n_fit": int,
        "fit_side": "L"|"R"|"ALL"
      }
    """
    out: dict = {}
    for shot_id, rows in (corrected_by_shot or {}).items():
        sides = (layer_results or {}).get(shot_id, {}) or {}
        fit_side, fit_res = _choose_shot_fit(sides)
        if not fit_res:
            continue

        if fit_side == "ALL":
            rows_fit = list(rows)
        else:
            rows_fit = [r for r in rows if r.get("side") == fit_side]
            if not rows_fit:
                rows_fit = list(rows)

        obs: list = []
        pred: list = []
        for r in rows_fit:
            x_abs = float(r.get("true_off_m", 0.0))
            t_obs = float(r.get("fb_interp_inline_ms", 0.0))
            t_pred = _predict_time_from_fit(x_abs, fit_res)
            if t_pred is None:
                continue
            obs.append(t_obs)
            pred.append(float(t_pred))

        if obs:
            oa = np.asarray(obs, dtype=float)
            pa = np.asarray(pred, dtype=float)
            rms = float(np.sqrt(np.mean((oa - pa) ** 2)))
        else:
            rms = 0.0

        segs_raw = list((fit_res or {}).get("segments", []) or [])
        segs: list = []
        for i, seg in enumerate(segs_raw):
            s = dict(seg)
            w = None
            wins = list((fit_res or {}).get("windows", []) or [])
            if i < len(wins):
                w = wins[i]
            if w is not None:
                s["x_start"] = float(w[0])
                s["x_end"] = float(w[1])
            else:
                s["x_start"] = float(seg.get("x0", 0.0))
                s["x_end"] = float(seg.get("x1", 0.0))
            segs.append(s)

        out[int(shot_id)] = {
            "segments": segs,
            "depths": {
                "ti1_ms": float((fit_res or {}).get("ti1_ms", 0.0)),
                "ti2_ms": float((fit_res or {}).get("ti2_ms", 0.0)),
                "h1_m": float((fit_res or {}).get("h1_m", 0.0)),
                "h2_m": float((fit_res or {}).get("h2_m", 0.0)),
            },
            "rms_ms": rms,
            "n_fit": int(len(obs)),
            "fit_side": fit_side or "",
        }

    return out


def export_fit_plot(profile_name: str,
                    corrected_by_shot: dict,
                    layer_results: dict,
                    filename_suffix: str = "") -> Path:
    """Save observed-vs-computed fit scatter with global RMS annotation."""
    c = _tc()
    fig, ax = plt.subplots(figsize=(8.2, 6.2))
    fig.patch.set_facecolor(c["fig_bg"])
    ax.set_facecolor(c["ax_bg"])

    pal = ["#e63946", "#2a9d8f", "#e9c46a", "#457b9d", "#f4a261", "#8d99ae"]
    all_obs: list = []
    all_pred: list = []

    for i, shot_id in enumerate(sorted(corrected_by_shot)):
        rows = corrected_by_shot.get(shot_id, [])
        side_map = (layer_results or {}).get(shot_id, {}) or {}
        fit_side, fit_res = _choose_shot_fit(side_map)
        if not rows or not fit_res:
            continue

        if fit_side == "ALL":
            rows_fit = rows
        else:
            rows_fit = [r for r in rows if r.get("side") == fit_side]
            if not rows_fit:
                rows_fit = rows

        obs: list = []
        pred: list = []
        for r in rows_fit:
            t_obs = float(r.get("fb_interp_inline_ms", 0.0))
            t_pred = _predict_time_from_fit(float(r.get("true_off_m", 0.0)), fit_res)
            if t_pred is None:
                continue
            obs.append(t_obs)
            pred.append(float(t_pred))

        if not obs:
            continue

        oa = np.asarray(obs, dtype=float)
        pa = np.asarray(pred, dtype=float)
        shot_rms = float(np.sqrt(np.mean((oa - pa) ** 2)))
        col = pal[i % len(pal)]
        ax.scatter(oa, pa, s=26, color=col, alpha=0.86,
                   label=f"Shot {shot_id} ({fit_side}) RMS={shot_rms:.2f} ms")
        all_obs.extend(obs)
        all_pred.extend(pred)

    if all_obs:
        obs_a = np.asarray(all_obs, dtype=float)
        pred_a = np.asarray(all_pred, dtype=float)
        rms_all = float(np.sqrt(np.mean((obs_a - pred_a) ** 2)))
        lo = float(min(obs_a.min(), pred_a.min()))
        hi = float(max(obs_a.max(), pred_a.max()))
        pad = max(1.0, 0.03 * (hi - lo))
        lo -= pad
        hi += pad
        ax.plot([lo, hi], [lo, hi], "--", color="#444444", lw=1.2, label="Ideal: y=x")
        ax.set_xlim(lo, hi)
        ax.set_ylim(lo, hi)
        ax.text(0.02, 0.98, f"Global RMS = {rms_all:.3f} ms\nN = {len(obs_a)}",
                transform=ax.transAxes, va="top", ha="left",
                fontsize=9, color=c["text"])
    else:
        ax.text(0.5, 0.5, "No fitted points available", transform=ax.transAxes,
                ha="center", va="center", color=c["text"])

    ax.set_xlabel("Observed FB time (ms)", color=c["label"])
    ax.set_ylabel("Computed FB time (ms)", color=c["label"])
    ax.set_title(f"Observed vs Computed Fit - Profile {profile_name}",
                 color=c["text"], fontsize=11)
    ax.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax.tick_params(colors=c["tick"])
    for sp in ax.spines.values():
        sp.set_edgecolor(c["spine"])
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(fontsize=8, facecolor=c["leg_face"], edgecolor=c["leg_edge"],
                  labelcolor=c["text"], loc="best")
    fig.tight_layout()

    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{profile_name}_fit_rms{filename_suffix}.png"
    fig.savefig(str(out), dpi=180, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  Fit plot  -> {out.relative_to(CWD.parent)}")
    return out


def export_corrected_qc_plot(profile_name: str,
                             corrected_by_shot: dict,
                             shot_label_pos: dict | None = None,
                             layer_results: dict | None = None,
                             filename_suffix: str = "",
                             show_plot: bool = False) -> Path:
    """QC visualization for corrected picks and true-offset mapping."""
    c = _tc()
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(17, 5), sharey=False)
    fig.patch.set_facecolor(c["fig_bg"])
    ax1.set_facecolor(c["ax_bg"])
    ax2.set_facecolor(c["ax_bg"])
    ax3.set_facecolor(c["ax_bg"])
    ax3_top = ax3.twiny()
    ax3_top.set_facecolor("none")

    pal = ["#e63946", "#2a9d8f", "#e9c46a", "#457b9d", "#f4a261", "#8d99ae"]

    for i, shot_id in enumerate(sorted(corrected_by_shot)):
        rows = corrected_by_shot.get(shot_id, [])
        if not rows:
            continue
        col = pal[i % len(pal)]
        label_pos = (shot_label_pos or {}).get(shot_id, 0.0)

        tr = [r["trace_no"] for r in rows]
        x_geom = [r["recv_pos_m"] for r in rows]
        t_interp = [r["fb_interp_inline_ms"] for r in rows]
        ax1.plot(tr, t_interp, "o-", ms=3, lw=1.0, color=col,
                 label=f"Shot {shot_id} (@{label_pos:.1f} m)")

        ax2.plot(x_geom, t_interp, "o-", ms=3, lw=1.0, color=col, alpha=0.9)

        ax3.plot(tr, t_interp, "o-", ms=3, lw=1.0, color=col, alpha=0.9,
                 label=f"Shot {shot_id}")
        ax3_top.plot(x_geom, t_interp, "--", lw=0.9, color=col, alpha=0.7)

    # Secondary top axis on geometry-x panel: channels
    sample_rows: list = []
    for sid in sorted(corrected_by_shot):
        rr = corrected_by_shot.get(sid, [])
        if rr:
            sample_rows = rr
            break
    if sample_rows:
        ch_from_x = sorted([(float(r["recv_pos_m"]), int(r["trace_no"]))
                            for r in sample_rows], key=lambda p: p[0])
        if ch_from_x:
            x_vals = np.array([p[0] for p in ch_from_x], dtype=float)
            ch_vals = [p[1] for p in ch_from_x]
            ax2_top = ax2.twiny()
            ax2_top.set_xlim(ax2.get_xlim())
            step = max(1, len(x_vals) // 10)
            idxs = list(range(0, len(x_vals), step))
            if (len(x_vals) - 1) not in idxs:
                idxs.append(len(x_vals) - 1)
            ax2_top.set_xticks(x_vals[idxs])
            ax2_top.set_xticklabels([str(ch_vals[i]) for i in idxs], fontsize=7)
            ax2_top.set_xlabel("Channel", color=c["label"], fontsize=8)
            ax2_top.tick_params(colors=c["tick"], labelsize=7)
            for sp in ax2_top.spines.values():
                sp.set_edgecolor(c["spine"])

    ax1.set_xlabel("Channel", color=c["label"])
    ax1.set_ylabel("First-break time (ms)", color=c["label"])
    ax1.set_title("Interpolated FB vs Channel", color=c["text"], fontsize=10)

    ax2.set_xlabel("Geometry x (m)", color=c["label"])
    ax2.set_title("Interpolated FB vs geometry x", color=c["text"], fontsize=10)

    ax3.set_xlabel("Channel", color=c["label"])
    ax3_top.set_xlabel("Geometry x (m)", color=c["label"], fontsize=8)
    ax3.set_ylabel("First-break time (ms)", color=c["label"])
    ax3.set_title("Interpolated FB vs channel / geometry", color=c["text"], fontsize=10)

    for ax in (ax1, ax2, ax3):
        ax.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
        ax.tick_params(colors=c["tick"])
        for sp in ax.spines.values():
            sp.set_edgecolor(c["spine"])

    ax3_top.tick_params(colors=c["tick"], labelsize=8)
    for sp in ax3_top.spines.values():
        sp.set_edgecolor(c["spine"])

    ax1.legend(fontsize=8, facecolor=c["leg_face"], edgecolor=c["leg_edge"],
               labelcolor=c["text"])
    ax1.set_ylim(T_MAX_MS, 0.0)
    ax2.set_ylim(T_MAX_MS, 0.0)
    ax3.set_ylim(T_MAX_MS, 0.0)

    if layer_results:
        txt_lines: list = []
        for sid in sorted(layer_results):
            shot_sides = layer_results.get(sid, {}) or {}
            ordered_sides = [s for s in ("L", "R", "ALL") if s in shot_sides]
            ordered_sides += [s for s in shot_sides if s not in ordered_sides]
            for side in ordered_sides:
                lr = shot_sides.get(side)
                if not lr:
                    continue
                txt_lines.append(
                    f"S{sid}-{side}: V0={lr.get('V0_m_s',0):.0f}, "
                    f"V1={lr.get('V1_m_s',0):.0f}, V2={lr.get('V2_m_s',0):.0f} m/s"
                )
        if txt_lines:
            ax2.text(0.01, 0.01, "\n".join(txt_lines[:8]), transform=ax2.transAxes,
                     fontsize=7, color=c["text"], va="bottom", ha="left",
                     bbox=dict(facecolor=c["ax_bg"], edgecolor=c["spine"], alpha=0.75))

    fig.suptitle(f"Corrected first-break QC  --  Profile {profile_name}",
                 color=c["text"], fontsize=11)
    fig.tight_layout()

    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{profile_name}_corrected_qc{filename_suffix}.png"
    if show_plot:
        try:
            backend = str(plt.get_backend()).lower()
        except Exception:
            backend = ""
        if "agg" in backend:
            show_plot = _ensure_interactive_backend()
    if show_plot:
        try:
            plt.show(block=True)
        except Exception:
            pass
    fig.savefig(str(out), dpi=180, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  corrected_qc -> {out.relative_to(CWD.parent)}")
    return out


# ---------------------------------------------------------------------------
# Interactive first-break picker
# ---------------------------------------------------------------------------

class FirstBreakPicker:
    """
    Matplotlib interactive first-break picker for one shot record.

    Key design decisions
    --------------------
    - run() uses plt.show(block=False) + a manual polling loop.
      This avoids the double-open artefact caused by plt.show(block=True)
      on Windows (figure re-appearing after plt.close was called by a key handler).
        - Prev/Next store picks in session and keep the picking loop active.
        - Save/Close ('s') finalizes the profile and exits to export stage.
    - The JSON always stores RAW picks; BULK_SHIFT_MS is applied later.
    """

    def __init__(self, data_raw: Any, data_filt: Any, dt_s: float,
                 recv_abs: Any, shot_id: int, profile_name: str,
                 shot_pos_m: float = 0.0,
                 delay_ms: float = 0.0,
                 existing_picks: dict | None = None,
                 qc_dir: Path | None = None,
                 save_callback: Any = None,
                 header_info: dict | None = None):
        self.data_raw   = data_raw
        self.data_filt  = data_filt
        self.dt_s       = dt_s
        self.recv_abs   = recv_abs      # absolute receiver positions (m)
        self.shot_pos_m = shot_pos_m
        self.delay_ms   = delay_ms      # time of first sample relative to trigger (ms)
        self.shot_id    = shot_id
        self.profile    = profile_name
        self.qc_dir     = qc_dir or (OUTPUT_DIR / profile_name)

        self._picks: dict = {}
        self._saved          = False
        self._cancelled      = False
        self._done           = False
        self._filter_mode    = "ormsby" if FILTER_ON else "none"  # none|ormsby|butter
        self._inverted       = True
        self._top_mode       = "channel"   # "channel" or "offset" (toggle with 'c')
        self.ax_top: Any     = None
        self._save_callback  = save_callback   # called on nav/finalize with picks dict
        self._header_info    = header_info or {}
        self._nav_action     = "stay"   # "next" | "prev" | "finalize" | "quit" | "stay"
        self._drag_pick      = False
        self._drag_delete    = False
        self._last_drag_idx: Any = None
        self._last_drag_t: float | None = None
        self._range_anchor: Any = None
        self._gain_mode      = GAIN_MODE
        self._agc_window_ms  = AGC_WINDOW_MS
        self._agc_stat       = AGC_STAT
        self._display_mode   = DISPLAY_MODE
        self._show_timelines = SHOW_TIMELINES
        self._wiggle_stretch = WIGGLE_STRETCH
        self._f1 = BP_F1
        self._f2 = BP_F2
        self._f3 = BP_F3
        self._f4 = BP_F4
        self._butter_order = BUTTER_ORDER
        self._filter_debounce_ms = max(0, int(FILTER_DEBOUNCE_MS))
        self._filter_timer: Any = None

        self.data_ormsby = self.data_filt
        self.data_butter = apply_butterworth_all_params(
            self.data_raw, self.dt_s,
            low_hz=self._f2, high_hz=self._f3,
            order=self._butter_order,
        )

        if existing_picks:
            self._picks = {int(k): float(v) for k, v in existing_picks.items()}

        self.n_traces = data_raw.shape[0]
        self.n_samp   = data_raw.shape[1]
        dt_ms         = dt_s * 1000.0

        # Absolute time axis:  first sample at delay_ms, increases by dt_ms
        self.times_ms = delay_ms + np.arange(self.n_samp) * dt_ms

        # Display window: T_DISPLAY_PRE_MS (< 0) to T_MAX_MS
        self.t_disp_start = max(delay_ms, T_DISPLAY_PRE_MS)
        self.t_disp_end   = T_MAX_MS
        self._t_view_start = float(self.t_disp_start)
        self._t_view_end = float(self.t_disp_end)
        self._t_full_start = float(delay_ms)
        self._t_full_end = float(delay_ms + (self.n_samp - 1) * dt_ms)
        self._t_view_end = min(self._t_view_end, self._t_full_end)
        t_start_idx = max(0, int((self.t_disp_start - delay_ms) / dt_ms))
        t_end_idx   = min(int((self.t_disp_end   - delay_ms) / dt_ms) + 2,
                          self.n_samp)
        self.t_slice = slice(t_start_idx, t_end_idx)

        # Trace spacing for wiggle display
        off_range = float(np.ptp(recv_abs)) if len(recv_abs) > 1 else 1.0
        self._dx  = off_range / max(len(recv_abs) - 1, 1)

        self._build_figure()

    # ---- figure ------------------------------------------------------------

    def _build_figure(self):
        c = _tc()
        self.fig, self.ax = plt.subplots(figsize=(16, 8), constrained_layout=False)
        self.fig.subplots_adjust(left=0.06, right=0.76, bottom=0.13, top=0.91)
        self.fig.patch.set_facecolor(c["fig_bg"])
        self.ax.set_facecolor(c["ax_bg"])
        try:
            self.fig.canvas.manager.set_window_title(
                f"LVL Picker  --  {self.profile}  Shot {self.shot_id}")
        except Exception:
            pass
        self.fig.canvas.mpl_connect("button_press_event", self._on_click)
        self.fig.canvas.mpl_connect("motion_notify_event", self._on_motion)
        self.fig.canvas.mpl_connect("button_release_event", self._on_release)
        self.fig.canvas.mpl_connect("key_press_event",    self._on_key)
        self._build_controls()
        self._redraw()   # creates ax_top secondary axis inside

    def _build_controls(self):
        c = _tc()
        x0 = 0.78
        w  = 0.20
        row_h = 0.032
        gap = 0.005
        row3_w = (w - 2 * gap) / 3.0

        self.ax_btn_prev = self.fig.add_axes([x0 + 0 * (row3_w + gap), 0.88, row3_w, row_h])
        self.ax_btn_next = self.fig.add_axes([x0 + 1 * (row3_w + gap), 0.88, row3_w, row_h])
        self.ax_btn_save = self.fig.add_axes([x0 + 2 * (row3_w + gap), 0.88, row3_w, row_h])

        self.ax_btn_auto = self.fig.add_axes([x0 + 0 * (row3_w + gap), 0.84, row3_w, row_h])
        self.ax_btn_inv  = self.fig.add_axes([x0 + 1 * (row3_w + gap), 0.84, row3_w, row_h])
        self.ax_btn_tl   = self.fig.add_axes([x0 + 2 * (row3_w + gap), 0.84, row3_w, row_h])

        btn_face = c["ax_bg"]
        btn_hover = "#e8e8e8" if THEME == "light" else "#2a2d3d"

        self.btn_prev = Button(self.ax_btn_prev, "Prev", color=btn_face, hovercolor=btn_hover)
        self.btn_next = Button(self.ax_btn_next, "Next", color=btn_face, hovercolor=btn_hover)
        self.btn_save = Button(self.ax_btn_save, "Save/Close", color=btn_face, hovercolor=btn_hover)
        self.btn_auto = Button(self.ax_btn_auto, "Auto Picker", color=btn_face, hovercolor=btn_hover)
        self.btn_inv  = Button(self.ax_btn_inv, "", color=btn_face, hovercolor=btn_hover)
        self.btn_tl   = Button(self.ax_btn_tl, "Timeline", color=btn_face, hovercolor=btn_hover)
        for btn in (self.btn_prev, self.btn_save, self.btn_next,
                    self.btn_auto, self.btn_inv, self.btn_tl):
            btn.label.set_color(c["text"])
            btn.label.set_fontsize(8)
            for sp in btn.ax.spines.values():
                sp.set_edgecolor(c["spine"])

        self._update_invert_button_label()
        self._refresh_toggle_buttons()

        self.btn_prev.on_clicked(lambda _e: self._go_prev())
        self.btn_save.on_clicked(lambda _e: self._save_and_finish("finalize"))
        self.btn_next.on_clicked(lambda _e: self._save_and_finish("next"))
        self.btn_auto.on_clicked(lambda _e: self._auto_then_redraw())
        self.btn_inv.on_clicked(lambda _e: self._toggle_invert())
        self.btn_tl.on_clicked(lambda _e: self._toggle_timelines())

        label_fs = 8
        value_fs = 8
        self.fig.text(x0, 0.80, "Gain Control:", fontsize=label_fs, fontweight="bold",
                  color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.76, "AGC stat:", fontsize=label_fs, fontweight="bold",
                  color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.72, "Display:", fontsize=label_fs, fontweight="bold",
              color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.620, "Filter mode:", fontsize=label_fs, fontweight="bold",
              color=c["text"], ha="left", va="bottom")
        self.fig.text(x0, 0.580, "Butter order:", fontsize=label_fs, fontweight="bold",
              color=c["text"], ha="left", va="bottom")

        self._gain_labels = ("none", "norm", "agc")
        self._gain_btns = self._make_mode_buttons(
            x0=x0 + 0.055, y=0.792, w=0.145, h=0.026,
            labels=self._gain_labels, callback=self._on_gain_mode,
        )

        self._stat_labels = ("rms", "mean")
        self._stat_btns = self._make_mode_buttons(
            x0=x0 + 0.055, y=0.752, w=0.095, h=0.026,
            labels=self._stat_labels, callback=self._on_agc_stat,
        )

        self._disp_labels = ("wiggle", "vd", "both")
        self._disp_btns = self._make_mode_buttons(
            x0=x0 + 0.055, y=0.712, w=0.145, h=0.026,
            labels=self._disp_labels, callback=self._on_display_mode,
        )

        self._filt_labels = ("none", "butter", "ormsby")
        self._filt_btns = self._make_mode_buttons(
            x0=x0 + 0.055, y=0.620, w=0.145, h=0.026,
            labels=self._filt_labels, callback=self._on_filter_mode,
        )

        self._butter_order_labels = ("2", "4", "6")
        self._butter_order_btns = self._make_mode_buttons(
            x0=x0 + 0.055, y=0.580, w=0.095, h=0.026,
            labels=self._butter_order_labels, callback=self._on_butter_order,
        )

        self.ax_agc = self.fig.add_axes([x0+0.02, 0.67, 0.25*w, 0.016], facecolor=c["ax_bg"])
        self.sl_agc = Slider(self.ax_agc, "AGC ", 5.0, 500.0,
                             valinit=float(self._agc_window_ms), valstep=1.0)
        self.sl_agc.label.set_fontweight("bold")
        self.sl_agc.label.set_fontsize(8)
        self.sl_agc.valtext.set_fontsize(8)
        self.sl_agc.on_changed(self._on_agc_window)

        self.ax_wig = self.fig.add_axes([x0+0.125, 0.67, 0.25*w, 0.016], facecolor=c["ax_bg"])
        self.sl_wig = Slider(self.ax_wig, "Scale ", 0.20, 5.00,
                     valinit=float(self._wiggle_stretch), valstep=0.01)
        self.sl_wig.label.set_fontweight("bold")
        self.sl_wig.label.set_fontsize(8)
        self.sl_wig.valtext.set_fontsize(8)
        self.sl_wig.on_changed(self._on_wiggle_stretch)

        self.fig.text(x0, 0.485, "Time View (ms)", fontsize=8, fontweight="bold",
                  color=c["text"], ha="left", va="bottom")
        tmin_hi = max(self._t_full_start + 1.0, self._t_full_end - 1.0)
        tmax_lo = min(self._t_full_end - 1.0, self._t_full_start + 1.0)
        self.ax_tmin = self.fig.add_axes([x0 + 0.02, 0.455, 0.25*w, 0.014], facecolor=c["ax_bg"])
        self.ax_tmax = self.fig.add_axes([x0 + 0.125, 0.455, 0.25*w, 0.014], facecolor=c["ax_bg"])
        self.sl_tmin = Slider(self.ax_tmin, "T min", self._t_full_start, tmin_hi,
                      valinit=float(self._t_view_start), valstep=1.0)
        self.sl_tmax = Slider(self.ax_tmax, "T max", tmax_lo, self._t_full_end,
                      valinit=float(self._t_view_end), valstep=1.0)
        self.sl_tmin.label.set_fontweight("bold")
        self.sl_tmax.label.set_fontweight("bold")
        self.sl_tmin.label.set_fontsize(8)
        self.sl_tmax.label.set_fontsize(8)
        self.sl_tmin.valtext.set_fontsize(8)
        self.sl_tmax.valtext.set_fontsize(8)
        self.sl_tmin.on_changed(self._on_tmin)
        self.sl_tmax.on_changed(self._on_tmax)

        self.ax_f1 = self.fig.add_axes([x0+0.0125, 0.550, 0.25*w, 0.014], facecolor=c["ax_bg"])
        self.ax_f2 = self.fig.add_axes([x0+0.125, 0.550, 0.25*w, 0.014], facecolor=c["ax_bg"])
        self.ax_f3 = self.fig.add_axes([x0+0.0125, 0.525, 0.25*w, 0.014], facecolor=c["ax_bg"])
        self.ax_f4 = self.fig.add_axes([x0+0.125, 0.525, 0.25*w, 0.014], facecolor=c["ax_bg"])

        self.sl_f1 = Slider(self.ax_f1, "f1 ", 0.0, 50.0, valinit=float(self._f1), valstep=0.5)
        self.sl_f2 = Slider(self.ax_f2, "f2 ", 0.5, 80.0, valinit=float(self._f2), valstep=0.5)
        self.sl_f3 = Slider(self.ax_f3, "f3 ", 20.0, 220.0, valinit=float(self._f3), valstep=1.0)
        self.sl_f4 = Slider(self.ax_f4, "f4 ", 40.0, 260.0, valinit=float(self._f4), valstep=1.0)
        self.sl_f1.on_changed(self._on_filter_sliders)
        self.sl_f2.on_changed(self._on_filter_sliders)
        self.sl_f3.on_changed(self._on_filter_sliders)
        self.sl_f4.on_changed(self._on_filter_sliders)

        # Header/parameter info panel
        self.ax_info = self.fig.add_axes([x0, 0.125, w, 0.23], facecolor=c["ax_bg"])
        self.ax_info.set_xticks([])
        self.ax_info.set_yticks([])
        for sp in self.ax_info.spines.values():
            sp.set_edgecolor(c["spine"])
        self.info_text = self.ax_info.text(0.02, 0.98, "", va="top", ha="left",
                                           fontsize=8, color=c["text"],
                                           transform=self.ax_info.transAxes)
        self._refresh_mode_buttons()

    def _make_mode_buttons(self, x0: float, y: float, w: float, h: float,
                           labels: tuple, callback: Any) -> list:
        c = _tc()
        value_fs = 8
        gap = 0.004
        n = max(1, len(labels))
        bw = (w - gap * (n - 1)) / n
        buttons: list = []
        for i, lbl in enumerate(labels):
            axb = self.fig.add_axes([x0 + i * (bw + gap), y, bw, h])
            btn = Button(axb, str(lbl), color=c["ax_bg"],
                         hovercolor="#e8e8e8" if THEME == "light" else "#2a2d3d")
            btn.label.set_color(c["text"])
            btn.label.set_fontsize(value_fs)
            for sp in btn.ax.spines.values():
                sp.set_edgecolor(c["spine"])
            btn.on_clicked(lambda _e, _lbl=str(lbl): callback(_lbl))
            buttons.append(btn)
        return buttons

    def _style_mode_buttons(self, buttons: list, labels: tuple, active_label: str):
        c = _tc()
        for btn, lbl in zip(buttons, labels):
            is_active = (str(lbl).lower() == str(active_label).lower())
            active_face = "#d9e8ff" if THEME == "light" else "#2a3b5c"
            inactive_face = c["ax_bg"]
            btn.ax.set_facecolor(active_face if is_active else inactive_face)
            btn.color = active_face if is_active else inactive_face
            btn.hovercolor = "#cfe2ff" if THEME == "light" else "#324972"
            btn.label.set_color(c["text"])

    def _refresh_mode_buttons(self):
        if hasattr(self, "_gain_btns"):
            self._style_mode_buttons(self._gain_btns, self._gain_labels, self._gain_mode)
        if hasattr(self, "_stat_btns"):
            self._style_mode_buttons(self._stat_btns, self._stat_labels, self._agc_stat)
        if hasattr(self, "_disp_btns"):
            self._style_mode_buttons(self._disp_btns, self._disp_labels, self._display_mode)
        if hasattr(self, "_filt_btns"):
            self._style_mode_buttons(self._filt_btns, self._filt_labels, self._filter_mode)
        if hasattr(self, "_butter_order_btns"):
            self._style_mode_buttons(
                self._butter_order_btns,
                self._butter_order_labels,
                str(int(self._butter_order)),
            )

    def _refresh_toggle_buttons(self):
        c = _tc()
        if hasattr(self, "btn_tl"):
            active_face = "#d9e8ff" if THEME == "light" else "#2a3b5c"
            inactive_face = c["ax_bg"]
            self.btn_tl.ax.set_facecolor(active_face if self._show_timelines else inactive_face)
            self.btn_tl.color = active_face if self._show_timelines else inactive_face
            self.btn_tl.hovercolor = "#cfe2ff" if THEME == "light" else "#324972"
            self.btn_tl.label.set_color(c["text"])
        if hasattr(self, "btn_inv"):
            active_face = "#d9e8ff" if THEME == "light" else "#2a3b5c"
            inactive_face = c["ax_bg"]
            self.btn_inv.ax.set_facecolor(active_face if self._inverted else inactive_face)
            self.btn_inv.color = active_face if self._inverted else inactive_face
            self.btn_inv.hovercolor = "#cfe2ff" if THEME == "light" else "#324972"
            self.btn_inv.label.set_color(c["text"])

    def _update_invert_button_label(self):
        if hasattr(self, "btn_inv"):
            self.btn_inv.label.set_text("Normal Pol" if self._inverted else "Inverse Pol")
            self._refresh_toggle_buttons()

    def _on_tmin(self, val: float):
        v = float(val)
        if v >= self._t_view_end - 1.0:
            v = self._t_view_end - 1.0
            self.sl_tmin.set_val(v)
            return
        self._t_view_start = v
        self._redraw()

    def _on_tmax(self, val: float):
        v = float(val)
        if v <= self._t_view_start + 1.0:
            v = self._t_view_start + 1.0
            self.sl_tmax.set_val(v)
            return
        self._t_view_end = v
        self._redraw()

    def _on_gain_mode(self, label: str):
        self._gain_mode = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _on_agc_stat(self, label: str):
        self._agc_stat = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _on_agc_window(self, val: float):
        self._agc_window_ms = float(val)
        if self._gain_mode == "agc":
            self._redraw()

    def _on_wiggle_stretch(self, val: float):
        self._wiggle_stretch = float(val)
        self._redraw()

    def _on_display_mode(self, label: str):
        self._display_mode = str(label)
        self._refresh_mode_buttons()
        self._redraw()

    def _on_filter_mode(self, label: str):
        self._filter_mode = str(label).lower()
        self._refresh_mode_buttons()
        self._redraw()

    def _on_butter_order(self, label: str):
        try:
            order = int(str(label).strip())
        except Exception:
            return
        if order < 1 or order == int(self._butter_order):
            return
        self._butter_order = int(order)
        self._schedule_filter_update()
        self._refresh_mode_buttons()

    def _schedule_filter_update(self):
        if self._filter_debounce_ms <= 0:
            self._apply_filter_update_now()
            return

        # Recreate timer each time to debounce rapid slider events.
        try:
            if self._filter_timer is not None:
                self._filter_timer.stop()
        except Exception:
            pass

        self._filter_timer = self.fig.canvas.new_timer(interval=int(self._filter_debounce_ms))
        self._filter_timer.single_shot = True
        self._filter_timer.add_callback(self._apply_filter_update_now)
        self._filter_timer.start()

    def _apply_filter_update_now(self):
        self._recompute_filter()
        self._redraw()

    def _toggle_invert(self):
        self._inverted = not self._inverted
        self._update_invert_button_label()
        self._redraw()

    def _toggle_timelines(self):
        self._show_timelines = not self._show_timelines
        self._refresh_toggle_buttons()
        self._redraw()

    def _on_filter_sliders(self, _val: float):
        f1 = float(self.sl_f1.val)
        f2 = float(self.sl_f2.val)
        f3 = float(self.sl_f3.val)
        f4 = float(self.sl_f4.val)
        if not (f1 < f2 < f3 < f4):
            # Ignore transient invalid states while dragging neighboring sliders.
            return
        self._f1, self._f2, self._f3, self._f4 = f1, f2, f3, f4
        self._schedule_filter_update()

    # ---- data --------------------------------------------------------------

    def _active_data(self) -> Any:
        if self._filter_mode == "ormsby":
            d = self.data_ormsby
        elif self._filter_mode == "butter":
            d = self.data_butter
        else:
            d = self.data_raw
        d = apply_gain(d, self.dt_s, mode=self._gain_mode,
                       window_ms=self._agc_window_ms, stat=self._agc_stat)
        return -d if self._inverted else d

    def _recompute_filter(self):
        self.data_ormsby = apply_ormsby_all_params(
            self.data_raw, self.dt_s, self._f1, self._f2, self._f3, self._f4
        )
        self.data_butter = apply_butterworth_all_params(
            self.data_raw, self.dt_s,
            low_hz=self._f2, high_hz=self._f3,
            order=self._butter_order,
        )

    # ---- drawing -----------------------------------------------------------

    def _draw_traces(self):
        c    = _tc()
        data = self._active_data()
        dt_ms = self.dt_s * 1000.0
        i0 = max(0, int((self._t_view_start - self.delay_ms) / dt_ms))
        i1 = min(self.n_samp, int((self._t_view_end - self.delay_ms) / dt_ms) + 2)
        if i1 <= i0:
            i1 = min(self.n_samp, i0 + 2)
        ts = slice(i0, i1)
        t_ms = self.times_ms[ts]
        dx   = self._dx

        if self._display_mode in ("vd", "both"):
            d_img = data[:, ts]
            vmax = float(np.percentile(np.abs(d_img), 98)) if d_img.size else 1.0
            if vmax < 1e-12:
                vmax = 1.0
            self.ax.imshow(
                d_img.T,
                aspect="auto",
                cmap="seismic",
                vmin=-vmax,
                vmax=vmax,
                extent=[self.recv_abs.min(), self.recv_abs.max(), t_ms[-1], t_ms[0]],
                alpha=0.45,
                interpolation="nearest",
                zorder=1,
            )

        if self._display_mode == "vd":
            return

        # Shared-median normalisation with per-trace clamping.
        # Using the global median keeps the filter effect visible (filtered
        # data appears quieter overall vs unfiltered).
        # Clamping each trace's own std to [median*0.3, median*3] prevents
        # loud near-offset traces from being over-squashed and dead traces
        # from being amplified into noise.
        stds  = np.array([data[i, ts].astype(float).std()
                          for i in range(self.n_traces)])
        valid = stds[stds > 1e-20]
        med   = float(np.median(valid)) if len(valid) else 1.0
        norms = np.clip(stds, med * 0.3, med * 3.0)
        norms = np.where(norms > 1e-20, norms, med)

        scale = norms * CLIP_FACTOR   # shape (n_traces,)

        for i, off in enumerate(self.recv_abs):
            tr   = data[i, ts].astype(float)
            tr_n = np.clip(tr / scale[i], -1.0, 1.0)
            x_w  = off + tr_n * dx * self._wiggle_stretch
            self.ax.plot(x_w, t_ms, color=c["trace"], lw=0.5, alpha=0.85)
            pos = np.where(tr_n > 0.0, tr_n, 0.0)
            self.ax.fill_betweenx(t_ms, off, off + pos * dx * self._wiggle_stretch,
                                  color=c["trace"], alpha=c["fill_alpha"])

    def _draw_picks(self):
        c = _tc()
        for idx, t in self._picks.items():
            if idx < len(self.recv_abs):
                self.ax.plot(self.recv_abs[idx], t, "v",
                             color=c["pick_clr"], ms=7, zorder=10,
                             markeredgecolor=c["pick_clr"],
                             markeredgewidth=0.6)

    def _redraw(self):
        # Remove the previous top axis BEFORE clearing the main axis;
        # twiny() axes are siblings in the figure and are NOT cleared by ax.clear().
        if getattr(self, 'ax_top', None) is not None:
            try:
                self.ax_top.remove()
            except Exception:
                pass
            self.ax_top = None

        c = _tc()
        self.ax.clear()
        self.ax.set_facecolor(c["ax_bg"])
        self._draw_traces()
        self._draw_picks()
        # Shot position marker (vertical) and trigger time (horizontal)
        self.ax.axvline(self.shot_pos_m, color="#ffcc00", lw=1.2,
                        ls="--", alpha=0.70, zorder=3)
        if self._show_timelines:
            self.ax.axhline(0.0, color="#00aa66", lw=1.0,
                            ls=":", alpha=0.75, zorder=3)
            for ms in range(10, int(self._t_view_end) + 1, 10):
                if self._t_view_start <= float(ms) <= self._t_view_end:
                    major = (ms % 20 == 0)
                    self.ax.axhline(
                        float(ms),
                        color="#666666" if THEME == "light" else "#aaaaaa",
                        lw=0.50 if major else 0.30,
                        ls=":" if major else "--",
                        alpha=0.35 if major else 0.20,
                        zorder=2,
                    )

        margin = self._dx * 0.5
        self.ax.set_xlim(self.recv_abs.min() - margin,
                         self.recv_abs.max() + margin)
        # Y-axis: time increases downward; 0 ms at top, T_MAX_MS at bottom
        self.ax.set_ylim(self._t_view_end, self._t_view_start)

        dt_ms  = self.dt_s * 1000.0
        n_samp = self.n_samp
        t_end  = self.delay_ms + (n_samp - 1) * dt_ms
        if self._filter_mode == "ormsby":
            filt_str = f"Ormsby {self._f1:.0f}-{self._f2:.0f}-{self._f3:.0f}-{self._f4:.0f} Hz"
        elif self._filter_mode == "butter":
            filt_str = f"Butterworth order {self._butter_order} ({self._f2:.0f}-{self._f3:.0f} Hz)"
        else:
            filt_str = "Filter none"
        if self._inverted:
            filt_str += "  [INV]"
        title = (f"Profile: {self.profile}  |  Shot {self.shot_id}  |"
                 f"  {n_samp} smp  dt={dt_ms:.4f} ms  "
                 f"delay={self.delay_ms:.1f} ms  end={t_end:.1f} ms  |"
                 f"  {len(self._picks)}/{self.n_traces} picks  |  {filt_str}"
                 f"  | gain={self._gain_mode}")
        self.ax.set_title(title, color=c["text"], fontsize=9)
        self.ax.set_xlabel("Receiver position (m)", color=c["label"], fontsize=8)
        self.ax.text(
            0.0, -0.14,
            "L:pick  R:delete  Shift+L:range  a:auto  f:filter mode  g:gain  v:polarity  l:timeline  c:top axis  s/n:save+next  p:prev  q:quit",
            transform=self.ax.transAxes,
            ha="left", va="top",
            color=c["label"], fontsize=7,
            clip_on=False,
        )
        self.ax.set_ylabel("Time  (ms)", color=c["label"])
        self.ax.grid(True, lw=0.3, alpha=0.30, color=c["grid"])
        self.ax.tick_params(colors=c["tick"])
        for sp in self.ax.spines.values():
            sp.set_edgecolor(c["spine"])
        handle = Line2D([0], [0], marker="v", linestyle="none",
                        color=c["pick_clr"], markersize=7,
                        label="First-break pick")
        self.ax.legend(handles=[handle], fontsize=8,
                       facecolor=c["leg_face"], edgecolor=c["leg_edge"],
                       labelcolor=c["text"], loc="lower right")

        # ---- Top axis: channel numbers (default) or signed offset from shot ----
        # twiny() shares the y-axis, giving an independent x-axis at the top.
        # We set matching xlim and place ticks at actual receiver positions so
        # the labels are exact regardless of whether spacing is uniform or not.
        ax_top = self.ax.twiny()
        ax_top.set_xlim(self.ax.get_xlim())
        n_recv = len(self.recv_abs)
        step   = max(1, n_recv // 8)          # aim for ~8-10 tick labels
        idxs   = list(range(0, n_recv, step))
        if (n_recv - 1) not in idxs:
            idxs.append(n_recv - 1)

        if self._top_mode == "channel":
            ax_top.set_xticks(self.recv_abs[idxs])
            ax_top.set_xticklabels([str(i + 1) for i in idxs], fontsize=7)
            ax_top.set_xlabel("Channel   [c -> offset]",
                              color=c["label"], fontsize=8)
        else:
            spm = self.shot_pos_m
            ax_top.set_xticks(self.recv_abs[idxs])
            ax_top.set_xticklabels(
                [f"{self.recv_abs[i] - spm:+.0f}" for i in idxs], fontsize=7)
            ax_top.set_xlabel("Signed offset from shot (m)   [c -> channel]",
                              color=c["label"], fontsize=8)

        ax_top.tick_params(colors=c["tick"], labelsize=7)
        for sp in ax_top.spines.values():
            sp.set_edgecolor(c["spine"])
        self.ax_top = ax_top

        ffid = self._header_info.get("ffid", "?")
        ntr = self._header_info.get("n_tr", self.n_traces)
        nsamp = self._header_info.get("n_samp", self.n_samp)
        s_hdr = self._header_info.get("shot_pos_hdr", None)
        info_block = (
            r"$\bf{Header\ Info}$" + "\n"
            f"FFID           : {ffid}\n"
            f"Traces         : {ntr}\n"
            f"Samples        : {nsamp}\n"
            f"dt (ms)        : {dt_ms:.4f}\n"
            f"Delay (ms)     : {self.delay_ms:.1f}\n"
            f"Shot (m)       : {self.shot_pos_m:.2f}\n"
            f"Gain           : {self._gain_mode} ({self._agc_stat}, {self._agc_window_ms:.0f} ms)\n"
            f"Display / Scale: {self._display_mode} / {self._wiggle_stretch:.2f}\n"
            f"Filter mode    : {self._filter_mode}\n"
            f"Ormsby (Hz)    : {self._f1:.1f}-{self._f2:.1f}-{self._f3:.1f}-{self._f4:.1f}\n"
            f"Butter (Hz)    : {self._f2:.1f}-{self._f3:.1f} (order {self._butter_order})\n"
            f"Time view (ms) : {self._t_view_start:.1f} .. {self._t_view_end:.1f}\n"
            f"Polarity       : {'inverse' if self._inverted else 'normal'}\n"
            f"Timelines      : {'on' if self._show_timelines else 'off'}"
        )
        if hasattr(self, "info_text"):
            self.info_text.set_text(info_block)

        self._refresh_mode_buttons()
        self._refresh_toggle_buttons()
        self.fig.patch.set_facecolor(c["fig_bg"])
        self.fig.canvas.draw_idle()

    # ---- events ------------------------------------------------------------

    def _nearest_idx(self, x: float) -> int:
        return int(np.abs(self.recv_abs - x).argmin())

    def _on_click(self, event: Any):
        # twiny() creates an overlay axis (ax_top) at the same figure position as
        # self.ax.  On Windows/TkAgg, mouse events are routed to whichever axes
        # was created last, so event.inaxes is often ax_top, not self.ax.
        # Fix: accept either axis, then convert pixel coordinates to self.ax
        # data space so xdata/ydata are always in the correct (geometry) units.
        ax_top_ref  = self.ax_top          # may be None before first _redraw
        valid_axes  = (self.ax,) if ax_top_ref is None else (self.ax, ax_top_ref)
        if event.inaxes not in valid_axes or event.x is None or event.y is None:
            return
        try:
            xdata, ydata = self.ax.transData.inverted().transform(
                (event.x, event.y))
        except Exception:
            if event.xdata is None or event.ydata is None:
                return
            xdata, ydata = event.xdata, event.ydata
        idx = self._nearest_idx(xdata)
        if event.key == "shift" and event.button == 1:
            self._range_fill(idx, float(ydata))
            self._redraw()
            return
        if event.button == 1:
            t = round(float(ydata), 2)
            if self._t_view_start <= t <= self._t_view_end:
                self._picks[idx] = t
                self._drag_pick = True
                self._last_drag_idx = idx
                self._last_drag_t = t
        elif event.button == 3:
            self._picks.pop(idx, None)
            self._drag_delete = True
            self._last_drag_idx = idx
            self._last_drag_t = None
        self._redraw()

    def _on_motion(self, event: Any):
        if not (self._drag_pick or self._drag_delete):
            return
        ax_top_ref = self.ax_top
        valid_axes = (self.ax,) if ax_top_ref is None else (self.ax, ax_top_ref)
        if event.inaxes not in valid_axes or event.x is None or event.y is None:
            return
        try:
            xdata, ydata = self.ax.transData.inverted().transform((event.x, event.y))
        except Exception:
            return
        idx = self._nearest_idx(xdata)
        if self._drag_pick:
            t = round(float(ydata), 2)
            if self._last_drag_idx is None:
                if self._t_view_start <= t <= self._t_view_end:
                    self._picks[idx] = t
                self._last_drag_idx = idx
                self._last_drag_t = t
                self._redraw()
                return

            prev_idx = int(self._last_drag_idx)
            prev_t = float(self._last_drag_t if self._last_drag_t is not None else t)
            if idx == prev_idx:
                return

            step = 1 if idx > prev_idx else -1
            span = abs(idx - prev_idx)
            for n, ii in enumerate(range(prev_idx + step, idx + step, step), start=1):
                frac = n / float(span)
                ti = round(float(prev_t + frac * (t - prev_t)), 2)
                if self._t_view_start <= ti <= self._t_view_end:
                    self._picks[ii] = ti
            self._last_drag_idx = idx
            self._last_drag_t = t
        elif self._drag_delete:
            if self._last_drag_idx is None:
                self._picks.pop(idx, None)
                self._last_drag_idx = idx
                self._redraw()
                return

            prev_idx = int(self._last_drag_idx)
            if idx == prev_idx:
                return
            lo, hi = (prev_idx, idx) if prev_idx < idx else (idx, prev_idx)
            for ii in range(lo, hi + 1):
                self._picks.pop(ii, None)
            self._last_drag_idx = idx
            self._last_drag_t = None
        self._redraw()

    def _on_release(self, _event: Any):
        self._drag_pick = False
        self._drag_delete = False
        self._last_drag_idx = None
        self._last_drag_t = None

    def _range_fill(self, idx: int, ydata: float):
        t = round(float(ydata), 2)
        if not (self._t_view_start <= t <= self._t_view_end):
            return
        if self._range_anchor is None:
            self._range_anchor = (idx, t)
            print(f"     Range anchor set at trace {idx+1}, t={t:.2f} ms")
            return
        i0, t0 = self._range_anchor
        i1, t1 = idx, t
        if i0 == i1:
            self._picks[i0] = t1
            self._range_anchor = None
            return
        lo, hi = (i0, i1) if i0 < i1 else (i1, i0)
        for ii in range(lo, hi + 1):
            frac = (ii - i0) / float(i1 - i0)
            tt = round(float(t0 + frac * (t1 - t0)), 2)
            if self._t_view_start <= tt <= self._t_view_end:
                self._picks[ii] = tt
        self._range_anchor = None
        print(f"     Range fill: traces {lo+1}-{hi+1}")

    def _save_and_finish(self, nav: str):
        if self._save_callback is not None:
            self._save_callback(self._picks)
        try:
            self._save_qc_image()
        except Exception:
            pass
        self._saved = True
        self._nav_action = nav
        self._done = True
        try:
            self.fig.canvas.stop_event_loop()
        except Exception:
            pass

    def _go_prev(self):
        self._save_and_finish("prev")

    def _auto_then_redraw(self):
        self._auto_pick()
        self._redraw()

    def _on_key(self, event: Any):
        global THEME
        key = event.key
        if key in ("q", "escape"):
            self._cancelled = True
            self._nav_action = "quit"
            self._done = True
            try: self.fig.canvas.stop_event_loop()
            except Exception: pass
        elif key == "s":
            self._save_and_finish("finalize")
        elif key == "n":
            self._save_and_finish("next")
        elif key == "p":
            self._go_prev()
        elif key == "k":
            # Save a per-shot QC screenshot (separate from pick-saving)
            self._save_qc_image()
        elif key == "f":
            self._filter_mode = {"none": "butter", "butter": "ormsby", "ormsby": "none"}.get(
                self._filter_mode, "none"
            )
            self._refresh_mode_buttons()
            self._redraw()
        elif key == "u":
            self._f2 = max(self._f1 + 0.5, self._f2 + 1.0)
            self._schedule_filter_update()
        elif key == "j":
            self._f2 = max(self._f1 + 0.5, self._f2 - 1.0)
            self._schedule_filter_update()
        elif key == "i":
            self._f3 = min(self._f4 - 1.0, self._f3 + 5.0)
            self._schedule_filter_update()
        elif key == "m":
            self._f3 = max(self._f2 + 1.0, self._f3 - 5.0)
            self._schedule_filter_update()
        elif key == "g":
            self._gain_mode = ({"none": "norm", "norm": "agc", "agc": "none"}
                               .get(self._gain_mode, "none"))
            self._redraw()
        elif key == "v":
            self._toggle_invert()
        elif key == "l":
            self._toggle_timelines()
        elif key == "t":
            THEME = "light" if THEME == "dark" else "dark"
            self._redraw()
        elif key == "a":
            self._auto_pick()
            self._redraw()
        elif key == "c":
            self._top_mode = "offset" if self._top_mode == "channel" else "channel"
            self._redraw()

    def _auto_pick(self):
        data = self._active_data()
        count = 0
        n_sta = max(1, int(STA_MS / 1000.0 / self.dt_s))
        n_lta = max(3, int(LTA_MS / 1000.0 / self.dt_s))
        start_idx = max(0, int((0.0 - self.delay_ms) / (self.dt_s * 1000.0)))
        end_idx = min(self.n_samp - n_sta - 1,
                      int((self._t_view_end - self.delay_ms) / (self.dt_s * 1000.0)))

        for i in range(self.n_traces):
            tr = data[i].astype(float)
            tr = tr - np.mean(tr[:max(1, n_lta)])
            char = np.maximum(tr, 0.0)
            search_lo = max(n_lta, start_idx)
            search_hi = max(search_lo + 2, end_idx)
            local_char = char[search_lo:search_hi]
            noise_floor = float(np.median(local_char)) if local_char.size else 0.0
            amp_floor = max(1e-12, 4.0 * noise_floor)
            picked_idx = None
            for k in range(search_lo, search_hi - 1):
                lta = char[k - n_lta:k].mean()
                if lta < 1e-30:
                    continue
                sta = char[k:k + n_sta].mean()
                ratio = sta / lta
                next_lta = char[k + 1 - n_lta:k + 1].mean()
                if next_lta < 1e-30:
                    continue
                next_sta = char[k + 1:k + 1 + n_sta].mean()
                next_ratio = next_sta / next_lta
                if ratio >= STALTA_TRIG and next_ratio >= STALTA_TRIG and sta >= amp_floor:
                    picked_idx = k
                    break
            if picked_idx is None:
                continue

            t_abs = round(self.delay_ms + picked_idx * self.dt_s * 1000.0, 2)
            if self._t_view_start <= t_abs <= self._t_view_end:
                self._picks[i] = t_abs
                count += 1

        print(f"     Auto-pick: {count}/{self.n_traces} placed (gain-aware).")

    def _save_qc_image(self):
        """Save the main plot area only (without control widgets) as PNG."""
        self.qc_dir.mkdir(parents=True, exist_ok=True)
        out = self.qc_dir / f"{self.profile}_shot{self.shot_id:02d}_picks.png"
        self.fig.canvas.draw()
        renderer = self.fig.canvas.get_renderer()
        bbox = self.ax.get_tightbbox(renderer).expanded(1.02, 1.04)
        bbox_inches = bbox.transformed(self.fig.dpi_scale_trans.inverted())
        self.fig.savefig(str(out), dpi=180, bbox_inches=bbox_inches,
                         facecolor=self.ax.get_facecolor())
        print(f"\n     QC image -> {out.name}")

    # ---- public ------------------------------------------------------------

    @property
    def picks_ms(self) -> dict:
        return self._picks

    def run(self) -> Any:
        """
        Show exactly ONE picker window and block until it is done.

        Why this approach:
        - plt.show(block=True) on Windows/TkAgg re-enters mainloop().
          If the first mainloop() was already stopped by plt.close(), the
          second call may not block, causing all 3 shot figures to be created
          before any are shown (6-window bug).
        - canvas.start_event_loop(timeout) runs the backend event loop for
          exactly `timeout` seconds per call, then returns.  This gives a
          tightly-controlled poll loop that is fully responsive.
        - Key handlers set self._done = True and call stop_event_loop() to
          return the current start_event_loop call immediately (no plt.close
          inside key handlers).
        - plt.close() is called by run() AFTER the loop exits, exactly once.

        Returns raw picks {trace_idx: ms from trigger} or None if cancelled.
        """
        self._done = False

        def _on_close(_evt: Any):
            # Fired when the user clicks the window X button
            if not (self._saved or self._cancelled):
                self._cancelled = True
            self._done = True
            try: self.fig.canvas.stop_event_loop()
            except Exception: pass

        self.fig.canvas.mpl_connect("close_event", _on_close)
        plt.show(block=False)   # display this figure (non-blocking)

        # Poll in short chunks; each chunk runs the native event loop so the
        # window remains fully interactive
        while not self._done:
            try:
                self.fig.canvas.start_event_loop(0.05)
            except Exception:
                self._cancelled = True
                break

        # Close the figure if still open (X button may have already closed it)
        try:
            if plt.fignum_exists(self.fig.number):
                plt.close(self.fig)
        except Exception:
            pass

        if self._cancelled:
            return {"status": "quit", "picks": self._picks}
        return {"status": self._nav_action or "next", "picks": self._picks}


# ---------------------------------------------------------------------------
# Picks persistence  (JSON -- always RAW, no static applied)
# ---------------------------------------------------------------------------

def _picks_json_path(profile_name: str) -> Path:
    return OUTPUT_DIR / profile_name / "picks.json"


def _session_picks_json_path(profile_name: str) -> Path:
    return OUTPUT_DIR / profile_name / "picks.session.json"


def _layer_json_path(profile_name: str) -> Path:
    return OUTPUT_DIR / profile_name / "layer_analysis.json"


def _layer_session_json_path(profile_name: str) -> Path:
    return OUTPUT_DIR / profile_name / "layer_analysis.session.json"


def _coerce_layer_results(raw: dict) -> dict:
    out: dict = {}
    for sid, side_map in (raw or {}).items():
        try:
            shot_id = int(sid)
        except Exception:
            continue
        if not isinstance(side_map, dict):
            continue
        out[shot_id] = {str(side): dict(payload) for side, payload in side_map.items()
                        if isinstance(payload, dict)}
    return out


def load_layer_json(profile_name: str) -> dict:
    p = _layer_json_path(profile_name)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as fh:
            raw = json.load(fh)
        return _coerce_layer_results(raw)
    except Exception as exc:
        print(f"  [WARN] Could not load layer_analysis.json: {exc}")
        return {}


def load_layer_session_json(profile_name: str) -> dict:
    p = _layer_session_json_path(profile_name)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as fh:
            raw = json.load(fh)
        return _coerce_layer_results(raw)
    except Exception as exc:
        print(f"  [WARN] Could not load layer_analysis.session.json: {exc}")
        return {}


def save_layer_json(profile_name: str, layer_results: dict):
    p = _layer_json_path(profile_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump({str(k): v for k, v in (layer_results or {}).items()}, fh, indent=2)
    print(f"     Layer analysis saved -> {p.relative_to(CWD.parent)}")


def save_layer_session_json(profile_name: str, layer_results: dict):
    p = _layer_session_json_path(profile_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump({str(k): v for k, v in (layer_results or {}).items()}, fh, indent=2)
    print(f"     Layer session saved -> {p.relative_to(CWD.parent)}")


def clear_layer_session_json(profile_name: str):
    p = _layer_session_json_path(profile_name)
    try:
        if p.exists():
            p.unlink()
            print(f"     Layer session file cleared -> {p.relative_to(CWD.parent)}")
    except Exception as exc:
        print(f"  [WARN] Could not clear layer session file: {exc}")


def load_picks_json(profile_name: str) -> dict:
    """Load raw picks: {shot_id: {trace_idx: ms}}."""
    p = _picks_json_path(profile_name)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as fh:
            raw = json.load(fh)
        return {int(k): {int(ti): float(tv) for ti, tv in v.items()}
                for k, v in raw.items()}
    except Exception as exc:
        print(f"  [WARN] Could not load picks.json: {exc}")
        return {}


def load_session_picks_json(profile_name: str) -> dict:
    """Load session picks: {shot_id: {trace_idx: ms}}."""
    p = _session_picks_json_path(profile_name)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as fh:
            raw = json.load(fh)
        return {int(k): {int(ti): float(tv) for ti, tv in v.items()}
                for k, v in raw.items()}
    except Exception as exc:
        print(f"  [WARN] Could not load picks.session.json: {exc}")
        return {}


def save_picks_json(profile_name: str, all_picks: dict):
    p = _picks_json_path(profile_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump({str(k): {str(ti): tv for ti, tv in v.items()}
                   for k, v in all_picks.items()},
                  fh, indent=2)
    print(f"     Picks saved -> {p.relative_to(CWD.parent)}")


def save_session_picks_json(profile_name: str, all_picks: dict):
    p = _session_picks_json_path(profile_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump({str(k): {str(ti): tv for ti, tv in v.items()}
                   for k, v in all_picks.items()},
                  fh, indent=2)
    print(f"     Session picks saved -> {p.relative_to(CWD.parent)}")


def clear_session_picks_json(profile_name: str):
    p = _session_picks_json_path(profile_name)
    try:
        if p.exists():
            p.unlink()
            print(f"     Session file cleared -> {p.relative_to(CWD.parent)}")
    except Exception as exc:
        print(f"  [WARN] Could not clear session file: {exc}")


# ---------------------------------------------------------------------------
# Export: picks_clean.txt  (compatible with lvl.ipynb)
# ---------------------------------------------------------------------------

def export_picks_txt(profile_name: str, shots_info: list,
                     all_picks: dict, recv_positions: Any) -> Path:
    """Write Ensemble / SOURCE / CHAN / OFFSET / FB_PICK text file."""
    out_dir  = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    txt_path = out_dir / f"{profile_name}_picks_clean.txt"
    header   = (f"{'Ensemble':>10} {'#':>4} {'SOURCE':>8} "
                f"{'CHAN':>6} {'OFFSET':>10} {'FB_PICK':>10}")

    with open(txt_path, "w", encoding="utf-8") as fh:
        fh.write(header + "\n")
        ens = 1
        for shot_idx, (shot_id, shot_pos_m) in enumerate(shots_info):
            corr = apply_bulk_static(all_picks.get(shot_id, {}))
            if not corr:
                continue
            if shot_idx > 0:
                fh.write("\n")
            for trace_idx in sorted(corr):
                if trace_idx >= len(recv_positions):
                    continue
                offset_m = recv_positions[trace_idx] - shot_pos_m
                fh.write(f"{ens:>10} {ens:>4} {shot_id:>8} "
                         f"{trace_idx+1:>6} {offset_m:>10.2f} "
                         f"{corr[trace_idx]:>10.3f}\n")
                ens += 1

    print(f"  picks_txt -> {txt_path.relative_to(CWD.parent)}")
    return txt_path


# ---------------------------------------------------------------------------
# Export: Excel workbook  (Config + per-shot formulas + Analysis)
# ---------------------------------------------------------------------------

_FILL_HDR  = PatternFill("solid", fgColor="2E4057")
_FILL_IN   = PatternFill("solid", fgColor="FFFACD")   # lemon  = editable input
_FILL_FORM = PatternFill("solid", fgColor="E8F4F8")   # blue   = Excel formula
_FILL_OK   = PatternFill("solid", fgColor="DDFFDD")   # green  = depth result
_FONT_HDR  = Font(bold=True, color="FFFFFF")
_FONT_BOLD = Font(bold=True)
_FONT_ITA  = Font(italic=True, color="888888")


def _chdr(ws: Any, row: int, col: int, val: str):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _FILL_HDR
    c.font = _FONT_HDR
    c.alignment = Alignment(horizontal="center")


def _autofit_xl(ws: Any, min_w: int = 6, max_w: int = 30):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None),
                   default=min_w)
        ws.column_dimensions[col[0].column_letter].width = min(best + 3, max_w)


def export_excel(profile_name: str, shots_info: list,
                 all_picks: dict, recv_positions: Any,
                 analysis: dict, cfg: dict,
                 corrected_by_shot: dict | None = None,
                 layer_results: dict | None = None,
                 perp_by_shot: dict | None = None,
                 filename_suffix: str = "") -> Path:
    """
    Write <profile>_picks.xlsx with four sheets.

    Config    Yellow cells = user input.
              B5 = perp_m, B6 = bulk_static_ms.
              All per-shot formula cells reference these two cells so
              changing perp_m or the static updates everything instantly.

    Shot_N    Per-shot picks with Excel formulas:
                True_off  = SQRT(Inline_off^2 + Config!$B$5^2)
                FB_corr   = FB_raw + Config!$B$6
              Summary rows: SLOPE / INTERCEPT / RSQ on corrected picks.

    Combined  All shots concatenated.

    Analysis  Python-computed velocities, intercept times, depths (green),
              R^2, RMS.  Depth formula reference below the table.
    """
    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    xl_path = out_dir / f"{profile_name}_picks{filename_suffix}.xlsx"
    wb      = openpyxl.Workbook()

    # -- Config sheet -------------------------------------------------------
    ws_cfg = wb.active
    ws_cfg.title = "Config"
    ws_cfg["A1"] = f"LVL Refraction Analysis  --  Profile {profile_name}"
    ws_cfg["A1"].font = Font(bold=True, size=13)
    ws_cfg.merge_cells("A1:D1")

    def _inp(row: int, name: str, value: Any, unit: str = "", note: str = ""):
        ws_cfg.cell(row=row, column=1, value=name).font = _FONT_BOLD
        c = ws_cfg.cell(row=row, column=2, value=value)
        c.fill = _FILL_IN
        c.alignment = Alignment(horizontal="right")
        ws_cfg.cell(row=row, column=3, value=unit)
        if note:
            ws_cfg.cell(row=row, column=4, value=note).font = _FONT_ITA

    _inp(3,  "Profile name",      profile_name,           "",   "Folder name under data/")
    _inp(4,  "Geometry type",     cfg.get("geom", "?"),   "m",  "100 = 100m spread,  200 = 200m spread")
    _inp(5,  "Perp. distance",    cfg.get("perp_m", 0.0), "m",
             "Perpendicular shot-to-line offset.  "
             "True_off = SQRT(inline^2 + perp^2)  <- referenced in Shot sheets col G")
    _inp(6,  "Bulk static",       BULK_SHIFT_MS,          "ms",
             "Added to raw FB times.  ProMAX: bulk shift static -> add.  "
             "Negative = earlier.  <- referenced in Shot sheets col I")
    _inp(7,  "N layers",          N_LAYERS,                "",   "2 or 3")
    _inp(8,  "Ormsby f1",         BP_F1,                  "Hz", "Low-cut ramp start")
    _inp(9,  "Ormsby f2",         BP_F2,                  "Hz", "Low-cut ramp end / pass start")
    _inp(10, "Ormsby f3",         BP_F3,                  "Hz", "Pass end / high-cut ramp start")
    _inp(11, "Ormsby f4",         BP_F4,                  "Hz", "High-cut ramp end")
    _inp(12, "FFT zero-padding",  int(BP_FFT_PAD * 100),  "%",  "25 = ProMAX default")
    _inp(13, "STA window",        STA_MS,                 "ms", "Short-term average (auto-pick)")
    _inp(14, "LTA window",        LTA_MS,                 "ms", "Long-term average (auto-pick)")
    _inp(15, "STA/LTA threshold", STALTA_TRIG,             "",  "Trigger ratio")
    if perp_by_shot:
        shot_po_str = ", ".join(f"{sid}:{perp_by_shot[sid]:.3f}" for sid in sorted(perp_by_shot))
        _inp(16, "Perp by shot", shot_po_str, "m",
             "Used for corrected offsets Ck=sqrt((xk-SP)^2+PO^2)")

    ws_cfg.column_dimensions["A"].width = 24
    ws_cfg.column_dimensions["B"].width = 14
    ws_cfg.column_dimensions["C"].width = 6
    ws_cfg.column_dimensions["D"].width = 60

    # -- Per-shot sheets ----------------------------------------------------
    combined_rows: list = []

    for shot_id, shot_pos_m in shots_info:
        corr_rows = (corrected_by_shot or {}).get(shot_id, [])
        if not corr_rows:
            raw_picks = all_picks.get(shot_id, {})
            if not raw_picks:
                continue
            local_po = float((perp_by_shot or {}).get(shot_id, cfg.get("perp_m", 0.0)))
            corr_rows = []
            for trace_idx in sorted(raw_picks):
                if trace_idx >= len(recv_positions):
                    continue
                recv_pos = float(recv_positions[trace_idx])
                inline   = recv_pos - shot_pos_m
                raw_fb = float(raw_picks[trace_idx])
                bulk_fb = float(apply_bulk_static({trace_idx: raw_fb})[trace_idx])
                corr_rows.append({
                    "trace_idx": trace_idx,
                    "trace_no": trace_idx + 1,
                    "recv_pos_m": recv_pos,
                    "shot_pos_m": shot_pos_m,
                    "inline_signed_m": inline,
                    "inline_abs_m": abs(inline),
                    "perp_m": local_po,
                    "true_off_m": float(true_offset(abs(inline), local_po)),
                    "fb_raw_ms": raw_fb,
                    "fb_bulk_ms": bulk_fb,
                    "fb_interp_inline_ms": bulk_fb,
                    "fb_interp_geom_ms": bulk_fb,
                    "side": "L" if inline < 0 else "R",
                })

        if not corr_rows:
            continue

        ws = wb.create_sheet(title=f"Shot_{shot_id}")
        col_headers = ["Trace", "FFID", "Recv_pos_m", "Shot_pos_m",
               "Inline_signed_m", "Inline_abs_m", "Inline_shift_m", "Inline_corr_m",
               "Perp_m", "True_off_m",
                   "FB_raw_ms", "FB_bulk_ms", "FB_interp_geom_ms", "Side"]
        for ci, h in enumerate(col_headers, 1):
            _chdr(ws, 1, ci, h)

        for row_i, row_d in enumerate(corr_rows, start=2):
            ws.cell(row=row_i, column=1, value=int(row_d["trace_no"]))
            ws.cell(row=row_i, column=2, value=shot_id)
            ws.cell(row=row_i, column=3, value=round(float(row_d["recv_pos_m"]), 3))
            ws.cell(row=row_i, column=4, value=round(float(row_d["shot_pos_m"]), 3))
            ws.cell(row=row_i, column=5, value=round(float(row_d["inline_signed_m"]), 3))
            ws.cell(row=row_i, column=6, value=round(float(row_d["inline_abs_m"]), 3))
            ws.cell(row=row_i, column=7, value=round(float(row_d.get("inline_shift_m", 0.0)), 3))
            ws.cell(row=row_i, column=8, value=round(float(row_d.get("inline_corr_m", row_d["inline_abs_m"])), 3))
            ws.cell(row=row_i, column=9, value=round(float(row_d["perp_m"]), 3))
            ws.cell(row=row_i, column=10, value=round(float(row_d["true_off_m"]), 3))
            ws.cell(row=row_i, column=11, value=round(float(row_d["fb_raw_ms"]), 3))
            ws.cell(row=row_i, column=12, value=round(float(row_d["fb_bulk_ms"]), 3))
            ws.cell(row=row_i, column=13, value=round(float(row_d.get("fb_interp_geom_ms", row_d["fb_interp_inline_ms"])), 3))
            ws.cell(row=row_i, column=14, value=str(row_d.get("side", "R")))

            combined_rows.append({
                "shot_id": shot_id,
                "trace": int(row_d["trace_no"]),
                "recv_pos": float(row_d["recv_pos_m"]),
                "shot_pos": float(row_d["shot_pos_m"]),
                "inline_abs": float(row_d["inline_abs_m"]),
                "inline_shift": float(row_d.get("inline_shift_m", 0.0)),
                "inline_corr": float(row_d.get("inline_corr_m", row_d["inline_abs_m"])),
                "true_off": float(row_d["true_off_m"]),
                "fb_raw": float(row_d["fb_raw_ms"]),
                "fb_bulk": float(row_d["fb_bulk_ms"]),
                "fb_interp": float(row_d.get("fb_interp_geom_ms", row_d["fb_interp_inline_ms"])),
                "side": str(row_d.get("side", "R")),
            })

        # Summary: SLOPE / INTERCEPT / RSQ on corrected picks vs true offset
        n_data  = len(corr_rows)
        sum_row = n_data + 3
        _chdr(ws, sum_row, 1, "Metric")
        _chdr(ws, sum_row, 2, "Value")
        _chdr(ws, sum_row, 3, "Range used")

        x_rng = f"J2:J{n_data + 1}"
        t_rng = f"M2:M{n_data + 1}"
        for ri, (label, formula, basis) in enumerate([
            ("Velocity (m/s)",
             f"=IFERROR(1000/SLOPE({t_rng},{x_rng}),\"n/a\")",
             "SLOPE(FB_interp_geom, True_off)"),
            ("Intercept (ms)",
             f"=IFERROR(INTERCEPT({t_rng},{x_rng}),\"n/a\")",
             "INTERCEPT(FB_interp_geom, True_off)"),
            ("R^2",
             f"=IFERROR(RSQ({t_rng},{x_rng}),\"n/a\")",
             "RSQ(FB_interp_geom, True_off)"),
        ], start=sum_row + 1):
            ws.cell(row=ri, column=1, value=label).font = _FONT_BOLD
            ws.cell(row=ri, column=2, value=formula).fill = _FILL_FORM
            ws.cell(row=ri, column=3, value=basis).font  = _FONT_ITA

        _autofit_xl(ws)

    # -- Combined sheet -----------------------------------------------------
    if combined_rows:
        ws_c = wb.create_sheet(title="Combined")
        for ci, h in enumerate(["Shot_ID", "Trace", "Recv_pos_m", "Shot_pos_m",
                     "Inline_abs_m", "Inline_shift_m", "Inline_corr_m", "True_off_m", "FB_raw_ms",
                     "FB_bulk_ms", "FB_interp_geom_ms", "Side"], 1):
            _chdr(ws_c, 1, ci, h)
        for ri, row in enumerate(combined_rows, start=2):
            ws_c.cell(row=ri, column=1, value=row["shot_id"])
            ws_c.cell(row=ri, column=2, value=row["trace"])
            ws_c.cell(row=ri, column=3, value=round(row["recv_pos"], 3))
            ws_c.cell(row=ri, column=4, value=round(row["shot_pos"], 3))
            ws_c.cell(row=ri, column=5, value=round(row["inline_abs"], 3))
            ws_c.cell(row=ri, column=6, value=round(row.get("inline_shift", 0.0), 3))
            ws_c.cell(row=ri, column=7, value=round(row.get("inline_corr", row["inline_abs"]), 3))
            ws_c.cell(row=ri, column=8, value=round(row["true_off"], 3))
            ws_c.cell(row=ri, column=9, value=round(row["fb_raw"], 3))
            ws_c.cell(row=ri, column=10, value=round(row["fb_bulk"], 3))
            ws_c.cell(row=ri, column=11, value=round(row["fb_interp"], 3))
            ws_c.cell(row=ri, column=12, value=row["side"])
        _autofit_xl(ws_c)

    # -- Layer picks sheet -------------------------------------------------
    if layer_results:
        ws_l = wb.create_sheet(title="Layer_Picks")
        l_hdrs = ["Shot_ID", "Side", "x1", "x2", "x3", "x4", "x5", "x6",
              "t1 (ms)", "t2 (ms)", "t3 (ms)", "t4 (ms)", "t5 (ms)", "t6 (ms)",
              "V0 (m/s)", "V1 (m/s)", "V2 (m/s)",
              "ti1 (ms)", "ti2 (ms)", "h1 (m)", "h2 (m)", "n_points"]
        for ci, h in enumerate(l_hdrs, 1):
            _chdr(ws_l, 1, ci, h)

        rr = 2
        for shot_id in sorted(layer_results):
            shot_sides = layer_results.get(shot_id, {})
            ordered_sides = [s for s in ("L", "R", "ALL") if s in shot_sides]
            ordered_sides += [s for s in shot_sides if s not in ordered_sides]
            for side in ordered_sides:
                res = shot_sides.get(side)
                if not res:
                    continue
                wins = res.get("windows", [None, None, None])
                pts = list(res.get("picked_points", []))
                pts += [(None, None)] * (6 - len(pts))

                def _w(i: int, j: int) -> Any:
                    w = wins[i] if i < len(wins) else None
                    return round(float(w[j]), 3) if w else ""

                def _px(i: int) -> Any:
                    xv = pts[i][0] if i < len(pts) else None
                    return round(float(xv), 3) if xv is not None else ""

                def _pt(i: int) -> Any:
                    tv = pts[i][1] if i < len(pts) else None
                    return round(float(tv), 3) if tv is not None else ""

                ws_l.cell(row=rr, column=1, value=shot_id)
                ws_l.cell(row=rr, column=2, value=side)
                ws_l.cell(row=rr, column=3, value=_px(0) if _px(0) != "" else _w(0, 0))
                ws_l.cell(row=rr, column=4, value=_px(1) if _px(1) != "" else _w(0, 1))
                ws_l.cell(row=rr, column=5, value=_px(2) if _px(2) != "" else _w(1, 0))
                ws_l.cell(row=rr, column=6, value=_px(3) if _px(3) != "" else _w(1, 1))
                ws_l.cell(row=rr, column=7, value=_px(4) if _px(4) != "" else _w(2, 0))
                ws_l.cell(row=rr, column=8, value=_px(5) if _px(5) != "" else _w(2, 1))
                ws_l.cell(row=rr, column=9, value=_pt(0))
                ws_l.cell(row=rr, column=10, value=_pt(1))
                ws_l.cell(row=rr, column=11, value=_pt(2))
                ws_l.cell(row=rr, column=12, value=_pt(3))
                ws_l.cell(row=rr, column=13, value=_pt(4))
                ws_l.cell(row=rr, column=14, value=_pt(5))
                ws_l.cell(row=rr, column=15, value=round(float(res.get("V0_m_s", 0.0)), 2))
                ws_l.cell(row=rr, column=16, value=round(float(res.get("V1_m_s", 0.0)), 2))
                ws_l.cell(row=rr, column=17, value=round(float(res.get("V2_m_s", 0.0)), 2))
                ws_l.cell(row=rr, column=18, value=round(float(res.get("ti1_ms", 0.0)), 3))
                ws_l.cell(row=rr, column=19, value=round(float(res.get("ti2_ms", 0.0)), 3))
                ws_l.cell(row=rr, column=20, value=round(float(res.get("h1_m", 0.0)), 3))
                ws_l.cell(row=rr, column=21, value=round(float(res.get("h2_m", 0.0)), 3))
                ws_l.cell(row=rr, column=22, value=int(res.get("n_points", 0)))

                for col in (20, 21):
                    cell = ws_l.cell(row=rr, column=col)
                    if cell.value not in ("", 0, 0.0):
                        cell.fill = _FILL_OK
                rr += 1

        _autofit_xl(ws_l)

    # -- Layer averages sheet ---------------------------------------------
    shot_pos_by_id = {int(sid): float(sp) for sid, sp in shots_info}
    layer_avg = compute_layer_averages(layer_results or {}, shot_pos_by_id)
    if layer_avg:
        ws_la = wb.create_sheet(title="Layer_Averages")
        hdrs = ["Layer", "V_off (m/s)", "V_center (m/s)", "V_avg (m/s)",
                "ti_off (ms)", "ti_center (ms)", "ti_avg (ms)"]
        for ci, h in enumerate(hdrs, 1):
            _chdr(ws_la, 1, ci, h)

        layer_map = [("V0", "ti1_ms", "Layer 1"), ("V1", "ti1_ms", "Layer 2"), ("V2", "ti2_ms", "Layer 3")]
        rr = 2
        for vkey, tkey, lname in layer_map:
            vv = layer_avg.get(vkey, {})
            tt = layer_avg.get(tkey, {})
            ws_la.cell(row=rr, column=1, value=lname)
            ws_la.cell(row=rr, column=2, value=round(float(vv.get("off", 0.0)), 3))
            ws_la.cell(row=rr, column=3, value=round(float(vv.get("center", 0.0)), 3))
            ws_la.cell(row=rr, column=4, value=round(float(vv.get("avg", 0.0)), 3))
            ws_la.cell(row=rr, column=5, value=round(float(tt.get("off", 0.0)), 3))
            ws_la.cell(row=rr, column=6, value=round(float(tt.get("center", 0.0)), 3))
            ws_la.cell(row=rr, column=7, value=round(float(tt.get("avg", 0.0)), 3))
            rr += 1

        ws_la.cell(row=rr + 1, column=1, value="Depth h1 (m)").font = _FONT_BOLD
        ws_la.cell(row=rr + 1, column=2, value=round(float(layer_avg.get("h1_m", 0.0)), 3)).fill = _FILL_OK
        ws_la.cell(row=rr + 2, column=1, value="Depth h2 (m)").font = _FONT_BOLD
        ws_la.cell(row=rr + 2, column=2, value=round(float(layer_avg.get("h2_m", 0.0)), 3)).fill = _FILL_OK
        ws_la.cell(row=rr + 4, column=1, value="Off-end shots").font = _FONT_BOLD
        ws_la.cell(row=rr + 4, column=2, value=", ".join(str(s) for s in layer_avg.get("off_shots", [])))
        ws_la.cell(row=rr + 5, column=1, value="Center shots").font = _FONT_BOLD
        ws_la.cell(row=rr + 5, column=2, value=", ".join(str(s) for s in layer_avg.get("center_shots", [])))
        _autofit_xl(ws_la)

    # -- Analysis sheet -----------------------------------------------------
    ws_a = wb.create_sheet(title="Analysis")
    a_hdrs = ["Shot_ID", "FFID", "V1 (m/s)", "V2 (m/s)", "V3 (m/s)",
              "ti_1 (ms)", "ti_2 (ms)", "h1 (m)", "h2 (m)", "RMS (ms)", "n_picks"]
    for ci, h in enumerate(a_hdrs, 1):
        _chdr(ws_a, 1, ci, h)

    row_a = 2
    for shot_id, _ in shots_info:
        res = (analysis or {}).get(shot_id)
        if res is None:
            continue
        deps = res.get("depths", {})
        segs = res.get("segments", [])

        def _vel(i: int) -> Any:
            return round(segs[i]["velocity_m_s"], 1) if i < len(segs) else ""

        def _dep(key: str) -> Any:
            val = deps.get(key)
            return round(val, 2) if val else ""

        ws_a.cell(row=row_a, column=1,  value=shot_id)
        ws_a.cell(row=row_a, column=2,  value=shot_id)
        ws_a.cell(row=row_a, column=3,  value=_vel(0))
        ws_a.cell(row=row_a, column=4,  value=_vel(1))
        ws_a.cell(row=row_a, column=5,  value=_vel(2))
        ws_a.cell(row=row_a, column=6,
                  value=round(deps["ti1_ms"], 3) if deps.get("ti1_ms") else "")
        ws_a.cell(row=row_a, column=7,
                  value=round(deps["ti2_ms"], 3) if deps.get("ti2_ms") else "")
        ws_a.cell(row=row_a, column=8,  value=_dep("h1_m"))
        ws_a.cell(row=row_a, column=9,  value=_dep("h2_m"))
        ws_a.cell(row=row_a, column=10, value=round(res.get("rms_ms", 0.0), 3))
        ws_a.cell(row=row_a, column=11, value=len(all_picks.get(shot_id, {})))

        for col in (8, 9):
            cell = ws_a.cell(row=row_a, column=col)
            if cell.value != "":
                cell.fill = _FILL_OK
        row_a += 1

    note_row = row_a + 2
    notes = [
        "Depth formulas (intercept-time method):",
        "  2-layer:  h1 = (ti1_ms/1000 x V1 x V2) / (2 x SQRT(V2^2 - V1^2))",
        "  3-layer:  h2 = (ti2_eff_ms/1000 x V1 x V3) / (2 x SQRT(V3^2 - V1^2))"
        "    where  ti2_eff = ti2 - 2 x h1 x cos(ic12) / V1 x 1000",
    ]
    for oi, note in enumerate(notes):
        r = note_row + oi
        ws_a.cell(row=r, column=1, value=note).font = (
            _FONT_BOLD if oi == 0 else _FONT_ITA)
        ws_a.merge_cells(f"A{r}:K{r}")

    _autofit_xl(ws_a)
    wb.save(str(xl_path))
    print(f"  Excel     -> {xl_path.relative_to(CWD.parent)}")
    return xl_path


# ---------------------------------------------------------------------------
# Export: T-X picks plot  (time increases downward + analysis overlay)
# ---------------------------------------------------------------------------

def export_tx_plot(profile_name: str, shots_info: list,
                   all_picks: dict, recv_positions: Any,
                   analysis: dict, perp_m: float = 0.0,
                   shot_label_pos: dict | None = None) -> Path:
    """
    Final T-X summary plot.
    - Scatter of corrected picks (absolute true offset vs corrected pick time)
    - Dashed fitted-line overlay with velocity labels
    - Y-axis inverted: t = 0 at top  (seismic convention: time grows downward)
    - Theme follows the current THEME setting at export time
    """
    c = _tc()
    fig, ax = plt.subplots(figsize=(10, 6))
    fig.patch.set_facecolor(c["fig_bg"])
    ax.set_facecolor(c["ax_bg"])

    pal = ["#e63946", "#2a9d8f", "#e9c46a", "#a8dadc", "#f4a261", "#457b9d"]
    mks = ["o", "s", "^", "D", "P", "X"]

    for shot_id, shot_pos_m in shots_info:
        corr = apply_bulk_static(all_picks.get(shot_id, {}))
        if not corr:
            continue
        col = pal[(shot_id - 1) % len(pal)]
        mk  = mks[(shot_id - 1) % len(mks)]
        label_pos_m = (shot_label_pos or {}).get(shot_id, shot_pos_m)

        x_geom, t_vals = [], []
        for idx in sorted(corr):
            if idx < len(recv_positions):
                x_geom.append(float(recv_positions[idx]))
                t_vals.append(corr[idx])

        ax.scatter(x_geom, t_vals, color=col, marker=mk, s=28, zorder=5,
               label=f"Shot {shot_id}  (@ {label_pos_m:.1f} m)")

        # Dashed fitted T-X lines (only when analysis data is available)
        res = (analysis or {}).get(shot_id)
        if res and res.get("segments"):
            for seg in res["segments"]:
                # Convert segment x-range (absolute offset) back to geometry x
                # using shot position so lines align with geometry x-axis
                xg0 = shot_pos_m + float(seg.get("x_start", seg.get("x0", 0.0)))
                xg1 = shot_pos_m + float(seg.get("x_end", seg.get("x1", 0.0)))
                sl  = seg["slope_ms_m"]
                ic  = seg["intercept_ms"]
                vel = seg["velocity_m_s"]
                xl  = np.array([xg0, xg1])
                # Predicted time at geometry x using offset from shot
                t_line = sl * np.abs(xl - shot_pos_m) + ic
                ax.plot(xl, t_line, "--", color=col,
                        lw=1.6, alpha=c["fit_alpha"])
                xm = (xg0 + xg1) / 2.0
                tm = sl * abs(xm - shot_pos_m) + ic
                ax.text(xm, tm - 3.5,
                        f"{vel:.0f} m/s", color=col, fontsize=7,
                        ha="center", fontweight="bold")

    # Seismic convention: t = 0 at top, time grows downward
    ax.set_ylim(T_MAX_MS, 0.0)
    ax.set_xlabel("Receiver position  (m)", color=c["label"])
    ax.set_ylabel("First-break time  (ms)", color=c["label"])
    ax.set_title(f"T-X first-break picks  --  Profile {profile_name}",
                 color=c["text"], fontsize=11)
    ax.grid(True, lw=0.3, alpha=0.3, color=c["grid"])
    ax.tick_params(colors=c["tick"])
    for sp in ax.spines.values():
        sp.set_edgecolor(c["spine"])
    ax.legend(fontsize=9, facecolor=c["leg_face"],
              edgecolor=c["leg_edge"], labelcolor=c["text"])
    fig.tight_layout()

    out_dir = OUTPUT_DIR / profile_name
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{profile_name}_tx_picks.png"
    fig.savefig(str(out), dpi=180, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  T-X plot  -> {out.relative_to(CWD.parent)}")
    return out


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------

def process_profile(profile_name: str, pick_mode: bool = True,
                    geom_override: int | None = None,
                    perp_by_shot_override: dict | None = None,
                    inline_shift_by_shot_override: dict | None = None,
                    perp_excel_cfg: dict | None = None,
                    enable_layer_pick: bool = True):
    cfg = PROFILES.get(profile_name)
    if cfg is None:
        print(f"[ERROR] Profile '{profile_name}' not in PROFILES. "
              f"Known: {list(PROFILES)}")
        return

    data_dir = DATA_DIR / profile_name
    if not data_dir.exists():
        print(f"[ERROR] Data folder not found: {data_dir}")
        return

    inferred_geom: int | None = None
    if geom_override is None and perp_excel_cfg and perp_excel_cfg.get("path"):
        try:
            inferred_geom = infer_geometry_from_field_report(
                path=Path(perp_excel_cfg.get("path")),
                profile_name=profile_name,
                sheet_name=perp_excel_cfg.get("sheet"),
            )
        except Exception as exc:
            print(f"  [WARN] Geometry inference from field report failed: {exc}")

    if geom_override is not None:
        geom_type = int(geom_override)
        geom_src = "CLI"
    elif inferred_geom in (100, 200):
        geom_type = int(inferred_geom)
        geom_src = "field-report"
    else:
        geom_type = 200
        geom_src = "default"

    perp_cfg       = cfg.get("perp_m", 0.0)
    perp_m         = float(perp_cfg if not isinstance(perp_cfg, dict) else 0.0)
    recv_positions = load_geometry(geom_type)
    print(f"  Geometry selected: {geom_type} m ({geom_src})")

    # Resolve shot positions: "auto" derives them from geometry
    shots_cfg = cfg.get("shots", "auto")
    if shots_cfg == "auto" or not isinstance(shots_cfg, dict):
        shots_cfg = auto_shot_positions(recv_positions)
        print(f"  Shot positions (auto from geometry): "
              + "  ".join(f"Shot{k}={v:.3f}m" for k, v in shots_cfg.items()))

        print(f"  Geometry {geom_type} m : {len(recv_positions)} receivers, "
            f"{recv_positions[0]:.2f} - {recv_positions[-1]:.2f} m  "
            f"|  perp(default) = {perp_m:.1f} m  "
            f"|  bulk static = {BULK_SHIFT_MS:+.1f} ms")

    def _ffid(p: Path) -> int:
        digits = "".join(c for c in p.stem if c.isdigit())
        return int(digits) if digits else 0

    seg2_candidates = list(data_dir.glob("*.seg2")) + list(data_dir.glob("*.SEG2"))
    seg2_unique = {str(p.resolve()).lower(): p for p in seg2_candidates}
    seg2_files = sorted(seg2_unique.values(), key=_ffid)
    if not seg2_files:
        print(f"[ERROR] No .seg2 files found in {data_dir}")
        return
    print(f"  Found {len(seg2_files)} SEG2 file(s)")

    final_picks: dict = load_picks_json(profile_name)
    session_picks: dict = load_session_picks_json(profile_name)
    if session_picks:
        all_picks: dict = session_picks
        print("  Session picks found: resuming from picks.session.json")
    else:
        all_picks = {sid: dict(vals) for sid, vals in final_picks.items()}
    shots_meta: list = []
    qc_dir = OUTPUT_DIR / profile_name
    finalized = False

    shot_cache: list = []
    for file_idx, seg2_path in enumerate(seg2_files):
        shot_id = file_idx + 1
        data_raw, dt_s, n_tr, n_samp, shot_pos_hdr, ffid_hdr, \
            delay_ms, recv_locs_hdr = read_seg2(seg2_path)
        n_show = min(n_tr, len(recv_positions))
        shot_pos_cfg = shots_cfg.get(shot_id)
        shot_pos_source = "default"
        shot_pos_m = 0.0

        if shot_pos_cfg is not None:
            shot_pos_m = float(shot_pos_cfg)
            shot_pos_source = "config"
        elif shot_pos_hdr is not None:
            shot_pos_m = float(shot_pos_hdr)
            shot_pos_source = "header"

        if USE_SEG2_SHOT_POSITION and shot_pos_hdr is not None:
            shot_pos_m = float(shot_pos_hdr)
            shot_pos_source = "header"
        shot_pos_nominal = (float(shot_pos_cfg)
                            if shot_pos_cfg is not None else float(shot_pos_m))
        shot_cache.append({
            "shot_id": shot_id,
            "seg2_path": seg2_path,
            "data_raw": data_raw,
            "dt_s": dt_s,
            "n_tr": n_tr,
            "n_samp": n_samp,
            "shot_pos_hdr": shot_pos_hdr,
            "ffid_hdr": ffid_hdr,
            "delay_ms": delay_ms,
            "n_show": n_show,
            "shot_pos_m": shot_pos_m,
            "shot_pos_source": shot_pos_source,
            "shot_pos_nominal": shot_pos_nominal,
        })

    idx = 0
    while 0 <= idx < len(shot_cache):
        shot = shot_cache[idx]
        shot_id = int(shot["shot_id"])
        seg2_path = shot["seg2_path"]
        print(f"\n  -- Shot {shot_id}  ({seg2_path.name}) --")

        data_raw = shot["data_raw"]
        dt_s = float(shot["dt_s"])
        n_tr = int(shot["n_tr"])
        n_samp = int(shot["n_samp"])
        shot_pos_hdr = shot["shot_pos_hdr"]
        shot_pos_source = shot.get("shot_pos_source", "default")
        ffid_hdr = shot["ffid_hdr"]
        delay_ms = float(shot["delay_ms"])
        n_show = int(shot["n_show"])
        shot_pos_m = float(shot["shot_pos_m"])

        dt_ms  = dt_s * 1000.0
        t_end  = delay_ms + (n_samp - 1) * dt_ms
        print(f"     FFID={ffid_hdr}  |  {n_tr} traces  |  "
              f"dt={dt_ms:.4f} ms  |  delay={delay_ms:.1f} ms  |  "
              f"{n_samp} smp  ({delay_ms:.1f} to {t_end:.1f} ms)")

        if shot_pos_source == "header":
            print(f"     Shot pos : {shot_pos_m:.2f} m  (SEG2 header)")
        elif shot_pos_source == "config":
            print(f"     Shot pos : {shot_pos_m:.2f} m  (CONFIG)")
        elif shot_pos_hdr is not None:
            print(f"     Shot pos : {shot_pos_m:.2f} m  (SEG2 header; out-of-range fallback)")
        else:
            print("     Shot pos : 0.0 m  [WARN: defaulting to 0]")

        if pick_mode:
            # Always use geometry file positions for the x-axis display and picking.
            # SEG2 RECEIVER_LOCATION may differ from the geometry file (e.g. a
            # 200-m spread recorded with 0-94 m header values).
            geom_slice = recv_positions[:n_show]

            data_slice = data_raw[:n_show]
            data_filt  = apply_ormsby_all(data_slice, dt_s)

            def _save_cb(picks_for_shot: dict, _sid: int = shot_id) -> None:
                """Called on nav/finalize to persist temporary session progress."""
                all_picks[_sid] = picks_for_shot
                save_session_picks_json(profile_name, all_picks)
                print(f"     ✓ {len(picks_for_shot)} pick(s) saved to session.")

            picker = FirstBreakPicker(
                data_slice, data_filt, dt_s,
                geom_slice, shot_id, profile_name,
                shot_pos_m=shot_pos_m,
                delay_ms=delay_ms,
                existing_picks=all_picks.get(shot_id, {}),
                qc_dir=qc_dir,
                save_callback=_save_cb,
                header_info={
                    "ffid": ffid_hdr,
                    "n_tr": n_tr,
                    "n_samp": n_samp,
                    "shot_pos_hdr": shot_pos_hdr,
                },
            )
            result = picker.run() or {"status": "quit", "picks": all_picks.get(shot_id, {})}
            status = result.get("status", "next")
            all_picks[shot_id] = dict(result.get("picks", {}))
            save_session_picks_json(profile_name, all_picks)

            if all_picks.get(shot_id):
                print(f"     Shot {shot_id}: {len(all_picks[shot_id])} pick(s) in memory.")
            else:
                print(f"     Picking cancelled / no picks for shot {shot_id}.")

            if status == "prev":
                if idx > 0:
                    idx -= 1
                else:
                    print("     Already at first shot; staying on current shot.")
                continue
            if status == "next":
                if idx < (len(shot_cache) - 1):
                    idx += 1
                else:
                    print("     Already at last shot; staying on current shot.")
                continue
            if status == "finalize":
                finalized = True
                break
            if status == "quit":
                break

        idx += 1

    preview_only = bool(pick_mode and not finalized)
    if preview_only:
        print("\n  Picking session ended without Save/Close finalization.")
        print("  Continuing to correction UI in PREVIEW mode.")
        print("  Final picks.json remains unchanged until Save/Close is used.")

    for shot in shot_cache:
        shot_id = int(shot["shot_id"])
        shots_meta.append((shot_id, float(shot["shot_pos_m"]), float(shot["shot_pos_nominal"])))

    excel_perp_by_shot: dict = {}
    excel_shift_by_shot: dict = {}
    if perp_excel_cfg and perp_excel_cfg.get("path"):
        try:
            ffid_by_shot = {
                int(s["shot_id"]): int(s.get("ffid_hdr", 0) or 0)
                for s in shot_cache
            }
            excel_perp_by_shot, excel_shift_by_shot = load_profile_offsets_from_excel(
                path=Path(perp_excel_cfg.get("path")),
                profile_name=profile_name,
                ffid_by_shot=ffid_by_shot,
                sheet_name=perp_excel_cfg.get("sheet"),
                ffid_col=str(perp_excel_cfg.get("ffid_col", "A")),
                perp_col=str(perp_excel_cfg.get("perp_col", "D")),
                profile_col=str(perp_excel_cfg.get("profile_col", "F")),
                inline_shift_col=perp_excel_cfg.get("inline_shift_col"),
            )
            if excel_perp_by_shot:
                print("  Excel defaults loaded for PO: "
                      + ", ".join(f"S{k}={v:.2f}" for k, v in sorted(excel_perp_by_shot.items())))
            if excel_shift_by_shot:
                print("  Excel defaults loaded for X-shift: "
                      + ", ".join(f"S{k}={v:.2f}" for k, v in sorted(excel_shift_by_shot.items())))
        except Exception as exc:
            print(f"  [WARN] Could not load PO Excel defaults: {exc}")

    perp_effective = dict(excel_perp_by_shot)
    perp_effective.update(perp_by_shot_override or {})
    shift_effective = dict(excel_shift_by_shot)
    shift_effective.update(inline_shift_by_shot_override or {})

    total_picks = sum(len(v) for v in all_picks.values())
    if total_picks == 0:
        print("\n  No picks to export.")
        return

    analysis: dict = {}

    layer_final: dict = load_layer_json(profile_name)
    layer_session: dict = load_layer_session_json(profile_name)
    existing_layer_results: dict = layer_session if layer_session else layer_final
    if layer_session:
        print("  Layer session found: resuming from layer_analysis.session.json")

    # Exports / correction UI
    print(f"\n  -- Exporting / correction  ({total_picks} picks) --")
    shots_info_proc = [(sid, sp_proc) for sid, sp_proc, _ in shots_meta]
    shot_label_pos  = {sid: sp_nom for sid, _, sp_nom in shots_meta}
    analysis_ui = AnalysisWorkflow(
        profile_name=profile_name,
        cfg=cfg,
        shots_info=shots_info_proc,
        shot_label_pos=shot_label_pos,
        all_picks=all_picks,
        recv_positions=recv_positions,
        perp_override=perp_effective,
        inline_shift_override=shift_effective,
        enable_layer_pick=enable_layer_pick,
        existing_layer_results=existing_layer_results,
    )
    analysis_bundle = analysis_ui.run()
    perp_by_shot = analysis_bundle["perp_by_shot"]
    corrected_by_shot = analysis_bundle["corrected_by_shot"]
    layer_results = analysis_bundle["layer_results"]
    analysis = build_analysis_from_layers(corrected_by_shot, layer_results)
    save_layer_session_json(profile_name, layer_results)

    qc_suffix = "_preview" if preview_only else ""
    export_corrected_qc_plot(profile_name, corrected_by_shot,
                             shot_label_pos=shot_label_pos,
                             layer_results=layer_results,
                             filename_suffix=qc_suffix,
                             show_plot=True)

    if preview_only:
        export_excel(profile_name, shots_info_proc, all_picks, recv_positions,
                     analysis, cfg,
                     corrected_by_shot=corrected_by_shot,
                     layer_results=layer_results,
                     perp_by_shot=perp_by_shot,
                     filename_suffix="_preview")
        export_tx_plot(profile_name, shots_info_proc, all_picks, recv_positions,
                       analysis, perp_m=perp_m, shot_label_pos=shot_label_pos)
        export_fit_plot(profile_name, corrected_by_shot, layer_results,
                        filename_suffix="_preview")
        print("  Preview files written; final picks.json not updated.")
        return

    export_picks_txt(profile_name, shots_info_proc, all_picks, recv_positions)
    export_excel(profile_name, shots_info_proc, all_picks, recv_positions,
                 analysis, cfg,
                 corrected_by_shot=corrected_by_shot,
                 layer_results=layer_results,
                 perp_by_shot=perp_by_shot)
    export_tx_plot(profile_name, shots_info_proc, all_picks, recv_positions,
                   analysis, perp_m=perp_m, shot_label_pos=shot_label_pos)
    export_fit_plot(profile_name, corrected_by_shot, layer_results)
    save_picks_json(profile_name, all_picks)
    save_layer_json(profile_name, layer_results)
    clear_session_picks_json(profile_name)
    clear_layer_session_json(profile_name)
    print(f"\n  Output -> {(OUTPUT_DIR / profile_name).relative_to(CWD.parent)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        prog="lvl_refraction.py",
        description="LVL refraction seismic: interactive picking + analysis.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python lvl_refraction.py 120\n"
            "  python lvl_refraction.py 120 --export-only\n"
            "  python lvl_refraction.py --all\n"
        ),
    )
    parser.add_argument("profile", nargs="?", default=None,
                        help="Profile folder name (e.g. 120 or 214_A)")
    parser.add_argument("geometry", nargs="?", default=None,
                        help="Legacy geometry override (positional): 100, 200, geometry100, geometry200")
    parser.add_argument("--geom", default=None,
                        help="Geometry override: 100, 200, geometry100, geometry200")
    parser.add_argument("--all", action="store_true",
                        help="Process all profiles in PROFILES")
    parser.add_argument("--export-only", action="store_true",
                        help="Skip picking; re-analyse and re-export existing picks")
    parser.add_argument("--perp-by-shot", default=None,
                        help="Override perpendicular offset per shot, e.g. 1:0,2:3.5,3:0")
    parser.add_argument("--inline-shift-by-shot", default=None,
                        help="Optional inline shift per shot (m), e.g. 1:50,2:0,3:50")
    parser.add_argument("--perp-excel", default=None,
                        help="Excel file with FFID/profile/perpendicular offsets defaults")
    parser.add_argument("--perp-sheet", default=None,
                        help="Excel sheet name/index for --perp-excel (default: first sheet)")
    parser.add_argument("--perp-col-ffid", default="A",
                        help="Excel column for FFID (default: A)")
    parser.add_argument("--perp-col-perp", default="D",
                        help="Excel column for perpendicular offset (default: D)")
    parser.add_argument("--perp-col-profile", default="F",
                        help="Excel column for profile token, e.g. LVL150 (default: F)")
    parser.add_argument("--perp-col-inline-shift", default=None,
                        help="Optional Excel column for inline X-shift in meters")
    parser.add_argument("--no-layer-pick", action="store_true",
                        help="Skip interactive x0..x5 layer-window picking")
    args = parser.parse_args()

    try:
        perp_override = parse_shot_value_map(args.perp_by_shot)
    except Exception as exc:
        print(f"[ERROR] Invalid --perp-by-shot value: {exc}")
        return

    try:
        inline_shift_override = parse_shot_value_map(args.inline_shift_by_shot)
    except Exception as exc:
        print(f"[ERROR] Invalid --inline-shift-by-shot value: {exc}")
        return

    geom_override: int | None = None
    geom_raw = args.geom if args.geom is not None else args.geometry
    if geom_raw is not None:
        g = str(geom_raw).strip().lower().replace(" ", "")
        if g in ("100", "geometry100"):
            geom_override = 100
        elif g in ("200", "geometry200"):
            geom_override = 200
        else:
            print("[ERROR] Invalid geometry override. Use one of:")
            print("        100, 200, geometry100, geometry200")
            return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    perp_excel_cfg = None
    if args.perp_excel:
        sheet_val: str | int | None = args.perp_sheet
        if sheet_val is not None:
            s_try = str(sheet_val).strip()
            if s_try.isdigit():
                sheet_val = int(s_try)
        perp_excel_cfg = {
            "path": args.perp_excel,
            "sheet": sheet_val,
            "ffid_col": args.perp_col_ffid,
            "perp_col": args.perp_col_perp,
            "profile_col": args.perp_col_profile,
            "inline_shift_col": args.perp_col_inline_shift,
        }
    else:
        auto_excel = discover_field_report_excel(DATA_DIR)
        if auto_excel is not None:
            perp_excel_cfg = {
                "path": str(auto_excel),
                "sheet": None,
                "ffid_col": args.perp_col_ffid,
                "perp_col": args.perp_col_perp,
                "profile_col": args.perp_col_profile,
                "inline_shift_col": args.perp_col_inline_shift,
            }
            print(f"  Auto field report detected: {auto_excel.relative_to(CWD.parent)}")

    if args.all:
        targets = list(PROFILES)
    elif args.profile:
        targets = [args.profile]
    else:
        chosen_profile = None
        try:
            root = tk.Tk()
            root.withdraw()
            sel = filedialog.askdirectory(
                title="Select profile data folder",
                initialdir=str(DATA_DIR),
                mustexist=True,
            )
            root.destroy()
            if sel:
                p = Path(sel)
                if p.parent.resolve() == DATA_DIR.resolve():
                    chosen_profile = p.name
        except Exception:
            chosen_profile = None

        if chosen_profile:
            print(f"  Selected folder -> profile '{chosen_profile}'")
            targets = [chosen_profile]
        else:
            print("\nAvailable profiles:")
            print(f"  {'Name':<12}  {'Geom':>6}  {'Perp_m':>7}  Data folder")
            print(f"  {'-'*12}  {'-'*6}  {'-'*7}  {'-'*30}")
            for pname, pcfg in PROFILES.items():
                folder = DATA_DIR / pname
                status = "found" if folder.exists() else "MISSING"
                n_seg2 = len(list(folder.glob("*.seg2"))) if folder.exists() else 0
                print(f"  {pname:<12}  {pcfg['geom']:>5}m  "
                      f"{pcfg.get('perp_m', 0.0):>7.1f}m  "
                      f"{status}  ({n_seg2} SEG2 files)")
            print(f"\nUsage:  python lvl_refraction.py <profile>  [--export-only]")
            return

    for pname in targets:
        print(f"\n{'='*60}\nProfile : {pname}\n{'='*60}")
        process_profile(pname, pick_mode=not args.export_only,
                        geom_override=geom_override,
                        perp_by_shot_override=perp_override,
                        inline_shift_by_shot_override=inline_shift_override,
                        perp_excel_cfg=perp_excel_cfg,
                        enable_layer_pick=not args.no_layer_pick)

    print("\nAll done.")


if __name__ == "__main__":
    main()

